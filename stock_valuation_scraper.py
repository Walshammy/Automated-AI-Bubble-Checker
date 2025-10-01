import pandas as pd
import numpy as np
import yfinance as yf
import requests
from bs4 import BeautifulSoup
import time
import os
from datetime import datetime, timedelta
import logging
import json
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import functools
from typing import Dict, Optional, Tuple

warnings.filterwarnings('ignore')

# Configuration constants
VALUATION_CONFIG = {
    'discount_rate': 0.08,
    'perpetual_growth_rate': 0.025,
    'projection_years': 10,
    'default_fcf_growth': 0.08,  # More conservative than 0.15
    'max_fcf_growth': 0.15,     # Cap at 15%
    'min_fcf_growth': 0.02,     # Minimum 2%
    'lynch_thresholds': {
        'VERY_UNDERVALUED': 2.0,
        'UNDERVALUED': 1.5,
        'FAIRLY_VALUED': 1.0,
        'OVERVALUED': 0.0
    },
    'api_delay': 0.2,  # 200ms delay between API calls
    'max_terminal_value_ratio': 0.8  # Terminal value shouldn't exceed 80% of total
}

def rate_limit(delay=VALUATION_CONFIG['api_delay']):
    """Rate limiting decorator for API calls"""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            time.sleep(delay)
            return func(*args, **kwargs)
        return wrapper
    return decorator

class StockValuationScraper:
    def __init__(self):
        # Setup logging
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger(__name__)

        # OneDrive directory (primary storage location)
        self.onedrive_dir = r"C:\Users\james\OneDrive - Silverdale Medical Limited\StockValuation"
        os.makedirs(self.onedrive_dir, exist_ok=True)

        # Second storage location - Downloads/Stock Valuation
        self.downloads_dir = self._get_downloads_directory()
        if self.downloads_dir:
            os.makedirs(self.downloads_dir, exist_ok=True)

        # Daily backups directory
        self.daily_backups_dir = os.path.join(self.onedrive_dir, "daily_backups")
        os.makedirs(self.daily_backups_dir, exist_ok=True)

        # Master dataset files
        self.master_file = os.path.join(self.onedrive_dir, "stock_valuation_dataset.xlsx")
        self.downloads_file = os.path.join(self.downloads_dir, "stock_valuation_dataset.xlsx") if self.downloads_dir else None
        
        # Historical dataset file
        self.historical_file = os.path.join(self.onedrive_dir, "stock_valuation_historical.xlsx")
        
        # Combined dataset file (historical + current)
        self.combined_file = os.path.join(self.onedrive_dir, "stock_valuation_combined.xlsx")
        
        # Historical progress tracking
        self.historical_progress_file = os.path.join(self.onedrive_dir, "valuation_historical_progress.json")

        # Focus stocks for valuation analysis - Complete NZX + International
        self.focus_stocks = {
            # NZX Mega/Large Cap (>$5B) - 13 companies
            'WBC.NZ': 'Westpac Banking Corporation',
            'ANZ.NZ': 'ANZ Group Holdings Limited',
            'FPH.NZ': 'Fisher & Paykel Healthcare Corporation Limited',
            'MEL.NZ': 'Meridian Energy Limited',
            'AIA.NZ': 'Auckland International Airport Limited',
            'IFT.NZ': 'Infratil Limited',
            'AFI.NZ': 'Australian Foundation Investment Company Limited',
            'MCY.NZ': 'Mercury NZ Limited',
            'EBO.NZ': 'EBOS Group Limited',
            'FCG.NZ': 'Fonterra Co-operative Group Limited',
            'CEN.NZ': 'Contact Energy Limited',
            'MFT.NZ': 'Mainfreight Limited',
            'ATM.NZ': 'The a2 Milk Company Limited',
            
            # NZX Mid Cap ($1B - $5B) - 16 companies
            'POT.NZ': 'Port of Tauranga Limited',
            'SPK.NZ': 'Spark New Zealand Limited',
            'VNT.NZ': 'Ventia Services Group Limited',
            'VCT.NZ': 'Vector Limited',
            'CNU.NZ': 'Chorus Limited',
            'FBU.NZ': 'Fletcher Building Limited',
            'GMT.NZ': 'Goodman Property Trust',
            'SUM.NZ': 'Summerset Group Holdings Limited',
            'GNE.NZ': 'Genesis Energy Limited',
            'RYM.NZ': 'Ryman Healthcare Limited',
            'FRW.NZ': 'Freightways Group Limited',
            'MNW.NZ': 'Manawa Energy Limited',
            'PCT.NZ': 'Precinct Properties NZ Ltd',
            'AIR.NZ': 'Air New Zealand Limited',
            'KPG.NZ': 'Kiwi Property Group Limited',
            'GTK.NZ': 'Gentrack Group Limited',
            
            # NZX Small Cap ($300M - $1B) - 29 companies
            'VHP.NZ': 'Vital Healthcare Property Trust',
            'PFI.NZ': 'Property For Industry Limited',
            'BGP.NZ': 'Briscoe Group Limited',
            'ARG.NZ': 'Argosy Property Limited',
            'SKL.NZ': 'Skellerup Holdings Limited',
            'CHI.NZ': 'Channel Infrastructure NZ Limited',
            'VSL.NZ': 'Vulcan Steel Limited',
            'VGL.NZ': 'Vista Group International Limited',
            'HGH.NZ': 'Heartland Group Holdings Limited',
            'SKC.NZ': 'SkyCity Entertainment Group Limited',
            'SCL.NZ': 'Scales Corporation Limited',
            'SPG.NZ': 'Stride Property Group',
            'NPH.NZ': 'Napier Port Holdings Limited',
            'TRA.NZ': 'Turners Automotive Group Limited',
            'WIN.NZ': 'Winton Land Limited',
            'OCA.NZ': 'Oceania Healthcare Limited',
            'TWR.NZ': 'Tower Limited',
            'SAN.NZ': 'Sanford Limited',
            'HLG.NZ': 'Hallenstein Glasson Holdings Limited',
            'NZX.NZ': 'NZX Limited',
            'THL.NZ': 'Tourism Holdings Limited',
            'IPL.NZ': 'Investore Property Limited',
            'DGL.NZ': 'Delegat Group Limited',
            'MCK.NZ': 'Millennium & Copthorne Hotels NZ Limited',
            'SKT.NZ': 'SKY Network Television Limited',
            'SKO.NZ': 'Serko Limited',
            'RBD.NZ': 'Restaurant Brands New Zealand Limited',
            'SML.NZ': 'Synlait Milk Limited',
            
            # NZX Micro Cap ($100M - $300M) - 22 companies
            'WHS.NZ': 'The Warehouse Group Limited',
            'AFT.NZ': 'AFT Pharmaceuticals Limited',
            'SPY.NZ': 'Smartpay Holdings Limited',
            'ERD.NZ': 'EROAD Limited',
            'CDI.NZ': 'CDL Investments New Zealand Limited',
            'TGG.NZ': 'T&G Global Limited',
            'CMO.NZ': 'The Colonial Motor Company Limited',
            'NZM.NZ': 'NZME Limited',
            'MLN.NZ': 'Marlin Global Limited',
            'KMD.NZ': 'KMD Brands Limited',
            'SPN.NZ': 'South Port New Zealand Limited',
            'MHJ.NZ': 'Michael Hill International Limited',
            'SEK.NZ': 'Seeka Limited',
            'SCT.NZ': 'Scott Technology Limited',
            'PGW.NZ': 'PGG Wrightson Limited',
            'RAK.NZ': 'Rakon Limited',
            'IKE.NZ': 'ikeGPS Group Limited',
            'LIC.NZ': 'Livestock Improvement Corporation Limited',
            'NZL.NZ': 'New Zealand Rural Land Company Limited',
            'GXH.NZ': 'Green Cross Health Limited',
            'STU.NZ': 'Steel & Tube Holdings Limited',
            'NZK.NZ': 'New Zealand King Salmon Investments Limited',
            
            # NZX Nano Cap (<$100M) - 35 companies
            'RAD.NZ': 'Radius Residential Care Limited',
            'NWF.NZ': 'NZ Windfarms Limited',
            'PEB.NZ': 'Pacific Edge Limited',
            'BPG.NZ': 'Black Pearl Group Limited',
            'ARB.NZ': 'ArborGen Holdings Limited',
            'APL.NZ': 'Asset Plus Limited',
            'MFB.NZ': 'My Food Bag Group Limited',
            'FWL.NZ': 'Foley Wines Limited',
            'BRW.NZ': 'Bremworth Limited',
            'AOF.NZ': 'AoFrio Limited',
            'NTL.NZ': 'New Talisman Gold Mines Limited',
            'TAH.NZ': 'Third Age Health Services Limited',
            'CVT.NZ': 'Comvita Limited',
            '2CC.NZ': '2 Cheap Cars Group Limited',
            'PYS.NZ': 'PaySauce Limited',
            'TWL.NZ': 'Trade Window Holdings Limited',
            'GEN.NZ': 'General Capital Limited',
            'BLT.NZ': 'BLIS Technologies Limited',
            'MOV.NZ': 'MOVE Logistics Group Limited',
            'ALF.NZ': 'Allied Farmers Limited',
            'PHL.NZ': 'Promisia Healthcare Limited',
            'WCO.NZ': 'WasteCo Group Limited',
            'VTL.NZ': 'Vital Limited',
            'CCC.NZ': 'Cooks Coffee Company Limited',
            'TRU.NZ': 'TruScreen Group Limited',
            'ENS.NZ': 'Enprise Group Limited',
            'SVR.NZ': 'Savor Limited',
            'BFG.NZ': 'Burger Fuel Group Limited',
            'BAI.NZ': 'Being AI Limited',
            'SDL.NZ': 'Solution Dynamics Limited',
            'MPG.NZ': 'Metro Performance Glass Limited',
            'AGL.NZ': 'Accordant Group Limited',
            'CRP.NZ': 'Chatham Rock Phosphate Limited',
            'RUA.NZ': 'Rua Bioscience Limited',
            'ME.NZ': 'Me Today Limited',
            'RTO.NZ': 'RTO Limited',
            
            # International Stocks
            'BRK-B': 'Berkshire Hathaway Class B',
            'IWM': 'iShares Russell 2000 ETF',
            'MSFT': 'Microsoft Corporation',
            'META': 'Meta Platforms',
            'AAPL': 'Apple Inc.',
            'NVDA': 'NVIDIA Corporation',
            'SNOW': 'Snowflake Inc.',
            'AMZN': 'Amazon.com Inc.',
            'LMT': 'Lockheed Martin Corporation',
            'TSM': 'Taiwan Semiconductor Manufacturing',
            'INTC': 'Intel Corporation',
            'GOOGL': 'Alphabet Inc.',
            'AMD': 'Advanced Micro Devices',
            'RKLB': 'Rocket Lab USA',
            'AMAT': 'Applied Materials',
            'NVO': 'Novo Nordisk',
            'NOC': 'Northrop Grumman Corporation',
            'BRK-A': 'Berkshire Hathaway Class A',
            'SMI.AX': 'Santos Limited'
        }

        # Market indices for context
        self.indices = {
            '^GSPC': 'S&P 500',
            '^IXIC': 'NASDAQ',
            '^VIX': 'VIX',
            '^TNX': '10-Year Treasury'
        }

        # Valuation parameters from configuration
        self.discount_rate = VALUATION_CONFIG['discount_rate']
        self.perpetual_growth_rate = VALUATION_CONFIG['perpetual_growth_rate']
        self.projection_years = VALUATION_CONFIG['projection_years']

    def _get_downloads_directory(self):
        """Get the Downloads/Stock Valuation directory with fallback to james.walsham search"""
        # Primary path
        primary_path = r"C:\Users\james\Downloads\Stock Valuation"
        if os.path.exists(os.path.dirname(primary_path)):
            return primary_path
        
        # Fallback: search for james.walsham in Users directory
        users_dir = r"C:\Users"
        if os.path.exists(users_dir):
            for item in os.listdir(users_dir):
                if "james.walsham" in item.lower():
                    fallback_path = os.path.join(users_dir, item, "Downloads", "Stock Valuation")
                    if os.path.exists(os.path.dirname(fallback_path)):
                        return fallback_path
        
        # If neither found, return None
        self.logger.warning("Could not find Downloads directory for james.walsham")
        return None

    def try_ticker_variations(self, ticker: str) -> Optional[str]:
        """Try different ticker variations to find valid data"""
        variations = [ticker, f"{ticker}.NZ", ticker.replace('.NZ', ''), ticker.replace('.AX', '')]
        
        for variant in variations:
            try:
                stock = yf.Ticker(variant)
                info = stock.info
                if info.get('regularMarketPrice') or info.get('currentPrice'):
                    self.logger.info(f"Found valid ticker: {variant} for {ticker}")
                    return variant
            except Exception as e:
                self.logger.debug(f"Ticker {variant} failed: {e}")
                continue
        
        self.logger.warning(f"No valid ticker found for {ticker}")
        return None

    def is_data_stale(self, stock_data: Dict) -> bool:
        """Check if market data is stale"""
        try:
            # For simplicity, assume data is stale if older than 15 minutes during market hours
            # In production, you'd check actual market hours
            data_age = datetime.now() - datetime.now()  # Placeholder - would need actual timestamp
            return False  # Placeholder implementation
        except Exception:
            return False

    @rate_limit()
    def get_stock_data(self, ticker: str, period: str = "1d") -> Optional[Dict]:
        """Get comprehensive stock data using yfinance with improved error handling"""
        try:
            # Try ticker variations first
            valid_ticker = self.try_ticker_variations(ticker)
            if not valid_ticker:
                raise ValueError(f"No valid ticker found for {ticker}")
            
            stock = yf.Ticker(valid_ticker)
            hist = stock.history(period=period)
            info = stock.info

            if hist.empty:
                raise ValueError(f"No historical data available for {valid_ticker}")

            current_price = hist['Close'].iloc[-1]
            
            # Validate essential data
            if not current_price or current_price <= 0:
                raise ValueError(f"Invalid current price for {valid_ticker}: {current_price}")
            
            # Get comprehensive financial data
            financial_data = {
                'ticker': ticker,
                'valid_ticker': valid_ticker,
                'current_price': current_price,
                'market_cap': info.get('marketCap', 0),
                'shares_outstanding': info.get('sharesOutstanding', 0),
                'trailing_pe': info.get('trailingPE', None),
                'forward_pe': info.get('forwardPE', None),
                'peg_ratio': info.get('pegRatio', None),
                'price_to_sales': info.get('priceToSalesTrailing12Months', None),
                'price_to_book': info.get('priceToBook', None),
                'dividend_yield': info.get('dividendYield', 0) if info.get('dividendYield') else 0,  # Already in percentage
                'dividend_rate': info.get('dividendRate', 0),
                'trailing_eps': info.get('trailingEps', None),
                'forward_eps': info.get('forwardEps', None),
                'revenue_growth': info.get('revenueGrowth', None),
                'earnings_growth': info.get('earningsGrowth', None),
                'free_cashflow': info.get('freeCashflow', None),
                'total_cash': info.get('totalCash', 0),
                'total_debt': info.get('totalDebt', 0),
                'book_value': info.get('bookValue', None),
                'return_on_equity': info.get('returnOnEquity', None),
                'return_on_assets': info.get('returnOnAssets', None),
                'debt_to_equity': info.get('debtToEquity', None),
                'current_ratio': info.get('currentRatio', None),
                'volume': hist['Volume'].iloc[-1] if 'Volume' in hist.columns else None,
                'avg_volume': info.get('averageVolume', None),
                'beta': info.get('beta', None),
                '52_week_high': info.get('fiftyTwoWeekHigh', None),
                '52_week_low': info.get('fiftyTwoWeekLow', None),
                'enterprise_value': info.get('enterpriseValue', None),
                'net_income': info.get('netIncomeToCommon', None),
                'revenue': info.get('totalRevenue', None),
                'ebitda': info.get('ebitda', None),
                'sector': info.get('sector', 'Unknown')
            }
            
            return financial_data
        except Exception as e:
            self.logger.error(f"Error getting data for {ticker}: {e}")
            return None

    @rate_limit()
    def get_historical_financial_data(self, ticker: str, period: str = "5y") -> Optional[Dict]:
        """Get historical financial data for growth rate calculations"""
        try:
            # Use the valid ticker if available
            valid_ticker = ticker
            if '.' in ticker:  # Try variations for NZ/AU stocks
                variations = [ticker, f"{ticker}.NZ", ticker.replace('.NZ', ''), ticker.replace('.AX', '')]
                for variant in variations:
                    try:
                        stock = yf.Ticker(variant)
                        hist = stock.history(period=period)
                        if not hist.empty:
                            valid_ticker = variant
                            break
                    except:
                        continue
            
            stock = yf.Ticker(valid_ticker)
            hist = stock.history(period=period)
            
            if hist.empty:
                return None
            
            # Calculate historical growth rates
            hist_data = {
                'price_history': hist['Close'].tolist(),
                'dates': hist.index.tolist(),
                'volume_history': hist['Volume'].tolist() if 'Volume' in hist.columns else []
            }
            
            # Calculate price growth rates
            if len(hist_data['price_history']) > 1:
                price_growth_rates = []
                for i in range(1, len(hist_data['price_history'])):
                    if hist_data['price_history'][i-1] > 0:  # Avoid division by zero
                        growth_rate = (hist_data['price_history'][i] - hist_data['price_history'][i-1]) / hist_data['price_history'][i-1]
                        price_growth_rates.append(growth_rate)
                
                hist_data['price_growth_rates'] = price_growth_rates
                hist_data['avg_price_growth_rate'] = np.mean(price_growth_rates) if price_growth_rates else 0
                
                # Calculate 5-year growth rate with division by zero protection
                if len(hist_data['price_history']) >= 5 and hist_data['price_history'][0] > 0:
                    hist_data['5_year_price_growth_rate'] = ((hist_data['price_history'][-1] / hist_data['price_history'][0]) ** (1/5)) - 1
                else:
                    hist_data['5_year_price_growth_rate'] = 0
            
            return hist_data
        except Exception as e:
            self.logger.error(f"Error getting historical data for {ticker}: {e}")
            return None

    def calculate_peter_lynch_valuation(self, stock_data, historical_data):
        """Calculate Peter Lynch valuation model"""
        try:
            if not stock_data:
                return None
            
            # Get required inputs
            current_price = stock_data.get('current_price', 0)
            trailing_pe = stock_data.get('trailing_pe', 0)
            forward_pe = stock_data.get('forward_pe', 0)
            dividend_yield = stock_data.get('dividend_yield', 0)
            earnings_growth = stock_data.get('earnings_growth', 0)
            sector = stock_data.get('sector', 'Unknown')
            
            # Use forward P/E if available, otherwise trailing P/E
            pe_ratio = forward_pe if forward_pe and forward_pe > 0 else trailing_pe
            
            if not pe_ratio or pe_ratio <= 0:
                self.logger.warning(f"No valid P/E ratio for Lynch calculation")
                return {
                    'lynch_ratio': None,
                    'valuation_status': 'N/A',
                    'valuation_color': 'gray',
                    'intrinsic_value': None,
                    'delta_percentage': None,
                    'eps_growth_rate': None,
                    'pe_ratio': None,
                    'dividend_yield': dividend_yield,
                    'weighted_growth_rate': None,
                    'advanced_lynch_ratio': None,
                    'advanced_intrinsic_value': None,
                    'advanced_delta_percentage': None
                }
            
            # Calculate earnings growth rate (keep as decimal, not percentage)
            if earnings_growth:
                eps_growth_rate = earnings_growth
            elif historical_data and historical_data.get('5_year_price_growth_rate'):
                # Use historical price growth as proxy for earnings growth
                eps_growth_rate = historical_data['5_year_price_growth_rate']
            else:
                # Conservative estimate
                eps_growth_rate = 0.10  # 10% growth
            
            # Keep as decimal for calculations, convert to percentage only for display
            eps_growth_rate_decimal = eps_growth_rate
            eps_growth_rate_percentage = eps_growth_rate * 100
            
            # Convert dividend yield from percentage to decimal for calculation
            dividend_yield_decimal = dividend_yield / 100 if dividend_yield else 0
            
            # Basic Lynch Ratio Calculation with division by zero protection
            if pe_ratio and pe_ratio > 0:
                lynch_ratio = (eps_growth_rate_decimal + dividend_yield_decimal) / pe_ratio
                # Cap extreme Lynch ratios to prevent unrealistic valuations
                lynch_ratio = min(lynch_ratio, 5.0)  # Cap at 5x (500% upside)
            else:
                # If no P/E ratio or negative earnings, use sector-specific default
                sector_defaults = {
                    'Technology': 25, 'Healthcare': 20, 'Financial Services': 12,
                    'Energy': 15, 'Utilities': 18, 'Consumer Staples': 18,
                    'Real Estate': 15, 'Unknown': 15
                }
                default_pe = sector_defaults.get(sector, 15)
                lynch_ratio = (eps_growth_rate_decimal + dividend_yield_decimal) / default_pe
                lynch_ratio = min(lynch_ratio, 3.0)  # More conservative cap for default P/E
            
            # Interpretation using configuration thresholds
            thresholds = VALUATION_CONFIG['lynch_thresholds']
            if lynch_ratio >= thresholds['VERY_UNDERVALUED']:
                valuation_status = "VERY UNDERVALUED"
                valuation_color = "green"
            elif lynch_ratio >= thresholds['UNDERVALUED']:
                valuation_status = "UNDERVALUED"
                valuation_color = "light_green"
            elif lynch_ratio >= thresholds['FAIRLY_VALUED']:
                valuation_status = "FAIRLY VALUED"
                valuation_color = "yellow"
            else:
                valuation_status = "OVERVALUED"
                valuation_color = "red"
            
            # Calculate intrinsic value per share with division by zero protection
            if current_price and current_price > 0 and lynch_ratio > 0:
                intrinsic_value = current_price * lynch_ratio
                delta = (intrinsic_value / current_price) - 1
            else:
                intrinsic_value = 0
                delta = 0
            
            # Advanced weighted growth rate calculation (keep as decimal)
            forward_eps_growth = eps_growth_rate_decimal
            historical_eps_growth = historical_data.get('5_year_price_growth_rate', 0) if historical_data else 0
            
            weighted_growth_rate = ((forward_eps_growth * 2) + historical_eps_growth) / 3
            
            # Advanced Lynch Ratio with weighted growth
            if pe_ratio and pe_ratio > 0:
                advanced_lynch_ratio = (weighted_growth_rate + dividend_yield_decimal) / pe_ratio
                # Cap extreme ratios
                advanced_lynch_ratio = min(advanced_lynch_ratio, 5.0)
            else:
                # Use sector default for advanced calculation too
                sector_defaults = {
                    'Technology': 25, 'Healthcare': 20, 'Financial Services': 12,
                    'Energy': 15, 'Utilities': 18, 'Consumer Staples': 18,
                    'Real Estate': 15, 'Unknown': 15
                }
                default_pe = sector_defaults.get(sector, 15)
                advanced_lynch_ratio = (weighted_growth_rate + dividend_yield_decimal) / default_pe
                advanced_lynch_ratio = min(advanced_lynch_ratio, 3.0)
            
            advanced_intrinsic_value = current_price * advanced_lynch_ratio
            advanced_delta = (advanced_intrinsic_value / current_price) - 1
            
            return {
                'lynch_ratio': lynch_ratio,
                'valuation_status': valuation_status,
                'valuation_color': valuation_color,
                'intrinsic_value': intrinsic_value,
                'delta_percentage': delta * 100,
                'eps_growth_rate': eps_growth_rate_percentage,  # Return percentage for display
                'pe_ratio': pe_ratio,
                'dividend_yield': dividend_yield,  # Already in percentage
                'weighted_growth_rate': weighted_growth_rate * 100,  # Convert to percentage for display
                'advanced_lynch_ratio': advanced_lynch_ratio,
                'advanced_intrinsic_value': advanced_intrinsic_value,
                'advanced_delta_percentage': advanced_delta * 100
            }
        except Exception as e:
            self.logger.error(f"Error calculating Peter Lynch valuation: {e}")
            return None

    def calculate_dcf_valuation(self, stock_data: Dict, historical_data: Optional[Dict]) -> Optional[Dict]:
        """Calculate Discounted Cash Flow (DCF) valuation with improved assumptions"""
        try:
            if not stock_data:
                return None
            
            # Get required inputs
            current_price = stock_data.get('current_price', 0)
            free_cashflow = stock_data.get('free_cashflow', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            total_cash = stock_data.get('total_cash', 0)
            total_debt = stock_data.get('total_debt', 0)
            
            if not free_cashflow or free_cashflow <= 0:
                self.logger.warning(f"No valid free cash flow for DCF calculation")
                return {
                    'fcf_growth_rate': None,
                    'discount_rate': self.discount_rate,
                    'perpetual_growth_rate': self.perpetual_growth_rate,
                    'projection_years': self.projection_years,
                    'terminal_value': None,
                    'enterprise_value': None,
                    'equity_value': None,
                    'intrinsic_value_per_share': None,
                    'delta_percentage': None,
                    'valuation_status': 'N/A',
                    'valuation_color': 'gray',
                    'sum_pv_fcfs': None,
                    'present_value_terminal': None,
                    'future_fcfs': [],
                    'present_value_fcfs': []
                }
            
            # Calculate FCF growth rate more conservatively
            fcf_growth_rate = VALUATION_CONFIG['default_fcf_growth']  # Start with 8%
            
            if historical_data and historical_data.get('avg_price_growth_rate'):
                # Use historical growth as proxy but cap it
                historical_growth = historical_data['avg_price_growth_rate']
                fcf_growth_rate = max(
                    VALUATION_CONFIG['min_fcf_growth'],
                    min(VALUATION_CONFIG['max_fcf_growth'], historical_growth * 0.7)  # 70% of price growth
                )
            
            # Additional conservative adjustments
            if stock_data.get('beta', 1) > 1.5:  # High beta stocks
                fcf_growth_rate *= 0.8  # Reduce growth by 20%
            elif stock_data.get('beta', 1) < 0.8:  # Low beta stocks
                fcf_growth_rate *= 1.1  # Increase growth by 10%
            
            # Cap the growth rate
            fcf_growth_rate = max(
                VALUATION_CONFIG['min_fcf_growth'],
                min(VALUATION_CONFIG['max_fcf_growth'], fcf_growth_rate)
            )
            
            # Project future free cash flows
            future_fcfs = []
            present_value_fcfs = []
            
            current_fcf = free_cashflow
            
            for year in range(1, self.projection_years + 1):
                # Project future FCF
                future_fcf = current_fcf * ((1 + fcf_growth_rate) ** year)
                future_fcfs.append(future_fcf)
                
                # Calculate present value
                present_value = future_fcf / ((1 + self.discount_rate) ** year)
                present_value_fcfs.append(present_value)
            
            # Calculate terminal value with sanity check
            last_year_fcf = future_fcfs[-1]
            terminal_value = (last_year_fcf * (1 + self.perpetual_growth_rate)) / (self.discount_rate - self.perpetual_growth_rate)
            present_value_terminal = terminal_value / ((1 + self.discount_rate) ** self.projection_years)
            
            # Calculate enterprise value
            sum_pv_fcfs = sum(present_value_fcfs)
            enterprise_value = sum_pv_fcfs + present_value_terminal
            
            # Sanity check: terminal value shouldn't exceed 80% of total value
            terminal_ratio = present_value_terminal / enterprise_value if enterprise_value > 0 else 0
            if terminal_ratio > VALUATION_CONFIG['max_terminal_value_ratio']:
                self.logger.warning(f"Terminal value ratio too high: {terminal_ratio:.2%}, capping at {VALUATION_CONFIG['max_terminal_value_ratio']:.2%}")
                # Recalculate with capped terminal value
                max_terminal_value = sum_pv_fcfs * VALUATION_CONFIG['max_terminal_value_ratio'] / (1 - VALUATION_CONFIG['max_terminal_value_ratio'])
                enterprise_value = sum_pv_fcfs + max_terminal_value
                present_value_terminal = max_terminal_value
            
            # Calculate equity value
            equity_value = enterprise_value + total_cash - total_debt
            
            # Calculate intrinsic value per share with division by zero protection
            if shares_outstanding and shares_outstanding > 0:
                intrinsic_value_per_share = equity_value / shares_outstanding
            else:
                intrinsic_value_per_share = 0
            
            # Calculate delta with division by zero protection
            if current_price and current_price > 0:
                delta = (intrinsic_value_per_share / current_price) - 1
            else:
                delta = 0
            
            # Valuation assessment
            if delta > 0.2:
                valuation_status = "SIGNIFICANTLY UNDERVALUED"
                valuation_color = "green"
            elif delta > 0.05:
                valuation_status = "UNDERVALUED"
                valuation_color = "light_green"
            elif delta > -0.05:
                valuation_status = "FAIRLY VALUED"
                valuation_color = "yellow"
            elif delta > -0.2:
                valuation_status = "OVERVALUED"
                valuation_color = "orange"
            else:
                valuation_status = "SIGNIFICANTLY OVERVALUED"
                valuation_color = "red"
            
            return {
                'fcf_growth_rate': fcf_growth_rate,
                'discount_rate': self.discount_rate,
                'perpetual_growth_rate': self.perpetual_growth_rate,
                'projection_years': self.projection_years,
                'terminal_value': terminal_value,
                'enterprise_value': enterprise_value,
                'equity_value': equity_value,
                'intrinsic_value_per_share': intrinsic_value_per_share,
                'delta_percentage': delta * 100,
                'valuation_status': valuation_status,
                'valuation_color': valuation_color,
                'sum_pv_fcfs': sum_pv_fcfs,
                'present_value_terminal': present_value_terminal,
                'future_fcfs': future_fcfs,
                'present_value_fcfs': present_value_fcfs,
                'terminal_ratio': terminal_ratio
            }
        except Exception as e:
            self.logger.error(f"Error calculating DCF valuation: {e}")
            return None

    def calculate_enhanced_dcf_valuation(self, stock_data):
        """Calculate enhanced DCF with multi-scenario analysis and industry-specific adjustments"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            free_cashflow = stock_data.get('free_cashflow', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            sector = stock_data.get('sector', 'Unknown')
            beta = stock_data.get('beta', 1.0)
            
            if not free_cashflow or free_cashflow <= 0 or not shares_outstanding:
                return {
                    'enhanced_dcf_intrinsic_value': None,
                    'enhanced_dcf_status': 'N/A',
                    'enhanced_dcf_delta': 0,
                    'enhanced_dcf_scenarios': None
                }
            
            # Industry-specific WACC adjustments
            base_discount_rate = self.discount_rate
            industry_adjustments = {
                'Technology': 0.5, 'Healthcare': 0.3, 'Financial Services': 0.4,
                'Energy': 0.6, 'Utilities': -0.5, 'Consumer Staples': -0.3,
                'Real Estate': 0.2
            }
            discount_rate = base_discount_rate + industry_adjustments.get(sector, 0)
            
            # Multi-scenario analysis
            scenarios = {
                'bear': {'probability': 0.25, 'growth_multiplier': 0.5, 'terminal_multiplier': 0.8},
                'base': {'probability': 0.5, 'growth_multiplier': 1.0, 'terminal_multiplier': 1.0},
                'bull': {'probability': 0.25, 'growth_multiplier': 1.5, 'terminal_multiplier': 1.2}
            }
            
            # Calculate base growth rate with fade periods
            historical_growth = stock_data.get('free_cashflow_growth_rate', 0)
            if historical_growth and historical_growth > 0:
                adjusted_growth = min(historical_growth * 0.7, 0.12)  # Convert to decimal
                if beta > 1.2:
                    adjusted_growth *= 0.8
                elif beta < 0.8:
                    adjusted_growth *= 1.1
                base_growth_rate = max(adjusted_growth, 0.02)  # Convert to decimal
            else:
                base_growth_rate = 0.05  # 5% as decimal
            
            scenario_values = {}
            
            for scenario_name, scenario_data in scenarios.items():
                # Calculate growth rates with fade periods
                growth_rates = []
                scenario_growth = base_growth_rate * scenario_data['growth_multiplier']
                
                # Fade period: growth gradually declines to perpetual rate
                fade_years = 5
                for year in range(1, self.projection_years + 1):
                    if year <= fade_years:
                        fade_factor = (fade_years - year + 1) / fade_years
                        year_growth = self.perpetual_growth_rate + (scenario_growth - self.perpetual_growth_rate) * fade_factor
                    else:
                        year_growth = self.perpetual_growth_rate
                    growth_rates.append(year_growth)
                
                # Calculate projected FCF with fade periods
                projected_fcf = []
                current_fcf = free_cashflow
                
                for year in range(self.projection_years):
                    current_fcf *= (1 + growth_rates[year])  # growth_rates is now decimal
                    projected_fcf.append(current_fcf)
                
                # Calculate terminal value with scenario adjustment
                last_year_fcf = projected_fcf[-1]
                terminal_growth = self.perpetual_growth_rate * scenario_data['terminal_multiplier']
                terminal_value = (last_year_fcf * (1 + terminal_growth)) / (discount_rate / 100 - terminal_growth)
                
                # Calculate present value of projected FCF
                pv_fcf = []
                for i, fcf in enumerate(projected_fcf):
                    pv = fcf / ((1 + discount_rate / 100) ** (i + 1))
                    pv_fcf.append(pv)
                
                # Calculate present value of terminal value
                pv_terminal = terminal_value / ((1 + discount_rate / 100) ** self.projection_years)
                
                # Total enterprise value for this scenario
                enterprise_value = sum(pv_fcf) + pv_terminal
                intrinsic_value_per_share = enterprise_value / shares_outstanding
                
                scenario_values[scenario_name] = {
                    'intrinsic_value': intrinsic_value_per_share,
                    'probability': scenario_data['probability'],
                    'growth_rate': scenario_growth,
                    'terminal_value': terminal_value
                }
            
            # Calculate probability-weighted intrinsic value
            weighted_value = sum(scenario_values[s]['intrinsic_value'] * scenario_values[s]['probability'] 
                                for s in scenario_values)
            
            # Determine valuation status
            if weighted_value > current_price * 1.2:
                valuation_status = "SIGNIFICANTLY UNDERVALUED"
            elif weighted_value > current_price * 1.05:
                valuation_status = "UNDERVALUED"
            elif weighted_value > current_price * 0.95:
                valuation_status = "FAIRLY VALUED"
            elif weighted_value > current_price * 0.8:
                valuation_status = "OVERVALUED"
            else:
                valuation_status = "SIGNIFICANTLY OVERVALUED"
            
            delta_percentage = ((weighted_value - current_price) / current_price) * 100
            
            return {
                'enhanced_dcf_intrinsic_value': weighted_value,
                'enhanced_dcf_status': valuation_status,
                'enhanced_dcf_delta': delta_percentage,
                'enhanced_dcf_scenarios': scenario_values,
                'enhanced_dcf_discount_rate': discount_rate,
                'enhanced_dcf_base_growth_rate': base_growth_rate
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating enhanced DCF valuation: {e}")
            return {
                'enhanced_dcf_intrinsic_value': None,
                'enhanced_dcf_status': 'N/A',
                'enhanced_dcf_delta': 0,
                'enhanced_dcf_scenarios': None
            }

    def calculate_relative_valuation(self, stock_data):
        """Calculate relative valuation using comparable company multiples"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            sector = stock_data.get('sector', 'Unknown')
            
            # Get financial metrics
            enterprise_value = stock_data.get('enterprise_value', 0)
            ebitda = stock_data.get('ebitda', 0)
            net_income = stock_data.get('net_income', 0)
            revenue = stock_data.get('revenue', 0)
            book_value = stock_data.get('book_value', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            
            if not shares_outstanding or shares_outstanding <= 0:
                return {
                    'relative_valuation_status': 'N/A',
                    'relative_valuation_delta': 0,
                    'ev_ebitda_multiple': None,
                    'pe_ratio': None,
                    'ps_ratio': None,
                    'pb_ratio': None
                }
            
            # Calculate multiples
            ev_ebitda_multiple = None
            pe_ratio = None
            ps_ratio = None
            pb_ratio = None
            
            if ebitda and ebitda > 0:
                ev_ebitda_multiple = enterprise_value / ebitda
            
            if net_income and net_income > 0 and shares_outstanding and shares_outstanding > 0:
                pe_ratio = (current_price * shares_outstanding) / net_income
            
            if revenue and revenue > 0 and shares_outstanding and shares_outstanding > 0:
                ps_ratio = (current_price * shares_outstanding) / revenue
            
            if book_value and book_value > 0 and shares_outstanding and shares_outstanding > 0:
                pb_ratio = (current_price * shares_outstanding) / book_value
            
            # Sector-specific median multiples (simplified - in practice, you'd use real sector data)
            sector_multiples = {
                'Technology': {'ev_ebitda': 15, 'pe': 25, 'ps': 5, 'pb': 3},
                'Healthcare': {'ev_ebitda': 12, 'pe': 20, 'ps': 4, 'pb': 2.5},
                'Financial Services': {'ev_ebitda': 8, 'pe': 12, 'ps': 2, 'pb': 1.2},
                'Energy': {'ev_ebitda': 6, 'pe': 15, 'ps': 1.5, 'pb': 1.5},
                'Utilities': {'ev_ebitda': 10, 'pe': 18, 'ps': 2.5, 'pb': 1.8},
                'Consumer Staples': {'ev_ebitda': 12, 'pe': 18, 'ps': 2, 'pb': 2},
                'Real Estate': {'ev_ebitda': 15, 'pe': 20, 'ps': 8, 'pb': 1.5}
            }
            
            sector_medians = sector_multiples.get(sector, {'ev_ebitda': 12, 'pe': 18, 'ps': 3, 'pb': 2})
            
            # Calculate valuation scores
            valuation_scores = []
            
            if ev_ebitda_multiple and sector_medians['ev_ebitda']:
                ev_ebitda_score = (sector_medians['ev_ebitda'] - ev_ebitda_multiple) / sector_medians['ev_ebitda']
                valuation_scores.append(ev_ebitda_score)
            
            if pe_ratio and sector_medians['pe']:
                pe_score = (sector_medians['pe'] - pe_ratio) / sector_medians['pe']
                valuation_scores.append(pe_score)
            
            if ps_ratio and sector_medians['ps']:
                ps_score = (sector_medians['ps'] - ps_ratio) / sector_medians['ps']
                valuation_scores.append(ps_score)
            
            if pb_ratio and sector_medians['pb']:
                pb_score = (sector_medians['pb'] - pb_ratio) / sector_medians['pb']
                valuation_scores.append(pb_score)
            
            # Calculate average valuation score
            if valuation_scores:
                avg_score = sum(valuation_scores) / len(valuation_scores)
                
                if avg_score > 0.2:
                    valuation_status = "SIGNIFICANTLY UNDERVALUED"
                elif avg_score > 0.05:
                    valuation_status = "UNDERVALUED"
                elif avg_score > -0.05:
                    valuation_status = "FAIRLY VALUED"
                elif avg_score > -0.2:
                    valuation_status = "OVERVALUED"
                else:
                    valuation_status = "SIGNIFICANTLY OVERVALUED"
                
                delta_percentage = avg_score * 100
            else:
                valuation_status = "N/A"
                delta_percentage = 0
            
            return {
                'relative_valuation_status': valuation_status,
                'relative_valuation_delta': delta_percentage,
                'ev_ebitda_multiple': ev_ebitda_multiple,
                'pe_ratio': pe_ratio,
                'ps_ratio': ps_ratio,
                'pb_ratio': pb_ratio,
                'sector_medians': sector_medians,
                'valuation_scores': valuation_scores
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating relative valuation: {e}")
            return {
                'relative_valuation_status': 'N/A',
                'relative_valuation_delta': 0,
                'ev_ebitda_multiple': None,
                'pe_ratio': None,
                'ps_ratio': None,
                'pb_ratio': None
            }

    def calculate_munger_farm_valuation(self, stock_data):
        """Calculate Charlie Munger's farm valuation concept"""
        try:
            if not stock_data:
                return None
            
            # Get required inputs
            current_price = stock_data.get('current_price', 0)
            trailing_eps = stock_data.get('trailing_eps', 0)
            market_cap = stock_data.get('market_cap', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            
            if not trailing_eps or trailing_eps <= 0:
                self.logger.warning(f"No valid EPS for Munger farm valuation")
                return {
                    'annual_profit_per_share': None,
                    'farm_valuations': {},
                    'total_company_valuations': {},
                    'current_price': current_price,
                    'market_cap': market_cap
                }
            
            # Calculate annual profit per share (EPS)
            annual_profit_per_share = trailing_eps
            
            # Different required return scenarios
            required_returns = [0.05, 0.07, 0.10, 0.12, 0.15]  # 5%, 7%, 10%, 12%, 15%
            
            farm_valuations = {}
            
            for required_return in required_returns:
                # Calculate intrinsic value based on required return
                intrinsic_value = annual_profit_per_share / required_return
                
                # Calculate delta
                delta = (intrinsic_value / current_price) - 1 if current_price > 0 else 0
                
                # Assessment
                if delta > 0.2:
                    assessment = "STRONG BUY"
                elif delta > 0.05:
                    assessment = "BUY"
                elif delta > -0.05:
                    assessment = "HOLD"
                elif delta > -0.2:
                    assessment = "SELL"
                else:
                    assessment = "STRONG SELL"
                
                farm_valuations[f"{int(required_return*100)}%_return"] = {
                    'required_return': required_return,
                    'intrinsic_value': intrinsic_value,
                    'delta_percentage': delta * 100,
                    'assessment': assessment
                }
            
            # Calculate total company valuation scenarios
            total_company_valuations = {}
            for required_return in required_returns:
                total_intrinsic_value = annual_profit_per_share * shares_outstanding / required_return
                total_delta = (total_intrinsic_value / market_cap) - 1 if market_cap > 0 else 0
                
                total_company_valuations[f"{int(required_return*100)}%_return"] = {
                    'required_return': required_return,
                    'total_intrinsic_value': total_intrinsic_value,
                    'total_delta_percentage': total_delta * 100
                }
            
            return {
                'annual_profit_per_share': annual_profit_per_share,
                'farm_valuations': farm_valuations,
                'total_company_valuations': total_company_valuations,
                'current_price': current_price,
                'market_cap': market_cap
            }
        except Exception as e:
            self.logger.error(f"Error calculating Munger farm valuation: {e}")
            return None

    def calculate_reverse_dcf(self, stock_data):
        """Calculate reverse DCF to determine what growth rate is priced in by the market"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            free_cashflow = stock_data.get('free_cashflow', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            sector = stock_data.get('sector', 'Unknown')
            
            if not free_cashflow or free_cashflow <= 0 or not shares_outstanding or not current_price:
                return {
                    'reverse_dcf_implied_growth': None,
                    'reverse_dcf_assessment': 'N/A',
                    'reverse_dcf_reasonable': None
                }
            
            # Industry-specific WACC adjustments
            base_discount_rate = self.discount_rate
            industry_adjustments = {
                'Technology': 0.5, 'Healthcare': 0.3, 'Financial Services': 0.4,
                'Energy': 0.6, 'Utilities': -0.5, 'Consumer Staples': -0.3,
                'Real Estate': 0.2
            }
            discount_rate = base_discount_rate + industry_adjustments.get(sector, 0)
            
            # Calculate implied growth rate using iterative approach
            market_cap = current_price * shares_outstanding
            target_enterprise_value = market_cap  # Simplified
            
            # Use binary search to find implied growth rate
            low_growth = 0.0
            high_growth = 0.5  # Cap at 50% growth
            relative_tolerance = 0.001  # 0.1% relative accuracy
            max_iterations = 100
            
            implied_growth = None
            
            for iteration in range(max_iterations):
                test_growth = (low_growth + high_growth) / 2
                
                # Calculate DCF with test growth rate
                projected_fcf = []
                current_fcf = free_cashflow
                
                for year in range(1, self.projection_years + 1):
                    current_fcf *= (1 + test_growth)
                    projected_fcf.append(current_fcf)
                
                # Calculate terminal value
                last_year_fcf = projected_fcf[-1]
                terminal_value = (last_year_fcf * (1 + self.perpetual_growth_rate)) / (discount_rate - self.perpetual_growth_rate)
                
                # Calculate present value of projected FCF
                pv_fcf = []
                for i, fcf in enumerate(projected_fcf):
                    pv = fcf / ((1 + discount_rate) ** (i + 1))
                    pv_fcf.append(pv)
                
                # Calculate present value of terminal value
                pv_terminal = terminal_value / ((1 + discount_rate) ** self.projection_years)
                
                # Total enterprise value
                calculated_enterprise_value = sum(pv_fcf) + pv_terminal
                
                # Check if we're close enough (use relative tolerance)
                if target_enterprise_value > 0 and abs(calculated_enterprise_value - target_enterprise_value) / target_enterprise_value < relative_tolerance:
                    implied_growth = test_growth
                    break
                
                # Adjust search range
                if calculated_enterprise_value < target_enterprise_value:
                    low_growth = test_growth
                else:
                    high_growth = test_growth
            
            if implied_growth is None:
                implied_growth = (low_growth + high_growth) / 2
            
            # Convert to percentage
            implied_growth_percentage = implied_growth * 100
            
            # Assess reasonableness
            historical_growth = stock_data.get('free_cashflow_growth_rate', 0)
            if historical_growth:
                historical_growth_percentage = historical_growth
            else:
                historical_growth_percentage = stock_data.get('avg_price_growth_rate', 0)
            
            # Determine if implied growth is reasonable
            if historical_growth_percentage > 0:
                growth_ratio = implied_growth_percentage / historical_growth_percentage
                
                if growth_ratio > 2.0:
                    assessment = "MARKET EXPECTS EXCESSIVE GROWTH"
                    reasonable = False
                elif growth_ratio > 1.5:
                    assessment = "MARKET EXPECTS HIGH GROWTH"
                    reasonable = False
                elif growth_ratio > 0.8:
                    assessment = "MARKET EXPECTS REASONABLE GROWTH"
                    reasonable = True
                elif growth_ratio > 0.5:
                    assessment = "MARKET EXPECTS LOW GROWTH"
                    reasonable = True
                else:
                    assessment = "MARKET EXPECTS DECLINING GROWTH"
                    reasonable = True
            else:
                assessment = "NO HISTORICAL GROWTH DATA"
                reasonable = None
            
            return {
                'reverse_dcf_implied_growth': implied_growth_percentage,
                'reverse_dcf_assessment': assessment,
                'reverse_dcf_reasonable': reasonable,
                'reverse_dcf_historical_growth': historical_growth_percentage,
                'reverse_dcf_growth_ratio': growth_ratio if historical_growth_percentage > 0 else None,
                'reverse_dcf_discount_rate': discount_rate
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating reverse DCF: {e}")
            return {
                'reverse_dcf_implied_growth': None,
                'reverse_dcf_assessment': 'N/A',
                'reverse_dcf_reasonable': None
            }

    def calculate_earnings_power_value(self, stock_data):
        """Calculate Bruce Greenwald's Earnings Power Value"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            net_income = stock_data.get('net_income', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            
            if not net_income or net_income <= 0 or not shares_outstanding:
                return {
                    'epv_intrinsic_value': None,
                    'epv_assessment': 'N/A',
                    'epv_delta': 0
                }
            
            # Normalize earnings (adjust for cyclicality)
            normalized_earnings = net_income  # Simplified - use multi-year average in practice
            
            # Capitalize at appropriate rate (conservative)
            epv_discount_rate = self.discount_rate + 0.02  # Add 2% for no-growth assumption
            
            # Calculate EPV
            epv_enterprise_value = normalized_earnings / epv_discount_rate
            epv_intrinsic_value = epv_enterprise_value / shares_outstanding
            
            # Determine valuation status
            if epv_intrinsic_value > current_price * 1.2:
                assessment = "SIGNIFICANTLY UNDERVALUED"
            elif epv_intrinsic_value > current_price * 1.05:
                assessment = "UNDERVALUED"
            elif epv_intrinsic_value > current_price * 0.95:
                assessment = "FAIRLY VALUED"
            elif epv_intrinsic_value > current_price * 0.8:
                assessment = "OVERVALUED"
            else:
                assessment = "SIGNIFICANTLY OVERVALUED"
            
            delta_percentage = ((epv_intrinsic_value - current_price) / current_price) * 100
            
            return {
                'epv_intrinsic_value': epv_intrinsic_value,
                'epv_assessment': assessment,
                'epv_delta': delta_percentage,
                'epv_normalized_earnings': normalized_earnings,
                'epv_discount_rate': epv_discount_rate
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating Earnings Power Value: {e}")
            return {
                'epv_intrinsic_value': None,
                'epv_assessment': 'N/A',
                'epv_delta': 0
            }

    def calculate_residual_income_model(self, stock_data):
        """Calculate Residual Income Model - particularly good for financial stocks"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            book_value = stock_data.get('book_value', 0)
            net_income = stock_data.get('net_income', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            sector = stock_data.get('sector', 'Unknown')
            roe = stock_data.get('return_on_equity', 0)
            
            if not book_value or book_value <= 0 or not shares_outstanding:
                return {
                    'rim_intrinsic_value': None,
                    'rim_assessment': 'N/A',
                    'rim_delta': 0
                }
            
            # Calculate book value per share
            book_value_per_share = book_value / shares_outstanding
            
            # Determine cost of equity based on sector
            sector_cost_of_equity = {
                'Financial Services': 0.10,  # 10% - higher risk
                'Technology': 0.12,          # 12% - high growth, high risk
                'Healthcare': 0.09,         # 9% - moderate risk
                'Energy': 0.11,             # 11% - cyclical risk
                'Utilities': 0.07,          # 7% - lower risk
                'Consumer Staples': 0.08,   # 8% - stable
                'Real Estate': 0.09         # 9% - moderate risk
            }
            
            cost_of_equity = sector_cost_of_equity.get(sector, 0.10)  # Default 10%
            
            # Calculate current ROE if not provided
            if not roe and net_income and book_value:
                roe = net_income / book_value
            
            # Project residual income for next 5 years
            projection_years = 5
            residual_income_projection = []
            current_book_value = book_value_per_share
            
            # Use current ROE or sector average if not available
            if not roe or roe <= 0:
                sector_roe_averages = {
                    'Financial Services': 0.12,
                    'Technology': 0.15,
                    'Healthcare': 0.10,
                    'Energy': 0.08,
                    'Utilities': 0.08,
                    'Consumer Staples': 0.12,
                    'Real Estate': 0.10
                }
                roe = sector_roe_averages.get(sector, 0.10)
            
            # Apply fade period - ROE gradually declines to cost of equity
            fade_years = 3
            terminal_roe = cost_of_equity
            
            for year in range(1, projection_years + 1):
                # Calculate ROE for this year (with fade)
                if year <= fade_years:
                    fade_factor = (fade_years - year + 1) / fade_years
                    year_roe = terminal_roe + (roe - terminal_roe) * fade_factor
                else:
                    year_roe = terminal_roe
                
                # Calculate expected earnings
                expected_earnings = current_book_value * year_roe
                
                # Calculate residual income
                residual_income = expected_earnings - (current_book_value * cost_of_equity)
                
                # Calculate present value of residual income
                pv_residual_income = residual_income / ((1 + cost_of_equity) ** year)
                
                residual_income_projection.append({
                    'year': year,
                    'roe': year_roe,
                    'expected_earnings': expected_earnings,
                    'residual_income': residual_income,
                    'pv_residual_income': pv_residual_income
                })
                
                # Update book value for next year
                current_book_value += expected_earnings
            
            # Calculate terminal value (assuming ROE = cost of equity)
            terminal_residual_income = residual_income_projection[-1]['residual_income']
            terminal_value = terminal_residual_income / cost_of_equity
            pv_terminal_value = terminal_value / ((1 + cost_of_equity) ** projection_years)
            
            # Calculate intrinsic value
            sum_pv_residual_income = sum(ri['pv_residual_income'] for ri in residual_income_projection)
            rim_intrinsic_value = book_value_per_share + sum_pv_residual_income + pv_terminal_value
            
            # Determine valuation status
            if rim_intrinsic_value > current_price * 1.2:
                assessment = "SIGNIFICANTLY UNDERVALUED"
            elif rim_intrinsic_value > current_price * 1.05:
                assessment = "UNDERVALUED"
            elif rim_intrinsic_value > current_price * 0.95:
                assessment = "FAIRLY VALUED"
            elif rim_intrinsic_value > current_price * 0.8:
                assessment = "OVERVALUED"
            else:
                assessment = "SIGNIFICANTLY OVERVALUED"
            
            delta_percentage = ((rim_intrinsic_value - current_price) / current_price) * 100
            
            return {
                'rim_intrinsic_value': rim_intrinsic_value,
                'rim_assessment': assessment,
                'rim_delta': delta_percentage,
                'rim_book_value_per_share': book_value_per_share,
                'rim_cost_of_equity': cost_of_equity,
                'rim_current_roe': roe,
                'rim_terminal_value': terminal_value,
                'rim_residual_income_projection': residual_income_projection,
                'rim_sum_pv_residual_income': sum_pv_residual_income
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating Residual Income Model: {e}")
            return {
                'rim_intrinsic_value': None,
                'rim_assessment': 'N/A',
                'rim_delta': 0
            }

    def collect_valuation_metrics(self, ticker):
        """Collect all valuation metrics for a specific ticker"""
        self.logger.info(f"Collecting valuation metrics for {ticker}")
        
        # Get current stock data
        stock_data = self.get_stock_data(ticker)
        if not stock_data:
            self.logger.error(f"Failed to get stock data for {ticker}")
            return None
        
        # Get historical data
        historical_data = self.get_historical_financial_data(ticker)
        
        # Calculate all valuation methods
        lynch_valuation = self.calculate_peter_lynch_valuation(stock_data, historical_data)
        dcf_valuation = self.calculate_dcf_valuation(stock_data, historical_data)
        munger_valuation = self.calculate_munger_farm_valuation(stock_data)
        
        # Calculate new Tier 1 and Tier 2 valuation methods
        enhanced_dcf_valuation = self.calculate_enhanced_dcf_valuation(stock_data)
        relative_valuation = self.calculate_relative_valuation(stock_data)
        reverse_dcf_valuation = self.calculate_reverse_dcf(stock_data)
        epv_valuation = self.calculate_earnings_power_value(stock_data)
        rim_valuation = self.calculate_residual_income_model(stock_data)
        
        # Combine all data
        valuation_data = {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'time': datetime.now().strftime('%H:%M:%S'),
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ticker': ticker,
            'company_name': self.focus_stocks.get(ticker, ticker)
        }
        
        # Add stock data
        valuation_data.update(stock_data)
        
        # Add historical data summary
        if historical_data:
            valuation_data.update({
                'avg_price_growth_rate': historical_data.get('avg_price_growth_rate', 0),
                '5_year_price_growth_rate': historical_data.get('5_year_price_growth_rate', 0),
                'price_history_length': len(historical_data.get('price_history', []))
            })
        
        # Add Lynch valuation
        if lynch_valuation:
            valuation_data.update({
                'lynch_ratio': lynch_valuation.get('lynch_ratio', 0),
                'lynch_valuation_status': lynch_valuation.get('valuation_status', 'N/A'),
                'lynch_intrinsic_value': lynch_valuation.get('intrinsic_value', 0),
                'lynch_delta_percentage': lynch_valuation.get('delta_percentage', 0),
                'lynch_eps_growth_rate': lynch_valuation.get('eps_growth_rate', 0),
                'advanced_lynch_ratio': lynch_valuation.get('advanced_lynch_ratio', 0),
                'advanced_lynch_intrinsic_value': lynch_valuation.get('advanced_intrinsic_value', 0),
                'advanced_lynch_delta_percentage': lynch_valuation.get('advanced_delta_percentage', 0)
            })
        
        # Add DCF valuation
        if dcf_valuation:
            valuation_data.update({
                'dcf_fcf_growth_rate': dcf_valuation.get('fcf_growth_rate', 0),
                'dcf_discount_rate': dcf_valuation.get('discount_rate', 0),
                'dcf_perpetual_growth_rate': dcf_valuation.get('perpetual_growth_rate', 0),
                'dcf_terminal_value': dcf_valuation.get('terminal_value', 0),
                'dcf_enterprise_value': dcf_valuation.get('enterprise_value', 0),
                'dcf_equity_value': dcf_valuation.get('equity_value', 0),
                'dcf_intrinsic_value_per_share': dcf_valuation.get('intrinsic_value_per_share', 0),
                'dcf_delta_percentage': dcf_valuation.get('delta_percentage', 0),
                'dcf_valuation_status': dcf_valuation.get('valuation_status', 'N/A')
            })
        
        # Add Munger valuation summary
        if munger_valuation:
            valuation_data.update({
                'munger_annual_profit_per_share': munger_valuation.get('annual_profit_per_share', 0),
                'munger_7pct_intrinsic_value': munger_valuation.get('farm_valuations', {}).get('7%_return', {}).get('intrinsic_value', 0),
                'munger_7pct_delta_percentage': munger_valuation.get('farm_valuations', {}).get('7%_return', {}).get('delta_percentage', 0),
                'munger_7pct_assessment': munger_valuation.get('farm_valuations', {}).get('7%_return', {}).get('assessment', 'N/A'),
                'munger_10pct_intrinsic_value': munger_valuation.get('farm_valuations', {}).get('10%_return', {}).get('intrinsic_value', 0),
                'munger_10pct_delta_percentage': munger_valuation.get('farm_valuations', {}).get('10%_return', {}).get('delta_percentage', 0),
                'munger_10pct_assessment': munger_valuation.get('farm_valuations', {}).get('10%_return', {}).get('assessment', 'N/A')
            })
        
        # Add Enhanced DCF valuation (Tier 1)
        if enhanced_dcf_valuation:
            valuation_data.update({
                'enhanced_dcf_intrinsic_value': enhanced_dcf_valuation.get('enhanced_dcf_intrinsic_value', 0),
                'enhanced_dcf_status': enhanced_dcf_valuation.get('enhanced_dcf_status', 'N/A'),
                'enhanced_dcf_delta': enhanced_dcf_valuation.get('enhanced_dcf_delta', 0),
                'enhanced_dcf_discount_rate': enhanced_dcf_valuation.get('enhanced_dcf_discount_rate', 0),
                'enhanced_dcf_base_growth_rate': enhanced_dcf_valuation.get('enhanced_dcf_base_growth_rate', 0)
            })
        
        # Add Relative Valuation (Tier 1)
        if relative_valuation:
            valuation_data.update({
                'relative_valuation_status': relative_valuation.get('relative_valuation_status', 'N/A'),
                'relative_valuation_delta': relative_valuation.get('relative_valuation_delta', 0),
                'ev_ebitda_multiple': relative_valuation.get('ev_ebitda_multiple', 0),
                'pe_ratio': relative_valuation.get('pe_ratio', 0),
                'ps_ratio': relative_valuation.get('ps_ratio', 0),
                'pb_ratio': relative_valuation.get('pb_ratio', 0)
            })
        
        # Add Reverse DCF (Tier 2)
        if reverse_dcf_valuation:
            valuation_data.update({
                'reverse_dcf_implied_growth': reverse_dcf_valuation.get('reverse_dcf_implied_growth', 0),
                'reverse_dcf_assessment': reverse_dcf_valuation.get('reverse_dcf_assessment', 'N/A'),
                'reverse_dcf_reasonable': reverse_dcf_valuation.get('reverse_dcf_reasonable', None),
                'reverse_dcf_historical_growth': reverse_dcf_valuation.get('reverse_dcf_historical_growth', 0),
                'reverse_dcf_growth_ratio': reverse_dcf_valuation.get('reverse_dcf_growth_ratio', 0)
            })
        
        # Add Earnings Power Value (Tier 2)
        if epv_valuation:
            valuation_data.update({
                'epv_intrinsic_value': epv_valuation.get('epv_intrinsic_value', 0),
                'epv_assessment': epv_valuation.get('epv_assessment', 'N/A'),
                'epv_delta': epv_valuation.get('epv_delta', 0),
                'epv_normalized_earnings': epv_valuation.get('epv_normalized_earnings', 0),
                'epv_discount_rate': epv_valuation.get('epv_discount_rate', 0)
            })
        
        # Add Residual Income Model (Tier 2)
        if rim_valuation:
            valuation_data.update({
                'rim_intrinsic_value': rim_valuation.get('rim_intrinsic_value', 0),
                'rim_assessment': rim_valuation.get('rim_assessment', 'N/A'),
                'rim_delta': rim_valuation.get('rim_delta', 0),
                'rim_book_value_per_share': rim_valuation.get('rim_book_value_per_share', 0),
                'rim_cost_of_equity': rim_valuation.get('rim_cost_of_equity', 0),
                'rim_current_roe': rim_valuation.get('rim_current_roe', 0)
            })
        
        self.logger.info(f"Collected {len(valuation_data)} valuation metrics for {ticker}")
        return valuation_data

    def load_existing_data(self):
        """Load existing valuation data if available and recent"""
        try:
            if not os.path.exists(self.master_file):
                return None
            
            # Load existing data
            df_existing = pd.read_excel(self.master_file, sheet_name='Valuation Data')
            
            # Check if data has required new fields
            required_fields = ['net_income', 'revenue', 'ebitda', 'sector']
            missing_fields = [field for field in required_fields if field not in df_existing.columns]
            
            if missing_fields:
                self.logger.info(f"Existing data missing required fields: {missing_fields}, will fetch fresh data")
                return None
            
            # Check if data is recent (within last 1 hour for fresh calculations)
            if 'analysis_timestamp' in df_existing.columns:
                latest_timestamp = pd.to_datetime(df_existing['analysis_timestamp']).max()
                time_diff = datetime.now() - latest_timestamp
                
                if time_diff.total_seconds() < 1 * 3600:  # 1 hour for fresh data
                    self.logger.info(f"Using existing data from {latest_timestamp} (age: {time_diff})")
                    return df_existing
            
            self.logger.info("Existing data is stale, will fetch fresh data")
            return None
            
        except Exception as e:
            self.logger.warning(f"Could not load existing data: {e}")
            return None

    def run(self):
        """Main execution method"""
        self.logger.info("Starting Stock Valuation Scraper")
        start_time = datetime.now()
        
        try:
            # Try to load existing data first
            existing_df = self.load_existing_data()
            
            if existing_df is not None:
                # Use existing data
                df = existing_df
                self.logger.info(f"Using existing data with {len(df)} stocks")
            else:
                # Collect fresh valuation metrics for focus stocks
                all_valuation_data = []
                
                for ticker, company_name in self.focus_stocks.items():
                    self.logger.info(f"Analyzing {company_name} ({ticker})")
                    valuation_data = self.collect_valuation_metrics(ticker)
                    if valuation_data:
                        all_valuation_data.append(valuation_data)
                
                if all_valuation_data:
                    # Create DataFrame
                    df = pd.DataFrame(all_valuation_data)
                    # Add timestamp
                    df['analysis_timestamp'] = datetime.now()
                else:
                    self.logger.error("No valuation data collected")
                    return
            
            # Save dataset
            self.save_dataset(df)
            
            # Print summary
            self.print_valuation_summary(df)
            
            end_time = datetime.now()
            duration = end_time - start_time
            self.logger.info(f"Stock valuation analysis completed in {duration}")
            
        except Exception as e:
            self.logger.error(f"Error during stock valuation analysis: {e}")

    def print_valuation_summary(self, df):
        """Print comprehensive valuation summary"""
        print(f"\n=== Stock Valuation Analysis Summary ===")
        print(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Stocks Analyzed: {len(df)}")
        
        for _, row in df.iterrows():
            ticker = row['ticker']
            company_name = row['company_name']
            current_price = row['current_price']
            
            print(f"\n--- {company_name} ({ticker}) ---")
            print(f"Current Price: ${current_price:,.2f}")
            
            # Lynch Valuation Summary
            if row.get('lynch_valuation_status') != 'N/A':
                print(f"Peter Lynch Valuation:")
                print(f"  Status: {row['lynch_valuation_status']}")
                print(f"  Lynch Ratio: {row.get('lynch_ratio', 0):.2f}")
                print(f"  Intrinsic Value: ${row.get('lynch_intrinsic_value', 0):,.2f}")
                print(f"  Delta: {row.get('lynch_delta_percentage', 0):+.1f}%")
            
            # DCF Valuation Summary
            if row.get('dcf_valuation_status') != 'N/A':
                print(f"DCF Valuation:")
                print(f"  Status: {row['dcf_valuation_status']}")
                print(f"  Intrinsic Value: ${row.get('dcf_intrinsic_value_per_share', 0):,.2f}")
                print(f"  Delta: {row.get('dcf_delta_percentage', 0):+.1f}%")
            
            # Munger Farm Valuation Summary
            if row.get('munger_7pct_assessment') != 'N/A':
                print(f"Munger Farm Valuation (7% return):")
                print(f"  Assessment: {row['munger_7pct_assessment']}")
                print(f"  Intrinsic Value: ${row.get('munger_7pct_intrinsic_value', 0):,.2f}")
                print(f"  Delta: {row.get('munger_7pct_delta_percentage', 0):+.1f}%")

    def save_dataset(self, df):
        """Save valuation dataset to Excel with formatting"""
        if df.empty:
            self.logger.warning("No data to save")
            return
        
        try:
            # Create timestamped filename for daily backup
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            daily_backup_file = os.path.join(self.daily_backups_dir, f"stock_valuation_dataset_{timestamp}.xlsx")
            
            # Try to load existing workbook first
            try:
                wb = load_workbook(self.master_file)
                self.logger.info("Loaded existing workbook, adding new data")
            except FileNotFoundError:
                # Create new workbook if file doesn't exist
                wb = Workbook()
                wb.remove(wb.active)
                self.logger.info("Created new workbook")
            
            # Create or update summary sheet
            if 'My Portfolio' in wb.sheetnames:
                wb.remove(wb['My Portfolio'])
            self.create_valuation_summary_sheet(df, wb)
            
            # Create or update prospects sheet
            if 'Prospects' in wb.sheetnames:
                wb.remove(wb['Prospects'])
            self.create_prospects_sheet(df, wb)
            
            # Create or update detailed data sheet
            if 'Valuation Data' in wb.sheetnames:
                # Check if we need to update headers for new columns
                ws_data = wb['Valuation Data']
                existing_headers = [cell.value for cell in ws_data[1]]
                new_headers = list(df.columns)
                
                # If headers don't match, recreate the sheet with new headers
                if existing_headers != new_headers:
                    self.logger.info("Updating Valuation Data sheet with new columns")
                    wb.remove(ws_data)
                    ws_data = wb.create_sheet('Valuation Data', 2)
                    # Add new headers
                    for col, header in enumerate(new_headers, 1):
                        ws_data.cell(row=1, column=col, value=header)
                    next_row = 2
                else:
                    # Find the next empty row
                    next_row = ws_data.max_row + 1
            else:
                ws_data = wb.create_sheet('Valuation Data', 2)
                next_row = 1
                # Add headers if new sheet
                headers = list(df.columns)
                for col, header in enumerate(headers, 1):
                    ws_data.cell(row=1, column=col, value=header)
                next_row = 2
            
            # Add new data to worksheet (skip header if not new sheet)
            for _, row in df.iterrows():
                for col, value in enumerate(row, 1):
                    ws_data.cell(row=next_row, column=col, value=value)
                next_row += 1
            
            # Auto-adjust column widths for Valuation Data sheet
            for column in ws_data.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 35)  # Cap at 35 for better readability
                ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # Save files
            wb.save(self.master_file)
            wb.save(daily_backup_file)
            
            if self.downloads_file:
                wb.save(self.downloads_file)
            
            self.logger.info(f"Valuation dataset saved to: {self.master_file}")
            self.logger.info(f"Daily backup saved to: {daily_backup_file}")
            
        except Exception as e:
            self.logger.error(f"Error saving valuation dataset: {e}")

    def create_valuation_summary_sheet(self, df, wb):
        """Create comprehensive valuation summary sheet with all 9 valuation methods"""
        try:
            ws_summary = wb.create_sheet('My Portfolio', 0)
            
            # Define styles
            header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=14)
            subheader_font = Font(bold=True, size=12)
            data_font = Font(size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Consistent conditional formatting fills (matching Prospects sheet)
            strong_buy_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            buy_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
            hold_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            sell_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            strong_sell_fill = PatternFill(start_color="FFA0A0", end_color="FFA0A0", fill_type="solid")
            no_data_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            row = 1
            
            # Title
            ws_summary[f'A{row}'] = "Stock Valuation Analysis Dashboard"
            ws_summary[f'A{row}'].font = Font(bold=True, size=16)
            ws_summary[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_summary.merge_cells(f'A{row}:I{row}')
            row += 1
            
            # Analysis date
            ws_summary[f'A{row}'] = f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_summary[f'A{row}'].font = Font(italic=True, size=10)
            row += 1
            
            ws_summary[f'A{row}'] = f"Stocks Analyzed: {len(df)}"
            ws_summary[f'A{row}'].font = Font(italic=True, size=10)
            row += 2
            
            # Headers for all 9 valuation methods
            ws_summary[f'A{row}'] = "Company (Ticker)"
            ws_summary[f'B{row}'] = "Peter Lynch"
            ws_summary[f'C{row}'] = "DCF Valuation"
            ws_summary[f'D{row}'] = "Munger Farm"
            ws_summary[f'E{row}'] = "Enhanced DCF"
            ws_summary[f'F{row}'] = "Relative Valuation"
            ws_summary[f'G{row}'] = "Reverse DCF"
            ws_summary[f'H{row}'] = "EPV/RIM"
            ws_summary[f'I{row}'] = "Current Price"
            
            # Format headers
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                cell = ws_summary[f'{col}{row}']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            row += 1
            
            # Separate stocks by completeness - most complete at top
            stocks_complete = []      # All 3 valuation methods
            stocks_partial = []       # 1-2 valuation methods
            stocks_without_data = []  # No valuation methods
            
            for _, stock_row in df.iterrows():
                lynch_status = stock_row.get('lynch_valuation_status', 'N/A')
                dcf_status = stock_row.get('dcf_valuation_status', 'N/A')
                munger_assessment = stock_row.get('munger_7pct_assessment', 'N/A')
                
                # Count how many valuation methods are available
                available_methods = 0
                if lynch_status != 'N/A':
                    available_methods += 1
                if dcf_status != 'N/A':
                    available_methods += 1
                if munger_assessment != 'N/A':
                    available_methods += 1
                
                if available_methods == 3:
                    stocks_complete.append(stock_row)
                elif available_methods > 0:
                    stocks_partial.append(stock_row)
                else:
                    stocks_without_data.append(stock_row)
            
            # Add stocks with complete data first (all 3 valuation methods)
            for stock_row in stocks_complete:
                ticker = stock_row['ticker']
                company_name = stock_row['company_name']
                current_price = stock_row['current_price']
                
                # Company name and ticker
                ws_summary[f'A{row}'] = f"{company_name} ({ticker})"
                ws_summary[f'A{row}'].font = data_font
                ws_summary[f'A{row}'].border = border
                ws_summary[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Current price
                ws_summary[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_summary[f'E{row}'].font = data_font
                ws_summary[f'E{row}'].border = border
                ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Peter Lynch Valuation (Column B)
                lynch_status = stock_row.get('lynch_valuation_status', 'N/A')
                lynch_delta = stock_row.get('lynch_delta_percentage', 0)
                
                if lynch_status != 'N/A':
                    lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
                    ws_summary[f'B{row}'] = lynch_text
                    ws_summary[f'B{row}'].font = data_font
                    ws_summary[f'B{row}'].border = border
                    ws_summary[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in lynch_status.upper() or lynch_delta > 20:
                        ws_summary[f'B{row}'].fill = strong_buy_fill
                    elif 'BUY' in lynch_status.upper() or lynch_delta > 5:
                        ws_summary[f'B{row}'].fill = buy_fill
                    elif 'HOLD' in lynch_status.upper() or abs(lynch_delta) <= 5:
                        ws_summary[f'B{row}'].fill = hold_fill
                    elif 'SELL' in lynch_status.upper() or lynch_delta < -5:
                        ws_summary[f'B{row}'].fill = sell_fill
                    elif 'STRONG SELL' in lynch_status.upper() or lynch_delta < -20:
                        ws_summary[f'B{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'B{row}'] = "Insufficient Data"
                    ws_summary[f'B{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'B{row}'].border = border
                    ws_summary[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'B{row}'].fill = no_data_fill
                
                # DCF Valuation (Column C)
                dcf_status = stock_row.get('dcf_valuation_status', 'N/A')
                dcf_delta = stock_row.get('dcf_delta_percentage', 0)
                
                if dcf_status != 'N/A':
                    dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
                    ws_summary[f'C{row}'] = dcf_text
                    ws_summary[f'C{row}'].font = data_font
                    ws_summary[f'C{row}'].border = border
                    ws_summary[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
                        ws_summary[f'C{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
                        ws_summary[f'C{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
                        ws_summary[f'C{row}'].fill = hold_fill
                    elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
                        ws_summary[f'C{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
                        ws_summary[f'C{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'C{row}'] = "Insufficient Data"
                    ws_summary[f'C{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'C{row}'].border = border
                    ws_summary[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'C{row}'].fill = no_data_fill
                
                # Munger Farm Valuation (Column D)
                munger_assessment = stock_row.get('munger_7pct_assessment', 'N/A')
                munger_delta = stock_row.get('munger_7pct_delta_percentage', 0)
                
                if munger_assessment != 'N/A':
                    munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
                    ws_summary[f'D{row}'] = munger_text
                    ws_summary[f'D{row}'].font = data_font
                    ws_summary[f'D{row}'].border = border
                    ws_summary[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
                        ws_summary[f'D{row}'].fill = strong_buy_fill
                    elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
                        ws_summary[f'D{row}'].fill = buy_fill
                    elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
                        ws_summary[f'D{row}'].fill = hold_fill
                    elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
                        ws_summary[f'D{row}'].fill = sell_fill
                    elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
                        ws_summary[f'D{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'D{row}'] = "Insufficient Data"
                    ws_summary[f'D{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'D{row}'].border = border
                    ws_summary[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'D{row}'].fill = no_data_fill
                
                # Enhanced DCF Valuation (Column E)
                enhanced_dcf_status = stock_row.get('enhanced_dcf_status', 'N/A')
                enhanced_dcf_delta = stock_row.get('enhanced_dcf_delta', 0)
                
                if enhanced_dcf_status != 'N/A':
                    enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
                    ws_summary[f'E{row}'] = enhanced_dcf_text
                    ws_summary[f'E{row}'].font = data_font
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
                        ws_summary[f'E{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
                        ws_summary[f'E{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
                        ws_summary[f'E{row}'].fill = hold_fill
                    elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
                        ws_summary[f'E{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
                        ws_summary[f'E{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'E{row}'] = "Insufficient Data"
                    ws_summary[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'E{row}'].fill = no_data_fill
                
                # Relative Valuation (Column F)
                relative_status = stock_row.get('relative_valuation_status', 'N/A')
                relative_delta = stock_row.get('relative_valuation_delta', 0)
                
                if relative_status != 'N/A':
                    relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
                    ws_summary[f'F{row}'] = relative_text
                    ws_summary[f'F{row}'].font = data_font
                    ws_summary[f'F{row}'].border = border
                    ws_summary[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
                        ws_summary[f'F{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
                        ws_summary[f'F{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
                        ws_summary[f'F{row}'].fill = hold_fill
                    elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
                        ws_summary[f'F{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
                        ws_summary[f'F{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'F{row}'] = "Insufficient Data"
                    ws_summary[f'F{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'F{row}'].border = border
                    ws_summary[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'F{row}'].fill = no_data_fill
                
                # Reverse DCF Valuation (Column G)
                reverse_dcf_assessment = stock_row.get('reverse_dcf_assessment', 'N/A')
                reverse_dcf_growth = stock_row.get('reverse_dcf_implied_growth', 0)
                
                if reverse_dcf_assessment != 'N/A':
                    reverse_dcf_text = f"{reverse_dcf_assessment}\nGrowth: {reverse_dcf_growth:+.1f}%"
                    ws_summary[f'G{row}'] = reverse_dcf_text
                    ws_summary[f'G{row}'].font = data_font
                    ws_summary[f'G{row}'].border = border
                    ws_summary[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'REASONABLE' in reverse_dcf_assessment.upper():
                        ws_summary[f'G{row}'].fill = hold_fill
                    elif 'UNREASONABLE' in reverse_dcf_assessment.upper():
                        ws_summary[f'G{row}'].fill = sell_fill
                    else:
                        ws_summary[f'G{row}'].fill = no_data_fill
                else:
                    ws_summary[f'G{row}'] = "Insufficient Data"
                    ws_summary[f'G{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'G{row}'].border = border
                    ws_summary[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'G{row}'].fill = no_data_fill
                
                # EPV/RIM Valuation (Column H)
                epv_assessment = stock_row.get('epv_assessment', 'N/A')
                epv_delta = stock_row.get('epv_delta', 0)
                rim_assessment = stock_row.get('rim_assessment', 'N/A')
                rim_delta = stock_row.get('rim_delta', 0)
                
                # Use EPV if available, otherwise RIM
                if epv_assessment != 'N/A':
                    epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
                    ws_summary[f'H{row}'] = epv_rim_text
                    ws_summary[f'H{row}'].font = data_font
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
                        ws_summary[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
                        ws_summary[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
                        ws_summary[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
                        ws_summary[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
                        ws_summary[f'H{row}'].fill = strong_sell_fill
                elif rim_assessment != 'N/A':
                    epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
                    ws_summary[f'H{row}'] = epv_rim_text
                    ws_summary[f'H{row}'].font = data_font
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
                        ws_summary[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
                        ws_summary[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
                        ws_summary[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
                        ws_summary[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
                        ws_summary[f'H{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'H{row}'] = "Insufficient Data"
                    ws_summary[f'H{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'H{row}'].fill = no_data_fill
                
                # Current Price (Column I)
                ws_summary[f'I{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_summary[f'I{row}'].font = data_font
                ws_summary[f'I{row}'].border = border
                ws_summary[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Add stocks with partial data (1-2 valuation methods)
            for stock_row in stocks_partial:
                ticker = stock_row['ticker']
                company_name = stock_row['company_name']
                current_price = stock_row['current_price']
                
                # Company name and ticker
                ws_summary[f'A{row}'] = f"{company_name} ({ticker})"
                ws_summary[f'A{row}'].font = data_font
                ws_summary[f'A{row}'].border = border
                ws_summary[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Current price
                ws_summary[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_summary[f'E{row}'].font = data_font
                ws_summary[f'E{row}'].border = border
                ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Peter Lynch Valuation (Column B)
                lynch_status = stock_row.get('lynch_valuation_status', 'N/A')
                lynch_delta = stock_row.get('lynch_delta_percentage', 0)
                
                if lynch_status != 'N/A':
                    lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
                    ws_summary[f'B{row}'] = lynch_text
                    ws_summary[f'B{row}'].font = data_font
                    ws_summary[f'B{row}'].border = border
                    ws_summary[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in lynch_status.upper() or lynch_delta > 20:
                        ws_summary[f'B{row}'].fill = strong_buy_fill
                    elif 'BUY' in lynch_status.upper() or lynch_delta > 5:
                        ws_summary[f'B{row}'].fill = buy_fill
                    elif 'HOLD' in lynch_status.upper() or abs(lynch_delta) <= 5:
                        ws_summary[f'B{row}'].fill = hold_fill
                    elif 'SELL' in lynch_status.upper() or lynch_delta < -5:
                        ws_summary[f'B{row}'].fill = sell_fill
                    elif 'STRONG SELL' in lynch_status.upper() or lynch_delta < -20:
                        ws_summary[f'B{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'B{row}'] = "Insufficient Data"
                    ws_summary[f'B{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'B{row}'].border = border
                    ws_summary[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'B{row}'].fill = no_data_fill
                
                # DCF Valuation (Column C)
                dcf_status = stock_row.get('dcf_valuation_status', 'N/A')
                dcf_delta = stock_row.get('dcf_delta_percentage', 0)
                
                if dcf_status != 'N/A':
                    dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
                    ws_summary[f'C{row}'] = dcf_text
                    ws_summary[f'C{row}'].font = data_font
                    ws_summary[f'C{row}'].border = border
                    ws_summary[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
                        ws_summary[f'C{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
                        ws_summary[f'C{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
                        ws_summary[f'C{row}'].fill = hold_fill
                    elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
                        ws_summary[f'C{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
                        ws_summary[f'C{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'C{row}'] = "Insufficient Data"
                    ws_summary[f'C{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'C{row}'].border = border
                    ws_summary[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'C{row}'].fill = no_data_fill
                
                # Munger Farm Valuation (Column D)
                munger_assessment = stock_row.get('munger_7pct_assessment', 'N/A')
                munger_delta = stock_row.get('munger_7pct_delta_percentage', 0)
                
                if munger_assessment != 'N/A':
                    munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
                    ws_summary[f'D{row}'] = munger_text
                    ws_summary[f'D{row}'].font = data_font
                    ws_summary[f'D{row}'].border = border
                    ws_summary[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
                        ws_summary[f'D{row}'].fill = strong_buy_fill
                    elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
                        ws_summary[f'D{row}'].fill = buy_fill
                    elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
                        ws_summary[f'D{row}'].fill = hold_fill
                    elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
                        ws_summary[f'D{row}'].fill = sell_fill
                    elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
                        ws_summary[f'D{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'D{row}'] = "Insufficient Data"
                    ws_summary[f'D{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'D{row}'].border = border
                    ws_summary[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'D{row}'].fill = no_data_fill
                
                # Enhanced DCF Valuation (Column E)
                enhanced_dcf_status = stock_row.get('enhanced_dcf_status', 'N/A')
                enhanced_dcf_delta = stock_row.get('enhanced_dcf_delta', 0)
                
                if enhanced_dcf_status != 'N/A':
                    enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
                    ws_summary[f'E{row}'] = enhanced_dcf_text
                    ws_summary[f'E{row}'].font = data_font
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
                        ws_summary[f'E{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
                        ws_summary[f'E{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
                        ws_summary[f'E{row}'].fill = hold_fill
                    elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
                        ws_summary[f'E{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
                        ws_summary[f'E{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'E{row}'] = "Insufficient Data"
                    ws_summary[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'E{row}'].fill = no_data_fill
                
                # Relative Valuation (Column F)
                relative_status = stock_row.get('relative_valuation_status', 'N/A')
                relative_delta = stock_row.get('relative_valuation_delta', 0)
                
                if relative_status != 'N/A':
                    relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
                    ws_summary[f'F{row}'] = relative_text
                    ws_summary[f'F{row}'].font = data_font
                    ws_summary[f'F{row}'].border = border
                    ws_summary[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
                        ws_summary[f'F{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
                        ws_summary[f'F{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
                        ws_summary[f'F{row}'].fill = hold_fill
                    elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
                        ws_summary[f'F{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
                        ws_summary[f'F{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'F{row}'] = "Insufficient Data"
                    ws_summary[f'F{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'F{row}'].border = border
                    ws_summary[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'F{row}'].fill = no_data_fill
                
                # Reverse DCF Valuation (Column G)
                reverse_dcf_assessment = stock_row.get('reverse_dcf_assessment', 'N/A')
                reverse_dcf_growth = stock_row.get('reverse_dcf_implied_growth', 0)
                
                if reverse_dcf_assessment != 'N/A':
                    reverse_dcf_text = f"{reverse_dcf_assessment}\nGrowth: {reverse_dcf_growth:+.1f}%"
                    ws_summary[f'G{row}'] = reverse_dcf_text
                    ws_summary[f'G{row}'].font = data_font
                    ws_summary[f'G{row}'].border = border
                    ws_summary[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'REASONABLE' in reverse_dcf_assessment.upper():
                        ws_summary[f'G{row}'].fill = hold_fill
                    elif 'UNREASONABLE' in reverse_dcf_assessment.upper():
                        ws_summary[f'G{row}'].fill = sell_fill
                    else:
                        ws_summary[f'G{row}'].fill = no_data_fill
                else:
                    ws_summary[f'G{row}'] = "Insufficient Data"
                    ws_summary[f'G{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'G{row}'].border = border
                    ws_summary[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'G{row}'].fill = no_data_fill
                
                # EPV/RIM Valuation (Column H)
                epv_assessment = stock_row.get('epv_assessment', 'N/A')
                epv_delta = stock_row.get('epv_delta', 0)
                rim_assessment = stock_row.get('rim_assessment', 'N/A')
                rim_delta = stock_row.get('rim_delta', 0)
                
                # Use EPV if available, otherwise RIM
                if epv_assessment != 'N/A':
                    epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
                    ws_summary[f'H{row}'] = epv_rim_text
                    ws_summary[f'H{row}'].font = data_font
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
                        ws_summary[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
                        ws_summary[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
                        ws_summary[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
                        ws_summary[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
                        ws_summary[f'H{row}'].fill = strong_sell_fill
                elif rim_assessment != 'N/A':
                    epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
                    ws_summary[f'H{row}'] = epv_rim_text
                    ws_summary[f'H{row}'].font = data_font
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
                        ws_summary[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
                        ws_summary[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
                        ws_summary[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
                        ws_summary[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
                        ws_summary[f'H{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'H{row}'] = "Insufficient Data"
                    ws_summary[f'H{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'H{row}'].fill = no_data_fill
                
                # Current Price (Column I)
                ws_summary[f'I{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_summary[f'I{row}'].font = data_font
                ws_summary[f'I{row}'].border = border
                ws_summary[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Add stocks with insufficient data at the bottom
            if stocks_without_data:
                row += 2
                ws_summary[f'A{row}'] = "STOCKS WITH INSUFFICIENT DATA"
                ws_summary[f'A{row}'].font = Font(bold=True, size=12, color="808080")
                ws_summary[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                ws_summary[f'A{row}'].border = border
                ws_summary[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for stock_row in stocks_without_data:
                    ticker = stock_row['ticker']
                    company_name = stock_row['company_name']
                    current_price = stock_row['current_price']
                    
                    # Company name and ticker
                    ws_summary[f'A{row}'] = f"{company_name} ({ticker})"
                    ws_summary[f'A{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'A{row}'].border = border
                    ws_summary[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Current price
                    ws_summary[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                    ws_summary[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # All valuation columns show "Insufficient Data"
                    for col in ['B', 'C', 'D']:
                        ws_summary[f'{col}{row}'] = "Insufficient Data"
                        ws_summary[f'{col}{row}'].font = Font(size=11, color="808080")  # Grey text
                        ws_summary[f'{col}{row}'].border = border
                        ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                        ws_summary[f'{col}{row}'].fill = no_data_fill
                    
                    row += 1
            
            # Add legend
            row += 1
            ws_summary[f'A{row}'] = "Legend:"
            ws_summary[f'A{row}'].font = Font(bold=True, size=12)
            ws_summary[f'A{row}'].border = border
            row += 1
            
            legend_items = [
                ("Strong Buy", strong_buy_fill),
                ("Buy", buy_fill),
                ("Hold", hold_fill),
                ("Sell", sell_fill),
                ("Strong Sell", strong_sell_fill),
                ("Insufficient Data", no_data_fill)
            ]
            
            for item_text, item_fill in legend_items:
                ws_summary[f'A{row}'] = item_text
                ws_summary[f'A{row}'].font = data_font
                ws_summary[f'A{row}'].fill = item_fill
                ws_summary[f'A{row}'].border = border
                ws_summary[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                row += 1
            
            # Set optimal column widths for readability
            ws_summary.column_dimensions['A'].width = 45  # Company name and ticker
            ws_summary.column_dimensions['B'].width = 35  # Peter Lynch valuation
            ws_summary.column_dimensions['C'].width = 35  # DCF valuation
            ws_summary.column_dimensions['D'].width = 35  # Munger Farm valuation
            ws_summary.column_dimensions['E'].width = 35  # Enhanced DCF valuation
            ws_summary.column_dimensions['F'].width = 35  # Relative Valuation
            ws_summary.column_dimensions['G'].width = 35  # Reverse DCF valuation
            ws_summary.column_dimensions['H'].width = 35  # EPV/RIM valuation
            ws_summary.column_dimensions['I'].width = 20  # Current price
            
            self.logger.info("Created valuation summary sheet with conditional formatting")
            
        except Exception as e:
            self.logger.error(f"Error creating valuation summary sheet: {e}")

    def create_prospects_sheet(self, df, wb):
        """Create prospects sheet ranked by undervaluation"""
        try:
            ws_prospects = wb.create_sheet('Prospects', 1)
            
            # Define consistent color scheme across all sheets
            header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=14)
            subheader_font = Font(bold=True, size=12)
            data_font = Font(size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Consistent conditional formatting fills
            strong_buy_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            buy_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
            hold_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            sell_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            strong_sell_fill = PatternFill(start_color="FFA0A0", end_color="FFA0A0", fill_type="solid")
            no_data_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            # Title
            ws_prospects['A1'] = "NZX PROSPECTS - RANKED BY UNDERVALUATION"
            ws_prospects['A1'].font = header_font
            ws_prospects['A1'].fill = header_fill
            ws_prospects['A1'].border = border
            ws_prospects['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_prospects.merge_cells('A1:I1')
            
            # Headers for all 9 valuation methods
            headers = ['Company (Ticker)', 'Peter Lynch', 'DCF Valuation', 'Munger Farm', 'Enhanced DCF', 'Relative Valuation', 'Reverse DCF', 'EPV/RIM', 'Current Price']
            row = 3
            for col, header in enumerate(headers, 1):
                cell = ws_prospects.cell(row=row, column=col)
                cell.value = header
                cell.font = subheader_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Calculate undervaluation scores for ranking
            def calculate_undervaluation_score(stock_row):
                """Calculate composite undervaluation score"""
                score = 0
                count = 0
                
                # Peter Lynch score
                lynch_delta = stock_row.get('lynch_delta_percentage', 0)
                if lynch_delta != 0:
                    score += lynch_delta
                    count += 1
                
                # DCF score
                dcf_delta = stock_row.get('dcf_delta_percentage', 0)
                if dcf_delta != 0:
                    score += dcf_delta
                    count += 1
                
                # Munger score
                munger_delta = stock_row.get('munger_7pct_delta_percentage', 0)
                if munger_delta != 0:
                    score += munger_delta
                    count += 1
                
                return score / count if count > 0 else -999  # -999 for insufficient data
            
            # Separate stocks by completeness and rank by undervaluation
            stocks_complete = []
            stocks_partial = []
            stocks_without_data = []
            
            for _, stock_row in df.iterrows():
                lynch_status = stock_row.get('lynch_valuation_status', 'N/A')
                dcf_status = stock_row.get('dcf_valuation_status', 'N/A')
                munger_assessment = stock_row.get('munger_7pct_assessment', 'N/A')
                
                # Count how many valuation methods are available
                available_methods = 0
                if lynch_status != 'N/A':
                    available_methods += 1
                if dcf_status != 'N/A':
                    available_methods += 1
                if munger_assessment != 'N/A':
                    available_methods += 1
                
                # Calculate undervaluation score
                undervaluation_score = calculate_undervaluation_score(stock_row)
                stock_row['undervaluation_score'] = undervaluation_score
                
                if available_methods == 3:
                    stocks_complete.append(stock_row)
                elif available_methods > 0:
                    stocks_partial.append(stock_row)
                else:
                    stocks_without_data.append(stock_row)
            
            # Sort by undervaluation score (highest first)
            stocks_complete.sort(key=lambda x: x['undervaluation_score'], reverse=True)
            stocks_partial.sort(key=lambda x: x['undervaluation_score'], reverse=True)
            
            # Add stocks with complete data first (ranked by undervaluation)
            for stock_row in stocks_complete:
                ticker = stock_row['ticker']
                company_name = stock_row['company_name']
                current_price = stock_row['current_price']
                
                # Company name and ticker
                ws_prospects[f'A{row}'] = f"{company_name} ({ticker})"
                ws_prospects[f'A{row}'].font = data_font
                ws_prospects[f'A{row}'].border = border
                ws_prospects[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Current price
                ws_prospects[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_prospects[f'E{row}'].font = data_font
                ws_prospects[f'E{row}'].border = border
                ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Peter Lynch Valuation (Column B)
                lynch_status = stock_row.get('lynch_valuation_status', 'N/A')
                lynch_delta = stock_row.get('lynch_delta_percentage', 0)
                
                if lynch_status != 'N/A':
                    lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
                    ws_prospects[f'B{row}'] = lynch_text
                    ws_prospects[f'B{row}'].font = data_font
                    ws_prospects[f'B{row}'].border = border
                    ws_prospects[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in lynch_status.upper() or lynch_delta > 20:
                        ws_prospects[f'B{row}'].fill = strong_buy_fill
                    elif 'BUY' in lynch_status.upper() or lynch_delta > 5:
                        ws_prospects[f'B{row}'].fill = buy_fill
                    elif 'HOLD' in lynch_status.upper() or abs(lynch_delta) <= 5:
                        ws_prospects[f'B{row}'].fill = hold_fill
                    elif 'SELL' in lynch_status.upper() or lynch_delta < -5:
                        ws_prospects[f'B{row}'].fill = sell_fill
                    elif 'STRONG SELL' in lynch_status.upper() or lynch_delta < -20:
                        ws_prospects[f'B{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'B{row}'] = "Insufficient Data"
                    ws_prospects[f'B{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'B{row}'].border = border
                    ws_prospects[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'B{row}'].fill = no_data_fill
                
                # DCF Valuation (Column C)
                dcf_status = stock_row.get('dcf_valuation_status', 'N/A')
                dcf_delta = stock_row.get('dcf_delta_percentage', 0)
                
                if dcf_status != 'N/A':
                    dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
                    ws_prospects[f'C{row}'] = dcf_text
                    ws_prospects[f'C{row}'].font = data_font
                    ws_prospects[f'C{row}'].border = border
                    ws_prospects[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
                        ws_prospects[f'C{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
                        ws_prospects[f'C{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
                        ws_prospects[f'C{row}'].fill = hold_fill
                    elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
                        ws_prospects[f'C{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
                        ws_prospects[f'C{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'C{row}'] = "Insufficient Data"
                    ws_prospects[f'C{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'C{row}'].border = border
                    ws_prospects[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'C{row}'].fill = no_data_fill
                
                # Munger Farm Valuation (Column D)
                munger_assessment = stock_row.get('munger_7pct_assessment', 'N/A')
                munger_delta = stock_row.get('munger_7pct_delta_percentage', 0)
                
                if munger_assessment != 'N/A':
                    munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
                    ws_prospects[f'D{row}'] = munger_text
                    ws_prospects[f'D{row}'].font = data_font
                    ws_prospects[f'D{row}'].border = border
                    ws_prospects[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
                        ws_prospects[f'D{row}'].fill = strong_buy_fill
                    elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
                        ws_prospects[f'D{row}'].fill = buy_fill
                    elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
                        ws_prospects[f'D{row}'].fill = hold_fill
                    elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
                        ws_prospects[f'D{row}'].fill = sell_fill
                    elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
                        ws_prospects[f'D{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'D{row}'] = "Insufficient Data"
                    ws_prospects[f'D{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'D{row}'].border = border
                    ws_prospects[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'D{row}'].fill = no_data_fill
                
                # Enhanced DCF Valuation (Column E)
                enhanced_dcf_status = stock_row.get('enhanced_dcf_status', 'N/A')
                enhanced_dcf_delta = stock_row.get('enhanced_dcf_delta', 0)
                
                if enhanced_dcf_status != 'N/A':
                    enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
                    ws_prospects[f'E{row}'] = enhanced_dcf_text
                    ws_prospects[f'E{row}'].font = data_font
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
                        ws_prospects[f'E{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
                        ws_prospects[f'E{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
                        ws_prospects[f'E{row}'].fill = hold_fill
                    elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
                        ws_prospects[f'E{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
                        ws_prospects[f'E{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'E{row}'] = "Insufficient Data"
                    ws_prospects[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'E{row}'].fill = no_data_fill
                
                # Relative Valuation (Column F)
                relative_status = stock_row.get('relative_valuation_status', 'N/A')
                relative_delta = stock_row.get('relative_valuation_delta', 0)
                
                if relative_status != 'N/A':
                    relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
                    ws_prospects[f'F{row}'] = relative_text
                    ws_prospects[f'F{row}'].font = data_font
                    ws_prospects[f'F{row}'].border = border
                    ws_prospects[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
                        ws_prospects[f'F{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
                        ws_prospects[f'F{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
                        ws_prospects[f'F{row}'].fill = hold_fill
                    elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
                        ws_prospects[f'F{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
                        ws_prospects[f'F{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'F{row}'] = "Insufficient Data"
                    ws_prospects[f'F{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'F{row}'].border = border
                    ws_prospects[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'F{row}'].fill = no_data_fill
                
                # Reverse DCF Valuation (Column G)
                reverse_dcf_assessment = stock_row.get('reverse_dcf_assessment', 'N/A')
                reverse_dcf_growth = stock_row.get('reverse_dcf_implied_growth', 0)
                
                if reverse_dcf_assessment != 'N/A':
                    reverse_dcf_text = f"{reverse_dcf_assessment}\nGrowth: {reverse_dcf_growth:+.1f}%"
                    ws_prospects[f'G{row}'] = reverse_dcf_text
                    ws_prospects[f'G{row}'].font = data_font
                    ws_prospects[f'G{row}'].border = border
                    ws_prospects[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'REASONABLE' in reverse_dcf_assessment.upper():
                        ws_prospects[f'G{row}'].fill = hold_fill
                    elif 'UNREASONABLE' in reverse_dcf_assessment.upper():
                        ws_prospects[f'G{row}'].fill = sell_fill
                    else:
                        ws_prospects[f'G{row}'].fill = no_data_fill
                else:
                    ws_prospects[f'G{row}'] = "Insufficient Data"
                    ws_prospects[f'G{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'G{row}'].border = border
                    ws_prospects[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'G{row}'].fill = no_data_fill
                
                # EPV/RIM Valuation (Column H)
                epv_assessment = stock_row.get('epv_assessment', 'N/A')
                epv_delta = stock_row.get('epv_delta', 0)
                rim_assessment = stock_row.get('rim_assessment', 'N/A')
                rim_delta = stock_row.get('rim_delta', 0)
                
                # Use EPV if available, otherwise RIM
                if epv_assessment != 'N/A':
                    epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
                    ws_prospects[f'H{row}'] = epv_rim_text
                    ws_prospects[f'H{row}'].font = data_font
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
                        ws_prospects[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
                        ws_prospects[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
                        ws_prospects[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
                        ws_prospects[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
                        ws_prospects[f'H{row}'].fill = strong_sell_fill
                elif rim_assessment != 'N/A':
                    epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
                    ws_prospects[f'H{row}'] = epv_rim_text
                    ws_prospects[f'H{row}'].font = data_font
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
                        ws_prospects[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
                        ws_prospects[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
                        ws_prospects[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
                        ws_prospects[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
                        ws_prospects[f'H{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'H{row}'] = "Insufficient Data"
                    ws_prospects[f'H{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'H{row}'].fill = no_data_fill
                
                # Current Price (Column I)
                ws_prospects[f'I{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_prospects[f'I{row}'].font = data_font
                ws_prospects[f'I{row}'].border = border
                ws_prospects[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Add stocks with partial data (ranked by undervaluation)
            for stock_row in stocks_partial:
                ticker = stock_row['ticker']
                company_name = stock_row['company_name']
                current_price = stock_row['current_price']
                
                # Company name and ticker
                ws_prospects[f'A{row}'] = f"{company_name} ({ticker})"
                ws_prospects[f'A{row}'].font = data_font
                ws_prospects[f'A{row}'].border = border
                ws_prospects[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Current price
                ws_prospects[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_prospects[f'E{row}'].font = data_font
                ws_prospects[f'E{row}'].border = border
                ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Peter Lynch Valuation (Column B)
                lynch_status = stock_row.get('lynch_valuation_status', 'N/A')
                lynch_delta = stock_row.get('lynch_delta_percentage', 0)
                
                if lynch_status != 'N/A':
                    lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
                    ws_prospects[f'B{row}'] = lynch_text
                    ws_prospects[f'B{row}'].font = data_font
                    ws_prospects[f'B{row}'].border = border
                    ws_prospects[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in lynch_status.upper() or lynch_delta > 20:
                        ws_prospects[f'B{row}'].fill = strong_buy_fill
                    elif 'BUY' in lynch_status.upper() or lynch_delta > 5:
                        ws_prospects[f'B{row}'].fill = buy_fill
                    elif 'HOLD' in lynch_status.upper() or abs(lynch_delta) <= 5:
                        ws_prospects[f'B{row}'].fill = hold_fill
                    elif 'SELL' in lynch_status.upper() or lynch_delta < -5:
                        ws_prospects[f'B{row}'].fill = sell_fill
                    elif 'STRONG SELL' in lynch_status.upper() or lynch_delta < -20:
                        ws_prospects[f'B{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'B{row}'] = "Insufficient Data"
                    ws_prospects[f'B{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'B{row}'].border = border
                    ws_prospects[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'B{row}'].fill = no_data_fill
                
                # DCF Valuation (Column C)
                dcf_status = stock_row.get('dcf_valuation_status', 'N/A')
                dcf_delta = stock_row.get('dcf_delta_percentage', 0)
                
                if dcf_status != 'N/A':
                    dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
                    ws_prospects[f'C{row}'] = dcf_text
                    ws_prospects[f'C{row}'].font = data_font
                    ws_prospects[f'C{row}'].border = border
                    ws_prospects[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
                        ws_prospects[f'C{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
                        ws_prospects[f'C{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
                        ws_prospects[f'C{row}'].fill = hold_fill
                    elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
                        ws_prospects[f'C{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
                        ws_prospects[f'C{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'C{row}'] = "Insufficient Data"
                    ws_prospects[f'C{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'C{row}'].border = border
                    ws_prospects[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'C{row}'].fill = no_data_fill
                
                # Munger Farm Valuation (Column D)
                munger_assessment = stock_row.get('munger_7pct_assessment', 'N/A')
                munger_delta = stock_row.get('munger_7pct_delta_percentage', 0)
                
                if munger_assessment != 'N/A':
                    munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
                    ws_prospects[f'D{row}'] = munger_text
                    ws_prospects[f'D{row}'].font = data_font
                    ws_prospects[f'D{row}'].border = border
                    ws_prospects[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
                        ws_prospects[f'D{row}'].fill = strong_buy_fill
                    elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
                        ws_prospects[f'D{row}'].fill = buy_fill
                    elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
                        ws_prospects[f'D{row}'].fill = hold_fill
                    elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
                        ws_prospects[f'D{row}'].fill = sell_fill
                    elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
                        ws_prospects[f'D{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'D{row}'] = "Insufficient Data"
                    ws_prospects[f'D{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'D{row}'].border = border
                    ws_prospects[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'D{row}'].fill = no_data_fill
                
                # Enhanced DCF Valuation (Column E)
                enhanced_dcf_status = stock_row.get('enhanced_dcf_status', 'N/A')
                enhanced_dcf_delta = stock_row.get('enhanced_dcf_delta', 0)
                
                if enhanced_dcf_status != 'N/A':
                    enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
                    ws_prospects[f'E{row}'] = enhanced_dcf_text
                    ws_prospects[f'E{row}'].font = data_font
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
                        ws_prospects[f'E{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
                        ws_prospects[f'E{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
                        ws_prospects[f'E{row}'].fill = hold_fill
                    elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
                        ws_prospects[f'E{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
                        ws_prospects[f'E{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'E{row}'] = "Insufficient Data"
                    ws_prospects[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'E{row}'].fill = no_data_fill
                
                # Relative Valuation (Column F)
                relative_status = stock_row.get('relative_valuation_status', 'N/A')
                relative_delta = stock_row.get('relative_valuation_delta', 0)
                
                if relative_status != 'N/A':
                    relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
                    ws_prospects[f'F{row}'] = relative_text
                    ws_prospects[f'F{row}'].font = data_font
                    ws_prospects[f'F{row}'].border = border
                    ws_prospects[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
                        ws_prospects[f'F{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
                        ws_prospects[f'F{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
                        ws_prospects[f'F{row}'].fill = hold_fill
                    elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
                        ws_prospects[f'F{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
                        ws_prospects[f'F{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'F{row}'] = "Insufficient Data"
                    ws_prospects[f'F{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'F{row}'].border = border
                    ws_prospects[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'F{row}'].fill = no_data_fill
                
                # Reverse DCF Valuation (Column G)
                reverse_dcf_assessment = stock_row.get('reverse_dcf_assessment', 'N/A')
                reverse_dcf_growth = stock_row.get('reverse_dcf_implied_growth', 0)
                
                if reverse_dcf_assessment != 'N/A':
                    reverse_dcf_text = f"{reverse_dcf_assessment}\nGrowth: {reverse_dcf_growth:+.1f}%"
                    ws_prospects[f'G{row}'] = reverse_dcf_text
                    ws_prospects[f'G{row}'].font = data_font
                    ws_prospects[f'G{row}'].border = border
                    ws_prospects[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'REASONABLE' in reverse_dcf_assessment.upper():
                        ws_prospects[f'G{row}'].fill = hold_fill
                    elif 'UNREASONABLE' in reverse_dcf_assessment.upper():
                        ws_prospects[f'G{row}'].fill = sell_fill
                    else:
                        ws_prospects[f'G{row}'].fill = no_data_fill
                else:
                    ws_prospects[f'G{row}'] = "Insufficient Data"
                    ws_prospects[f'G{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'G{row}'].border = border
                    ws_prospects[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'G{row}'].fill = no_data_fill
                
                # EPV/RIM Valuation (Column H)
                epv_assessment = stock_row.get('epv_assessment', 'N/A')
                epv_delta = stock_row.get('epv_delta', 0)
                rim_assessment = stock_row.get('rim_assessment', 'N/A')
                rim_delta = stock_row.get('rim_delta', 0)
                
                # Use EPV if available, otherwise RIM
                if epv_assessment != 'N/A':
                    epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
                    ws_prospects[f'H{row}'] = epv_rim_text
                    ws_prospects[f'H{row}'].font = data_font
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
                        ws_prospects[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
                        ws_prospects[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
                        ws_prospects[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
                        ws_prospects[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
                        ws_prospects[f'H{row}'].fill = strong_sell_fill
                elif rim_assessment != 'N/A':
                    epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
                    ws_prospects[f'H{row}'] = epv_rim_text
                    ws_prospects[f'H{row}'].font = data_font
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
                        ws_prospects[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
                        ws_prospects[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
                        ws_prospects[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
                        ws_prospects[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
                        ws_prospects[f'H{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'H{row}'] = "Insufficient Data"
                    ws_prospects[f'H{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'H{row}'].fill = no_data_fill
                
                # Current Price (Column I)
                ws_prospects[f'I{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_prospects[f'I{row}'].font = data_font
                ws_prospects[f'I{row}'].border = border
                ws_prospects[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Add stocks with insufficient data at the bottom
            if stocks_without_data:
                row += 2
                ws_prospects[f'A{row}'] = "STOCKS WITH INSUFFICIENT DATA"
                ws_prospects[f'A{row}'].font = Font(bold=True, size=12, color="808080")
                ws_prospects[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                ws_prospects[f'A{row}'].border = border
                ws_prospects[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for stock_row in stocks_without_data:
                    ticker = stock_row['ticker']
                    company_name = stock_row['company_name']
                    current_price = stock_row['current_price']
                    
                    # Company name and ticker
                    ws_prospects[f'A{row}'] = f"{company_name} ({ticker})"
                    ws_prospects[f'A{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'A{row}'].border = border
                    ws_prospects[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Current price
                    ws_prospects[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                    ws_prospects[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # All valuation columns show "Insufficient Data"
                    for col in ['B', 'C', 'D']:
                        ws_prospects[f'{col}{row}'] = "Insufficient Data"
                        ws_prospects[f'{col}{row}'].font = Font(size=11, color="808080")  # Grey text
                        ws_prospects[f'{col}{row}'].border = border
                        ws_prospects[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                        ws_prospects[f'{col}{row}'].fill = no_data_fill
                    
                    row += 1
            
            # Add legend
            row += 1
            ws_prospects[f'A{row}'] = "Legend:"
            ws_prospects[f'A{row}'].font = Font(bold=True, size=12)
            ws_prospects[f'A{row}'].border = border
            row += 1
            
            legend_items = [
                ("Strong Buy", strong_buy_fill),
                ("Buy", buy_fill),
                ("Hold", hold_fill),
                ("Sell", sell_fill),
                ("Strong Sell", strong_sell_fill),
                ("Insufficient Data", no_data_fill)
            ]
            
            for i, (label, fill) in enumerate(legend_items):
                col = i + 1
                ws_prospects.cell(row=row, column=col, value=label)
                ws_prospects.cell(row=row, column=col).font = Font(size=10)
                ws_prospects.cell(row=row, column=col).fill = fill
                ws_prospects.cell(row=row, column=col).border = border
                ws_prospects.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            # Set optimal column widths for readability
            ws_prospects.column_dimensions['A'].width = 45  # Company name and ticker
            ws_prospects.column_dimensions['B'].width = 35  # Peter Lynch valuation
            ws_prospects.column_dimensions['C'].width = 35  # DCF valuation
            ws_prospects.column_dimensions['D'].width = 35  # Munger Farm valuation
            ws_prospects.column_dimensions['E'].width = 35  # Enhanced DCF valuation
            ws_prospects.column_dimensions['F'].width = 35  # Relative Valuation
            ws_prospects.column_dimensions['G'].width = 35  # Reverse DCF valuation
            ws_prospects.column_dimensions['H'].width = 35  # EPV/RIM valuation
            ws_prospects.column_dimensions['I'].width = 20  # Current price
            
            self.logger.info("Created prospects sheet with undervaluation ranking")
            
        except Exception as e:
            self.logger.error(f"Error creating prospects sheet: {e}")

def main():
    """Main function"""
    scraper = StockValuationScraper()
    scraper.run()

if __name__ == "__main__":
    main()
