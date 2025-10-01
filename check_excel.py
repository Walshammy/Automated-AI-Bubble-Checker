import pandas as pd
from openpyxl import load_workbook
import os

def check_excel():
    """Check if all valuation methods are properly populated in Excel sheets"""
    
    excel_file = r"C:\Users\james\Downloads\Stock Valuation\stock_valuation_dataset.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"[ERROR] Excel file not found: {excel_file}")
        return
    
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        print(f"[SUCCESS] Successfully loaded Excel file: {excel_file}")
        
        # Check available sheets
        print(f"\n[INFO] Available sheets: {wb.sheetnames}")
        
        # Check each sheet
        for sheet_name in wb.sheetnames:
            print(f"\n[ANALYZING] Sheet: '{sheet_name}'")
            ws = wb[sheet_name]
            
            # Get the dimensions
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"   Dimensions: {max_row} rows x {max_col} columns")
            
            # Get headers - for My Portfolio sheet, headers are in row 5, for others in row 1
            if sheet_name == 'My Portfolio':
                header_row = 5
            else:
                header_row = 1
            headers = [cell.value for cell in ws[header_row]]
            print(f"   Headers (row {header_row}): {headers}")
            
            # Check for data in key columns
            if sheet_name in ['My Portfolio', 'Prospects']:
                print(f"   [CHECKING] Valuation data population...")
                
                # Check first few rows of data - start after headers
                data_start_row = header_row + 1
                for row in range(data_start_row, min(data_start_row + 4, max_row + 1)):  # Check first 4 data rows
                    row_data = []
                    for col in range(1, max_col + 1):
                        cell_value = ws.cell(row=row, column=col).value
                        row_data.append(str(cell_value))
                    print(f"   Row {row}: {row_data}")
                
                # Define expected valuation methods
                valuation_methods = ["Peter Lynch", "DCF Valuation", "Munger Farm", 
                                     "Enhanced DCF", "Relative Valuation", "Reverse DCF", "EPV/RIM"]
                
                found_methods = []
                for method in valuation_methods:
                    for header in headers:
                        if header and method.lower() in header.lower():
                            found_methods.append(method)
                            break
                
                print(f"   [FOUND] Valuation methods: {found_methods}")
                print(f"   [MISSING] Methods: {set(valuation_methods) - set(found_methods)}")
        
        # Also check using pandas for the Valuation Data sheet
        if 'Valuation Data' in wb.sheetnames:
            print(f"\n[PANDAS] Analyzing 'Valuation Data' sheet...")
            try:
                df = pd.read_excel(excel_file, sheet_name='Valuation Data')
                print(f"   Shape: {df.shape}")
                print(f"   Columns: {list(df.columns)}")
                
                new_methods = [
                    'enhanced_dcf_intrinsic_value', 'enhanced_dcf_status', 'enhanced_dcf_delta',
                    'relative_valuation_status', 'relative_valuation_delta',
                    'reverse_dcf_implied_growth', 'reverse_dcf_assessment',
                    'epv_intrinsic_value', 'epv_assessment', 'epv_delta',
                    'rim_intrinsic_value', 'rim_assessment', 'rim_delta'
                ]
                
                found_new_methods = []
                missing_new_methods = []
                for method in new_methods:
                    if method in df.columns:
                        found_new_methods.append(method)
                        # Check if column has data
                        non_null_count = df[method].notna().sum()
                        print(f"   [FOUND] {method}: {non_null_count} non-null values")
                    else:
                        missing_new_methods.append(method)
                
                print(f"\n   [FOUND] New methods: {len(found_new_methods)}")
                print(f"   [MISSING] New methods: {len(missing_new_methods)}")
                
                if missing_new_methods:
                    print(f"   Missing: {missing_new_methods}")
                
            except Exception as e:
                print(f"   [ERROR] Error reading with pandas: {e}")
        
        wb.close()
        
    except Exception as e:
        print(f"[ERROR] Error analyzing Excel file: {e}")

if __name__ == "__main__":
    check_excel()
