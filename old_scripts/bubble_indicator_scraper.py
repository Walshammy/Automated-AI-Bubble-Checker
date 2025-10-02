import pandas as pd
import numpy as np
import yfinance as yf
import requests
from bs4 import BeautifulSoup
import time
import os
from datetime import datetime, timedelta
import logging
import json
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings

warnings.filterwarnings('ignore')

class BubbleIndicatorScraper:
    def __init__(self):
        # Setup logging
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger(__name__)

        # OneDrive directory (primary storage location)
        self.onedrive_dir = r"C:\Users\james\OneDrive - Silverdale Medical Limited\AIbubble"
        os.makedirs(self.onedrive_dir, exist_ok=True)

        # Second storage location - Downloads/AI Bubble
        self.downloads_dir = self._get_downloads_directory()
        if self.downloads_dir:
            os.makedirs(self.downloads_dir, exist_ok=True)

        # Daily backups directory
        self.daily_backups_dir = os.path.join(self.onedrive_dir, "daily_backups")
        os.makedirs(self.daily_backups_dir, exist_ok=True)

        # Master dataset files
        self.master_file = os.path.join(self.onedrive_dir, "bubble_indicators_dataset.xlsx")
        self.downloads_file = os.path.join(self.downloads_dir, "bubble_indicators_dataset.xlsx") if self.downloads_dir else None
        
        # Historical dataset file
        self.historical_file = os.path.join(self.onedrive_dir, "bubble_indicators_historical.xlsx")
        
        # Combined dataset file (historical + current)
        self.combined_file = os.path.join(self.onedrive_dir, "bubble_indicators_combined.xlsx")
        
        # Historical progress tracking
        self.historical_progress_file = os.path.join(self.onedrive_dir, "historical_progress.json")

        # Key stock tickers for tracking
        self.ai_stocks = {
            'NVDA': 'NVIDIA',
            'MSFT': 'Microsoft',
            'GOOGL': 'Alphabet',
            'AAPL': 'Apple',
            'AMZN': 'Amazon',
            'META': 'Meta',
            'TSLA': 'Tesla'
        }

        # Market indices
        self.indices = {
            '^GSPC': 'S&P 500',
            '^IXIC': 'NASDAQ',
            '^VIX': 'VIX',
            '^TNX': '10-Year Treasury'
        }

    def _get_downloads_directory(self):
        """Get the Downloads/AI Bubble directory with fallback to james.walsham search"""
        # Primary path
        primary_path = r"C:\Users\james\Downloads\AI Bubble"
        if os.path.exists(os.path.dirname(primary_path)):
            return primary_path
        
        # Fallback: search for james.walsham in Users directory
        users_dir = r"C:\Users"
        if os.path.exists(users_dir):
            for item in os.listdir(users_dir):
                if "james.walsham" in item.lower():
                    fallback_path = os.path.join(users_dir, item, "Downloads", "AI Bubble")
                    if os.path.exists(os.path.dirname(fallback_path)):
                        return fallback_path
        
        # If neither found, return None
        self.logger.warning("Could not find Downloads directory for james.walsham")
        return None

    def clean_and_format_data(self, df):
        """Clean and format data for proper Excel number formatting"""
        try:
            # Create a copy to avoid modifying original
            cleaned_df = df.copy()
            
            # Clean numeric columns
            numeric_columns = ['vix_level', 'sp500_price', 'concentration_ratio', 'ten_year_treasury', 
                             'fed_funds_rate_approx', 'bubble_risk_score', 'total_ai_market_cap', 
                             'nvidia_dominance_ratio', 'top_10_market_cap', 'sp500_total_market_cap']
            
            # Clean text columns that might contain JSON
            text_columns = ['company_breakdown']
            
            for col in numeric_columns:
                if col in cleaned_df.columns:
                    cleaned_df[col] = cleaned_df[col].apply(lambda x: self.clean_number(x))
            
            # Clean text columns
            for col in text_columns:
                if col in cleaned_df.columns:
                    cleaned_df[col] = cleaned_df[col].apply(lambda x: str(x) if x is not None else '')
            
            # Clean AI stock price columns
            for ticker in self.ai_stocks.keys():
                price_col = f"{self.ai_stocks[ticker].lower().replace(' ', '_')}_price"
                market_cap_col = f"{self.ai_stocks[ticker].lower().replace(' ', '_')}_market_cap"
                pe_col = f"{self.ai_stocks[ticker].lower().replace(' ', '_')}_pe"
                
                if price_col in cleaned_df.columns:
                    cleaned_df[price_col] = cleaned_df[price_col].apply(lambda x: self.clean_number(x))
                if market_cap_col in cleaned_df.columns:
                    cleaned_df[market_cap_col] = cleaned_df[market_cap_col].apply(lambda x: self.clean_number(x))
                if pe_col in cleaned_df.columns:
                    cleaned_df[pe_col] = cleaned_df[pe_col].apply(lambda x: self.clean_number(x))
            
            # Clean index price columns
            for ticker, name in self.indices.items():
                price_col = f"{name.lower().replace(' ', '_').replace('-', '_')}_price"
                pe_col = f"{name.lower().replace(' ', '_').replace('-', '_')}_pe"
                
                if price_col in cleaned_df.columns:
                    cleaned_df[price_col] = cleaned_df[price_col].apply(lambda x: self.clean_number(x))
                if pe_col in cleaned_df.columns:
                    cleaned_df[pe_col] = cleaned_df[pe_col].apply(lambda x: self.clean_number(x))
            
            self.logger.info("Data cleaned and formatted for Excel")
            return cleaned_df
            
        except Exception as e:
            self.logger.error(f"Error cleaning data: {e}")
            return df

    def clean_number(self, value):
        """Clean a number value, return number if valid, otherwise return original"""
        if pd.isna(value) or value == 'N/A' or value == '' or value is None:
            return value
        
        try:
            # Convert to float if possible
            if isinstance(value, (int, float)):
                return float(value)
            
            # Remove any non-digit characters except decimal point and minus sign
            cleaned = str(value).replace(',', '').replace(' ', '')
            # Extract just the number part
            number_match = re.search(r'(-?\d+(?:\.\d+)?)', cleaned)
            if number_match:
                return float(number_match.group(1))
        except:
            pass
        
        return value

    def get_historical_data(self, ticker, period="10y"):
        """Get historical data for a ticker over specified period with API rate limiting"""
        try:
            import time
            # Small delay to respect API limits
            time.sleep(0.1)  # 100ms delay between requests
            
            stock = yf.Ticker(ticker)
            hist = stock.history(period=period)
            return hist
        except Exception as e:
            self.logger.error(f"Error getting historical data for {ticker}: {e}")
            return None

    def load_historical_progress(self):
        """Load historical data collection progress"""
        try:
            if os.path.exists(self.historical_progress_file):
                with open(self.historical_progress_file, 'r') as f:
                    return json.load(f)
            else:
                # Initialize progress tracking
                return {
                    'last_historical_date': None,
                    'total_historical_records': 0,
                    'target_start_date': '2015-01-01',
                    'days_per_run': 180  # Add 180 days (6 months) of historical data per run
                }
        except Exception as e:
            self.logger.error(f"Error loading historical progress: {e}")
            return {
                'last_historical_date': None,
                'total_historical_records': 0,
                'target_start_date': '2015-01-01',
                'days_per_run': 180  # Add 180 days (6 months) of historical data per run
            }

    def save_historical_progress(self, progress):
        """Save historical data collection progress"""
        try:
            with open(self.historical_progress_file, 'w') as f:
                json.dump(progress, f, indent=2)
        except Exception as e:
            self.logger.error(f"Error saving historical progress: {e}")

    def create_combined_dataset(self):
        """Create combined historical + current dataset with incremental historical data"""
        try:
            self.logger.info("Creating combined historical + current dataset")
            
            # Get current data
            current_data = self.collect_all_metrics()
            current_df = pd.DataFrame([current_data])
            
            # Load existing historical data
            existing_historical_df = self.load_existing_historical_data()
            
            # Get progress tracking
            progress = self.load_historical_progress()
            
            # Add incremental historical data
            new_historical_data = self.add_incremental_historical_data(progress)
            
            # Combine all historical data
            if not existing_historical_df.empty and not new_historical_data.empty:
                all_historical_df = pd.concat([existing_historical_df, new_historical_data], ignore_index=True)
            elif not existing_historical_df.empty:
                all_historical_df = existing_historical_df
            elif not new_historical_data.empty:
                all_historical_df = new_historical_data
            else:
                all_historical_df = pd.DataFrame()
            
            # Remove duplicates and sort
            if not all_historical_df.empty:
                all_historical_df = all_historical_df.drop_duplicates(subset=['date'])
                all_historical_df = all_historical_df.sort_values('date')
                all_historical_df = all_historical_df.reset_index(drop=True)
            
            # Mark current data
            current_df['is_historical'] = False
            
            # Combine historical and current data
            if not all_historical_df.empty:
                combined_df = pd.concat([all_historical_df, current_df], ignore_index=True)
            else:
                combined_df = current_df
            
            # Sort by date (oldest first)
            combined_df = combined_df.sort_values('date')
            combined_df = combined_df.reset_index(drop=True)
            
            # Clean and format data
            combined_df = self.clean_and_format_data(combined_df)
            
            # Save combined dataset
            self.save_combined_dataset(combined_df)
            
            # Update progress
            if not new_historical_data.empty:
                progress['total_historical_records'] = len(all_historical_df)
                progress['last_historical_date'] = all_historical_df['date'].max()
                self.save_historical_progress(progress)
            
            self.logger.info(f"Created combined dataset with {len(combined_df)} records ({len(all_historical_df)} historical + {len(current_df)} current)")
            return combined_df
            
        except Exception as e:
            self.logger.error(f"Error creating combined dataset: {e}")
            return pd.DataFrame()

    def load_existing_historical_data(self):
        """Load existing historical data from combined file"""
        try:
            if os.path.exists(self.combined_file):
                df = pd.read_excel(self.combined_file, sheet_name='Combined Data')
                historical_df = df[df['is_historical'] == True] if 'is_historical' in df.columns else pd.DataFrame()
                return historical_df
            return pd.DataFrame()
        except Exception as e:
            self.logger.error(f"Error loading existing historical data: {e}")
            return pd.DataFrame()

    def add_incremental_historical_data(self, progress):
        """Add a few days of historical data incrementally"""
        try:
            # Determine date range for this run
            if progress['last_historical_date']:
                start_date = pd.to_datetime(progress['last_historical_date']) + pd.Timedelta(days=1)
            else:
                start_date = pd.to_datetime(progress['target_start_date'])
            
            end_date = start_date + pd.Timedelta(days=progress['days_per_run'])
            today = pd.Timestamp.now().normalize()
            
            # Don't go beyond today
            if end_date > today:
                end_date = today
            
            if start_date >= today:
                self.logger.info("Historical data collection complete - reached current date")
                return pd.DataFrame()
            
            self.logger.info(f"Adding historical data from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
            
            historical_data = []
            
            # Get S&P 500 historical data for the date range
            sp500_hist = self.get_historical_data("^GSPC", "10y")
            if sp500_hist is not None and not sp500_hist.empty:
                # Convert timezone-aware index to timezone-naive for comparison
                sp500_hist.index = sp500_hist.index.tz_localize(None) if sp500_hist.index.tz else sp500_hist.index
                
                # Filter to our date range
                date_range_data = sp500_hist[(sp500_hist.index >= start_date) & (sp500_hist.index <= end_date)]
                
                for date, row in date_range_data.iterrows():
                    data = {
                        'date': date.strftime('%Y-%m-%d'),
                        'time': '00:00:00',
                        'timestamp': date.strftime('%Y-%m-%d %H:%M:%S'),
                        'sp500_price': row['Close'],
                        'sp500_pe_estimate': None,
                        'vix_level': None,
                        'vix_interpretation': None,
                        'concentration_ratio': None,
                        'ten_year_treasury': None,
                        'fed_funds_rate_approx': None,
                        'bubble_risk_score': None,
                        'bubble_risk_level': None,
                        'risk_factors': None,
                        'total_ai_market_cap': None,
                        'nvidia_dominance_ratio': None,
                        'company_breakdown': None,
                        'top_10_market_cap': None,
                        'sp500_total_market_cap': None,
                        'is_historical': True
                    }
                    
                    # Add AI stock data for this date (simplified for incremental approach)
                    for ticker, company in self.ai_stocks.items():
                        try:
                            stock_hist = self.get_historical_data(ticker, "10y")
                            if stock_hist is not None and not stock_hist.empty:
                                if date in stock_hist.index:
                                    stock_row = stock_hist.loc[date]
                                    data[f"{company.lower().replace(' ', '_')}_price"] = stock_row['Close']
                                else:
                                    data[f"{company.lower().replace(' ', '_')}_price"] = None
                            else:
                                data[f"{company.lower().replace(' ', '_')}_price"] = None
                            data[f"{company.lower().replace(' ', '_')}_market_cap"] = None
                            data[f"{company.lower().replace(' ', '_')}_pe"] = None
                        except Exception as e:
                            self.logger.warning(f"Error getting {ticker} data for {date}: {e}")
                            data[f"{company.lower().replace(' ', '_')}_price"] = None
                            data[f"{company.lower().replace(' ', '_')}_market_cap"] = None
                            data[f"{company.lower().replace(' ', '_')}_pe"] = None
                    
                    # Add index data
                    for ticker, name in self.indices.items():
                        try:
                            index_hist = self.get_historical_data(ticker, "10y")
                            if index_hist is not None and not index_hist.empty:
                                if date in index_hist.index:
                                    index_row = index_hist.loc[date]
                                    safe_name = name.lower().replace(' ', '_').replace('-', '_')
                                    data[f"{safe_name}_price"] = index_row['Close']
                                else:
                                    safe_name = name.lower().replace(' ', '_').replace('-', '_')
                                    data[f"{safe_name}_price"] = None
                            else:
                                safe_name = name.lower().replace(' ', '_').replace('-', '_')
                                data[f"{safe_name}_price"] = None
                        except Exception as e:
                            self.logger.warning(f"Error getting {ticker} data for {date}: {e}")
                            safe_name = name.lower().replace(' ', '_').replace('-', '_')
                            data[f"{safe_name}_price"] = None
                    
                    historical_data.append(data)
            
            # Add VIX and Treasury data for the date range
            try:
                vix_hist = self.get_historical_data("^VIX", "10y")
                if vix_hist is not None and not vix_hist.empty:
                    # Convert timezone-aware index to timezone-naive for comparison
                    vix_hist.index = vix_hist.index.tz_localize(None) if vix_hist.index.tz else vix_hist.index
                    date_range_vix = vix_hist[(vix_hist.index >= start_date) & (vix_hist.index <= end_date)]
                    for hist_data in historical_data:
                        target_date = pd.to_datetime(hist_data['date'])
                        if target_date in date_range_vix.index:
                            hist_data['vix_level'] = date_range_vix.loc[target_date]['Close']
                            hist_data['vix_interpretation'] = self.interpret_vix(hist_data['vix_level'])
            except Exception as e:
                self.logger.warning(f"Error getting VIX historical data: {e}")
            
            try:
                tnx_hist = self.get_historical_data("^TNX", "10y")
                if tnx_hist is not None and not tnx_hist.empty:
                    # Convert timezone-aware index to timezone-naive for comparison
                    tnx_hist.index = tnx_hist.index.tz_localize(None) if tnx_hist.index.tz else tnx_hist.index
                    date_range_tnx = tnx_hist[(tnx_hist.index >= start_date) & (tnx_hist.index <= end_date)]
                    for hist_data in historical_data:
                        target_date = pd.to_datetime(hist_data['date'])
                        if target_date in date_range_tnx.index:
                            hist_data['ten_year_treasury'] = date_range_tnx.loc[target_date]['Close']
            except Exception as e:
                self.logger.warning(f"Error getting Treasury historical data: {e}")
            
            # Convert to DataFrame
            new_historical_df = pd.DataFrame(historical_data)
            self.logger.info(f"Added {len(new_historical_df)} new historical records")
            return new_historical_df
            
        except Exception as e:
            self.logger.error(f"Error adding incremental historical data: {e}")
            return pd.DataFrame()

    def save_combined_dataset(self, combined_df):
        """Save combined dataset to Excel file"""
        try:
            if combined_df.empty:
                self.logger.warning("No combined data to save")
                return
            
            # Create workbook with multiple sheets
            wb = Workbook()
            wb.remove(wb.active)
            
            # Create combined summary sheet
            self.create_combined_summary_sheet(combined_df, wb)
            
            # Create combined dataset sheet
            ws_data = wb.create_sheet('Combined Data', 1)
            
            # Define the proper column order for the combined dataset
            proper_columns = [
                'date', 'time', 'timestamp', 'is_historical', 'vix_level', 'vix_interpretation',
                'sp500_price', 'sp500_pe_estimate', 'top_10_market_cap', 'sp500_total_market_cap',
                'concentration_ratio', 'company_breakdown', 'ten_year_treasury', 'fed_funds_rate_approx',
                'nvidia_price', 'nvidia_market_cap', 'nvidia_pe',
                'microsoft_price', 'microsoft_market_cap', 'microsoft_pe',
                'alphabet_price', 'alphabet_market_cap', 'alphabet_pe',
                'apple_price', 'apple_market_cap', 'apple_pe',
                'amazon_price', 'amazon_market_cap', 'amazon_pe',
                'meta_price', 'meta_market_cap', 'meta_pe',
                'tesla_price', 'tesla_market_cap', 'tesla_pe',
                'total_ai_market_cap', 'nvidia_dominance_ratio',
                'bubble_risk_score', 'bubble_risk_level', 'risk_factors',
                's&p_500_price', 'nasdaq_price', 'vix_price', '10_year_treasury_price'
            ]
            
            # Add headers
            ws_data.append(proper_columns)
            
            # Add combined data
            for _, row in combined_df.iterrows():
                data_row = [row.get(col, '') for col in proper_columns]
                ws_data.append(data_row)
            
            # Save combined file to OneDrive
            wb.save(self.combined_file)
            self.logger.info(f"Combined dataset saved to: {self.combined_file}")
            
            # Apply intelligent conditional formatting to OneDrive file
            self.apply_combined_formatting(self.combined_file, combined_df)
            
            # Also save combined dataset to Downloads location
            if self.downloads_file:
                try:
                    wb.save(self.downloads_file)
                    self.logger.info(f"Combined dataset also saved to: {self.downloads_file}")
                    
                    # Apply intelligent conditional formatting to Downloads file
                    self.apply_combined_formatting(self.downloads_file, combined_df)
                except Exception as e:
                    self.logger.error(f"Error saving combined dataset to Downloads: {e}")
            
        except Exception as e:
            self.logger.error(f"Error saving combined dataset: {e}")

    def create_combined_summary_sheet(self, combined_df, wb):
        """Create summary sheet for combined data"""
        try:
            ws_summary = wb.create_sheet('Combined Summary', 0)
            
            # Define styles
            header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=14)
            subheader_font = Font(bold=True, size=12)
            data_font = Font(size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            row = 1
            
            # Title
            ws_summary.merge_cells(f'A{row}:D{row}')
            ws_summary[f'A{row}'] = "AI Bubble Indicator - Combined Historical & Current Dataset"
            ws_summary[f'A{row}'].font = Font(bold=True, size=16, color="2F4F4F")
            ws_summary[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2
            
            # Dataset info
            historical_count = len(combined_df[combined_df['is_historical'] == True])
            current_count = len(combined_df[combined_df['is_historical'] == False])
            
            ws_summary[f'A{row}'] = f"Total Records: {len(combined_df)} (Historical: {historical_count}, Current: {current_count})"
            ws_summary[f'A{row}'].font = Font(italic=True, size=10)
            row += 1
            
            ws_summary[f'A{row}'] = f"Date Range: {combined_df['date'].min()} to {combined_df['date'].max()}"
            ws_summary[f'A{row}'].font = Font(italic=True, size=10)
            row += 2
            
            # Current vs Historical Analysis
            ws_summary[f'A{row}'] = "CURRENT vs HISTORICAL ANALYSIS"
            ws_summary[f'A{row}'].font = subheader_font
            ws_summary[f'A{row}'].fill = header_fill
            ws_summary[f'A{row}'].font = Font(bold=True, color="FFFFFF")
            row += 1
            
            # Get current data
            current_data = combined_df[combined_df['is_historical'] == False].iloc[0] if len(combined_df[combined_df['is_historical'] == False]) > 0 else None
            historical_data = combined_df[combined_df['is_historical'] == True]
            
            if current_data is not None and not historical_data.empty:
                # S&P 500 analysis
                sp500_hist = historical_data['sp500_price'].dropna()
                if not sp500_hist.empty and current_data.get('sp500_price'):
                    current_sp500 = current_data['sp500_price']
                    hist_min = sp500_hist.min()
                    hist_max = sp500_hist.max()
                    hist_avg = sp500_hist.mean()
                    hist_percentile = (current_sp500 - hist_min) / (hist_max - hist_min) * 100
                    
                    ws_summary[f'A{row}'] = "S&P 500 Current:"
                    ws_summary[f'B{row}'] = f"${current_sp500:,.2f}"
                    ws_summary[f'C{row}'] = f"Percentile: {hist_percentile:.1f}%"
                    row += 1
                    
                    ws_summary[f'A{row}'] = "S&P 500 Historical Range:"
                    ws_summary[f'B{row}'] = f"${hist_min:,.2f} - ${hist_max:,.2f}"
                    ws_summary[f'C{row}'] = f"Average: ${hist_avg:,.2f}"
                    row += 2
                
                # VIX analysis
                vix_hist = historical_data['vix_level'].dropna()
                if not vix_hist.empty and current_data.get('vix_level'):
                    current_vix = current_data['vix_level']
                    hist_min = vix_hist.min()
                    hist_max = vix_hist.max()
                    hist_avg = vix_hist.mean()
                    hist_percentile = (current_vix - hist_min) / (hist_max - hist_min) * 100
                    
                    ws_summary[f'A{row}'] = "VIX Current:"
                    ws_summary[f'B{row}'] = f"{current_vix:.2f}"
                    ws_summary[f'C{row}'] = f"Percentile: {hist_percentile:.1f}%"
                    row += 1
                    
                    ws_summary[f'A{row}'] = "VIX Historical Range:"
                    ws_summary[f'B{row}'] = f"{hist_min:.2f} - {hist_max:.2f}"
                    ws_summary[f'C{row}'] = f"Average: {hist_avg:.2f}"
                    row += 2
            
            # Apply borders and formatting
            for row_num in range(1, row):
                for col in ['A', 'B', 'C', 'D']:
                    cell = ws_summary[f'{col}{row_num}']
                    cell.border = border
                    if row_num > 1:
                        cell.alignment = Alignment(vertical='center')
            
            # Auto-adjust column widths
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20
            ws_summary.column_dimensions['C'].width = 25
            ws_summary.column_dimensions['D'].width = 20
            
            self.logger.info("Created combined summary sheet")
            
        except Exception as e:
            self.logger.error(f"Error creating combined summary sheet: {e}")

    def apply_combined_formatting(self, filepath, combined_df):
        """Apply intelligent conditional formatting based on historical ranges"""
        try:
            wb = load_workbook(filepath)
            
            # Apply formatting to Combined Data sheet
            if 'Combined Data' in wb.sheetnames:
                ws = wb['Combined Data']
                
                # Calculate historical ranges for conditional formatting
                historical_data = combined_df[combined_df['is_historical'] == True]
                
                # Define historical percentiles for key metrics
                metrics_ranges = {}
                
                # S&P 500
                sp500_hist = historical_data['sp500_price'].dropna()
                if not sp500_hist.empty:
                    metrics_ranges['sp500_price'] = {
                        'min': sp500_hist.min(),
                        'max': sp500_hist.max(),
                        'p25': sp500_hist.quantile(0.25),
                        'p75': sp500_hist.quantile(0.75),
                        'p90': sp500_hist.quantile(0.90),
                        'p95': sp500_hist.quantile(0.95)
                    }
                
                # VIX
                vix_hist = historical_data['vix_level'].dropna()
                if not vix_hist.empty:
                    metrics_ranges['vix_level'] = {
                        'min': vix_hist.min(),
                        'max': vix_hist.max(),
                        'p25': vix_hist.quantile(0.25),
                        'p75': vix_hist.quantile(0.75),
                        'p90': vix_hist.quantile(0.90),
                        'p95': vix_hist.quantile(0.95)
                    }
                
                # Format header row
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # Apply conditional formatting to data rows
                for row in range(2, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        column_name = ws.cell(row=1, column=col).value
                        
                        # Check if this is a current (non-historical) row
                        is_current = ws.cell(row=row, column=4).value == False  # is_historical column
                        
                        if column_name in metrics_ranges and cell.value is not None and isinstance(cell.value, (int, float)):
                            ranges = metrics_ranges[column_name]
                            value = float(cell.value)
                            
                            # Determine percentile
                            if value <= ranges['p25']:
                                percentile = 'very_low'
                            elif value <= ranges['p75']:
                                percentile = 'normal'
                            elif value <= ranges['p90']:
                                percentile = 'high'
                            elif value <= ranges['p95']:
                                percentile = 'very_high'
                            else:
                                percentile = 'extreme'
                            
                            # Apply color based on percentile and whether it's current data
                            if is_current:  # Current data gets more prominent colors
                                if percentile == 'very_low':
                                    cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")  # Light green
                                elif percentile == 'normal':
                                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
                                elif percentile == 'high':
                                    cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
                                elif percentile == 'very_high':
                                    cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
                                else:  # extreme
                                    cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Bright red
                            else:  # Historical data gets lighter colors
                                if percentile == 'very_low':
                                    cell.fill = PatternFill(start_color="F0F8F0", end_color="F0F8F0", fill_type="solid")  # Very light green
                                elif percentile == 'normal':
                                    cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")  # Very light gray
                                elif percentile == 'high':
                                    cell.fill = PatternFill(start_color="FFF8E0", end_color="FFF8E0", fill_type="solid")  # Very light orange
                                elif percentile == 'very_high':
                                    cell.fill = PatternFill(start_color="FFE8E8", end_color="FFE8E8", fill_type="solid")  # Very light red
                                else:  # extreme
                                    cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")  # Light red
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value is not None:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filepath)
                self.logger.info("Applied intelligent conditional formatting to Combined Data sheet")
                return
                
        except Exception as e:
            self.logger.error(f"Error applying combined formatting: {e}")

    def create_summary_sheet(self, df, wb):
        """Create a summary sheet with key metrics and definitions"""
        try:
            # Create or get summary sheet
            if 'Summary' in wb.sheetnames:
                ws_summary = wb['Summary']
                ws_summary.delete_rows(1, ws_summary.max_row)  # Clear existing data
            else:
                ws_summary = wb.create_sheet('Summary', 0)  # Insert as first sheet
            
            # Define styles
            header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=14)
            subheader_font = Font(bold=True, size=12)
            data_font = Font(size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            row = 1
            
            # Title
            ws_summary.merge_cells(f'A{row}:D{row}')
            ws_summary[f'A{row}'] = "AI Bubble Indicator Dashboard"
            ws_summary[f'A{row}'].font = Font(bold=True, size=16, color="2F4F4F")
            ws_summary[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 2
            
            # Last updated
            ws_summary[f'A{row}'] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_summary[f'A{row}'].font = Font(italic=True, size=10)
            row += 2
            
            # Data collection summary
            if not df.empty:
                latest = df.iloc[0]
                total_days = len(df)
                
                # Key metrics section
                ws_summary[f'A{row}'] = "KEY METRICS"
                ws_summary[f'A{row}'].font = subheader_font
                ws_summary[f'A{row}'].fill = header_fill
                ws_summary[f'A{row}'].font = Font(bold=True, color="FFFFFF")
                row += 1
                
                # Current date and days recorded
                ws_summary[f'A{row}'] = "Current Date:"
                ws_summary[f'B{row}'] = latest.get('date', 'N/A')
                ws_summary[f'C{row}'] = "Days Recorded:"
                ws_summary[f'D{row}'] = total_days
                row += 1
                
                # VIX
                ws_summary[f'A{row}'] = "VIX Level:"
                ws_summary[f'B{row}'] = f"{latest.get('vix_level', 'N/A'):.2f}" if latest.get('vix_level') else 'N/A'
                ws_summary[f'C{row}'] = "VIX Interpretation:"
                ws_summary[f'D{row}'] = latest.get('vix_interpretation', 'N/A')
                row += 1
                
                # Bubble risk
                ws_summary[f'A{row}'] = "Bubble Risk Level:"
                ws_summary[f'B{row}'] = latest.get('bubble_risk_level', 'N/A')
                ws_summary[f'C{row}'] = "Risk Score:"
                ws_summary[f'D{row}'] = f"{latest.get('bubble_risk_score', 'N/A'):.1f}" if latest.get('bubble_risk_score') else 'N/A'
                row += 1
                
                # Market concentration
                ws_summary[f'A{row}'] = "Market Concentration:"
                ws_summary[f'B{row}'] = f"{latest.get('concentration_ratio', 'N/A'):.1f}%" if latest.get('concentration_ratio') else 'N/A'
                ws_summary[f'C{row}'] = "Top 10 Market Cap:"
                ws_summary[f'D{row}'] = f"${latest.get('top_10_market_cap', 'N/A'):,.0f}" if latest.get('top_10_market_cap') else 'N/A'
                row += 1
                
                # S&P 500
                ws_summary[f'A{row}'] = "S&P 500 Price:"
                ws_summary[f'B{row}'] = f"${latest.get('sp500_price', 'N/A'):,.2f}" if latest.get('sp500_price') else 'N/A'
                ws_summary[f'C{row}'] = "S&P 500 P/E:"
                ws_summary[f'D{row}'] = f"{latest.get('sp500_pe_estimate', 'N/A'):.2f}" if latest.get('sp500_pe_estimate') else 'N/A'
                row += 1
                
                # Interest rates
                ws_summary[f'A{row}'] = "10-Year Treasury:"
                ws_summary[f'B{row}'] = f"{latest.get('ten_year_treasury', 'N/A'):.2f}%" if latest.get('ten_year_treasury') else 'N/A'
                ws_summary[f'C{row}'] = "Fed Funds Rate:"
                ws_summary[f'D{row}'] = f"{latest.get('fed_funds_rate_approx', 'N/A'):.2f}%" if latest.get('fed_funds_rate_approx') else 'N/A'
                row += 1
                
                # NVIDIA dominance
                ws_summary[f'A{row}'] = "NVIDIA Dominance:"
                ws_summary[f'B{row}'] = f"{latest.get('nvidia_dominance_ratio', 'N/A'):.1f}%" if latest.get('nvidia_dominance_ratio') else 'N/A'
                ws_summary[f'C{row}'] = "Total AI Market Cap:"
                ws_summary[f'D{row}'] = f"${latest.get('total_ai_market_cap', 'N/A'):,.0f}" if latest.get('total_ai_market_cap') else 'N/A'
                row += 2
                
                # Risk factors
                risk_factors = latest.get('risk_factors', '')
                if risk_factors:
                    ws_summary[f'A{row}'] = "CURRENT RISK FACTORS:"
                    ws_summary[f'A{row}'].font = subheader_font
                    ws_summary[f'A{row}'].fill = header_fill
                    ws_summary[f'A{row}'].font = Font(bold=True, color="FFFFFF")
                    row += 1
                    
                    # Split risk factors and display
                    factors = risk_factors.split('; ')
                    for factor in factors:
                        if factor.strip():
                            ws_summary[f'A{row}'] = f"• {factor.strip()}"
                            ws_summary[f'A{row}'].font = data_font
                            row += 1
                    row += 1
                
                # AI Stock Prices
                ws_summary[f'A{row}'] = "AI STOCK PRICES"
                ws_summary[f'A{row}'].font = subheader_font
                ws_summary[f'A{row}'].fill = header_fill
                ws_summary[f'A{row}'].font = Font(bold=True, color="FFFFFF")
                row += 1
                
                # Headers for AI stocks
                ws_summary[f'A{row}'] = "Company"
                ws_summary[f'B{row}'] = "Price"
                ws_summary[f'C{row}'] = "Market Cap"
                ws_summary[f'D{row}'] = "P/E Ratio"
                for col in ['A', 'B', 'C', 'D']:
                    ws_summary[f'{col}{row}'].font = Font(bold=True)
                    ws_summary[f'{col}{row}'].fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                row += 1
                
                # AI stock data
                for ticker, company in self.ai_stocks.items():
                    price_col = f"{company.lower().replace(' ', '_')}_price"
                    market_cap_col = f"{company.lower().replace(' ', '_')}_market_cap"
                    pe_col = f"{company.lower().replace(' ', '_')}_pe"
                    
                    ws_summary[f'A{row}'] = company
                    ws_summary[f'B{row}'] = f"${latest.get(price_col, 'N/A'):,.2f}" if latest.get(price_col) else 'N/A'
                    ws_summary[f'C{row}'] = f"${latest.get(market_cap_col, 'N/A'):,.0f}" if latest.get(market_cap_col) else 'N/A'
                    ws_summary[f'D{row}'] = f"{latest.get(pe_col, 'N/A'):.2f}" if latest.get(pe_col) else 'N/A'
                    row += 1
                
                row += 2
            
            # Definitions section
            ws_summary[f'A{row}'] = "DEFINITIONS & RISK THRESHOLDS"
            ws_summary[f'A{row}'].font = subheader_font
            ws_summary[f'A{row}'].fill = header_fill
            ws_summary[f'A{row}'].font = Font(bold=True, color="FFFFFF")
            row += 1
            
            definitions = [
                ("VIX (Volatility Index)", "Measures market fear/volatility. Lower = complacency risk, Higher = panic"),
                ("Market Concentration", "Top 10 companies as % of S&P 500. Higher = concentration risk"),
                ("Bubble Risk Score", "0-10 scale. 0-2=Low, 3-4=Elevated, 5-6=Moderate, 7+=High risk"),
                ("S&P 500 P/E Ratio", "Price-to-earnings ratio. Higher = overvaluation risk"),
                ("NVIDIA Dominance", "NVIDIA market cap as % of total AI sector. Higher = single-stock risk"),
                ("10-Year Treasury", "Government bond yield. Lower = asset inflation risk"),
                ("Risk Color Coding", "White=Low, Orange=Medium, Red=High, Bright Red=Extreme risk")
            ]
            
            for term, definition in definitions:
                ws_summary[f'A{row}'] = term
                ws_summary[f'A{row}'].font = Font(bold=True)
                ws_summary[f'B{row}'] = definition
                ws_summary[f'B{row}'].font = data_font
                row += 1
            
            # Apply borders and formatting
            for row_num in range(1, row):
                for col in ['A', 'B', 'C', 'D']:
                    cell = ws_summary[f'{col}{row_num}']
                    cell.border = border
                    if row_num > 1:  # Skip title row
                        cell.alignment = Alignment(vertical='center')
            
            # Auto-adjust column widths
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 20
            ws_summary.column_dimensions['C'].width = 25
            ws_summary.column_dimensions['D'].width = 20
            
            self.logger.info("Created summary sheet")
            
        except Exception as e:
            self.logger.error(f"Error creating summary sheet: {e}")

    def get_stock_data(self, ticker, period="1d"):
        """Get current stock data using yfinance"""
        try:
            stock = yf.Ticker(ticker)
            hist = stock.history(period=period)
            info = stock.info

            if hist.empty:
                self.logger.warning(f"No historical data for {ticker}")
                return None

            current_price = hist['Close'].iloc[-1]
            market_cap = info.get('marketCap', 0)
            pe_ratio = info.get('trailingPE', None)
            forward_pe = info.get('forwardPE', None)
            price_to_sales = info.get('priceToSalesTrailing12Months', None)
            volume = hist['Volume'].iloc[-1] if 'Volume' in hist.columns else None

            return {
                'ticker': ticker,
                'price': current_price,
                'market_cap': market_cap,
                'pe_ratio': pe_ratio,
                'forward_pe': forward_pe,
                'price_to_sales': price_to_sales,
                'volume': volume
            }
        except Exception as e:
            self.logger.error(f"Error getting data for {ticker}: {e}")
            return None

    def get_vix_data(self):
        """Get VIX (Volatility Index) data"""
        try:
            vix = yf.Ticker("^VIX")
            hist = vix.history(period="1d")
            if not hist.empty:
                current_vix = hist['Close'].iloc[-1]
                return {
                    'vix_level': current_vix,
                    'vix_interpretation': self.interpret_vix(current_vix)
                }
        except Exception as e:
            self.logger.error(f"Error getting VIX data: {e}")
        return {'vix_level': None, 'vix_interpretation': 'N/A'}

    def interpret_vix(self, vix_level):
        """Interpret VIX levels for bubble analysis"""
        try:
            vix_level = float(vix_level)
        except Exception:
            return "N/A"
        if vix_level < 12:
            return "Very Low - Complacency Risk"
        elif vix_level < 20:
            return "Low - Normal Market"
        elif vix_level < 30:
            return "Elevated - Increased Volatility"
        elif vix_level < 40:
            return "High - Market Stress"
        else:
            return "Very High - Panic/Crisis"

    def get_sp500_pe_ratio(self):
        """Get S&P 500 P/E ratio from multiple sources"""
        try:
            sp500 = yf.Ticker("^GSPC")
            hist = sp500.history(period="1d")
            info = sp500.info
            current_price = hist['Close'].iloc[-1] if not hist.empty else None
            
            # Try multiple methods to get P/E ratio
            pe_ratio = None
            
            # Method 1: Direct from info
            if info.get('trailingPE'):
                pe_ratio = info.get('trailingPE')
            # Method 2: Calculate from market cap and earnings
            elif info.get('marketCap') and info.get('trailingEps'):
                market_cap = info.get('marketCap')
                trailing_eps = info.get('trailingEps')
                if market_cap and trailing_eps and trailing_eps > 0:
                    pe_ratio = market_cap / (trailing_eps * 1000000000)  # Convert to billions
            # Method 3: Use SPY as proxy (more reliable for P/E)
            else:
                try:
                    spy = yf.Ticker("SPY")
                    spy_info = spy.info
                    if spy_info.get('trailingPE'):
                        pe_ratio = spy_info.get('trailingPE')
                except:
                    pass
            
            # Method 4: Use historical average if still no P/E
            if not pe_ratio:
                # S&P 500 historical average P/E is around 15-20
                pe_ratio = 18.5  # Conservative estimate
                self.logger.warning("Using estimated S&P 500 P/E ratio")
            
            return {
                'sp500_price': current_price,
                'sp500_pe_estimate': pe_ratio
            }
        except Exception as e:
            self.logger.error(f"Error getting S&P 500 P/E: {e}")
            return {'sp500_price': None, 'sp500_pe_estimate': 18.5}  # Fallback estimate

    def calculate_market_concentration(self):
        """Calculate market concentration of top 10 companies vs S&P 500"""
        try:
            # Top 10 companies by market cap (as of 2024)
            top_companies = ['AAPL', 'MSFT', 'NVDA', 'GOOGL', 'AMZN', 'META', 'TSLA', 'BRK-B', 'LLY', 'V']
            total_top_10_cap = 0
            company_data = {}
            
            for ticker in top_companies:
                data = self.get_stock_data(ticker)
                if data and data['market_cap']:
                    total_top_10_cap += data['market_cap']
                    company_data[ticker] = data['market_cap']
            
            # Get S&P 500 total market cap (more accurate than approximation)
            try:
                # Use SPY as proxy for S&P 500 market cap
                spy = yf.Ticker("SPY")
                spy_info = spy.info
                sp500_market_cap = spy_info.get('totalAssets', 0)  # SPY total assets approximate S&P 500 market cap
                
                # If SPY data not available, use historical S&P 500 market cap (~$40-50 trillion)
                if not sp500_market_cap or sp500_market_cap < 1000000000000:  # Less than 1 trillion
                    sp500_market_cap = 45000000000000  # ~$45 trillion estimate
                    self.logger.warning("Using estimated S&P 500 market cap")
            except:
                sp500_market_cap = 45000000000000  # Fallback estimate
            
            # Calculate concentration ratio: Top 10 market cap / S&P 500 total market cap
            concentration_ratio = (total_top_10_cap / sp500_market_cap) * 100 if sp500_market_cap > 0 else 0
            
            return {
                'top_10_market_cap': total_top_10_cap,
                'sp500_total_market_cap': sp500_market_cap,
                'concentration_ratio': concentration_ratio,
                'company_breakdown': json.dumps(company_data)  # Convert dict to JSON string
            }
        except Exception as e:
            self.logger.error(f"Error calculating market concentration: {e}")
            return {'top_10_market_cap': 0, 'sp500_total_market_cap': 0, 'concentration_ratio': 0, 'company_breakdown': '{}'}

    def get_interest_rates(self):
        """Get current interest rates"""
        try:
            tnx = yf.Ticker("^TNX")
            hist = tnx.history(period="1d")
            ten_year_yield = hist['Close'].iloc[-1] if not hist.empty else None

            three_month = yf.Ticker("^IRX")
            hist_3m = three_month.history(period="1d")
            fed_funds_approx = hist_3m['Close'].iloc[-1] if not hist_3m.empty else None

            return {
                'ten_year_treasury': ten_year_yield,
                'fed_funds_rate_approx': fed_funds_approx
            }
        except Exception as e:
            self.logger.error(f"Error getting interest rates: {e}")
            return {'ten_year_treasury': None, 'fed_funds_rate_approx': None}

    def get_ai_sector_metrics(self):
        """Get specific AI sector bubble indicators"""
        try:
            ai_metrics = {}
            total_ai_market_cap = 0
            for ticker, company in self.ai_stocks.items():
                data = self.get_stock_data(ticker)
                safe_name = company.lower().replace(' ', '_')
                if data:
                    ai_metrics[f"{safe_name}_price"] = data['price']
                    ai_metrics[f"{safe_name}_market_cap"] = data['market_cap']
                    ai_metrics[f"{safe_name}_pe"] = data['pe_ratio']
                    if data['market_cap']:
                        total_ai_market_cap += data['market_cap']
            ai_metrics['total_ai_market_cap'] = total_ai_market_cap
            nvidia_data = self.get_stock_data('NVDA')
            if nvidia_data and nvidia_data['market_cap'] and total_ai_market_cap > 0:
                ai_metrics['nvidia_dominance_ratio'] = (nvidia_data['market_cap'] / total_ai_market_cap) * 100
            else:
                ai_metrics['nvidia_dominance_ratio'] = 0
            return ai_metrics
        except Exception as e:
            self.logger.error(f"Error getting AI sector metrics: {e}")
            return {}

    def assess_bubble_risk(self, data):
        """Assess overall bubble risk based on collected metrics"""
        risk_factors = []
        risk_score = 0

        vix_level = data.get('vix_level', 50)
        try:
            vix_level = float(vix_level)
        except Exception:
            vix_level = 50

        if vix_level < 12:
            risk_factors.append("VIX extremely low - complacency risk")
            risk_score += 2
        elif vix_level < 15:
            risk_factors.append("VIX low - reduced fear")
            risk_score += 1

        concentration = data.get('concentration_ratio', 0)
        try:
            concentration = float(concentration)
        except Exception:
            concentration = 0

        if concentration > 35:
            risk_factors.append("High market concentration risk")
            risk_score += 2
        elif concentration > 30:
            risk_factors.append("Elevated market concentration")
            risk_score += 1

        nvidia_dominance = data.get('nvidia_dominance_ratio', 0)
        try:
            nvidia_dominance = float(nvidia_dominance)
        except Exception:
            nvidia_dominance = 0

        if nvidia_dominance > 40:
            risk_factors.append("NVIDIA market dominance risk")
            risk_score += 2
        elif nvidia_dominance > 30:
            risk_factors.append("High NVIDIA concentration")
            risk_score += 1

        sp500_pe = data.get('sp500_pe_estimate', 20)
        try:
            sp500_pe = float(sp500_pe)
        except Exception:
            sp500_pe = 20

        if sp500_pe > 30:
            risk_factors.append("Elevated S&P 500 P/E ratio")
            risk_score += 2
        elif sp500_pe > 25:
            risk_factors.append("High S&P 500 P/E ratio")
            risk_score += 1

        ten_year = data.get('ten_year_treasury', 5)
        try:
            ten_year = float(ten_year)
        except Exception:
            ten_year = 5

        if ten_year < 2:
            risk_factors.append("Very low interest rates - asset inflation risk")
            risk_score += 1

        if risk_score >= 6:
            risk_level = "HIGH BUBBLE RISK"
        elif risk_score >= 4:
            risk_level = "MODERATE BUBBLE RISK"
        elif risk_score >= 2:
            risk_level = "ELEVATED BUBBLE RISK"
        else:
            risk_level = "LOW BUBBLE RISK"

        return {
            'bubble_risk_score': risk_score,
            'bubble_risk_level': risk_level,
            'risk_factors': '; '.join(risk_factors)
        }

    def collect_all_metrics(self):
        """Collect all bubble indicator metrics"""
        self.logger.info("Starting bubble indicator data collection")

        data = {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'time': datetime.now().strftime('%H:%M:%S'),
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

        self.logger.info("Collecting VIX data...")
        vix_data = self.get_vix_data()
        data.update(vix_data)

        self.logger.info("Collecting S&P 500 data...")
        sp500_data = self.get_sp500_pe_ratio()
        data.update(sp500_data)

        self.logger.info("Calculating market concentration...")
        concentration_data = self.calculate_market_concentration()
        data.update(concentration_data)

        self.logger.info("Collecting interest rate data...")
        rates_data = self.get_interest_rates()
        data.update(rates_data)

        self.logger.info("Collecting AI sector metrics...")
        ai_data = self.get_ai_sector_metrics()
        data.update(ai_data)

        self.logger.info("Assessing bubble risk...")
        risk_assessment = self.assess_bubble_risk(data)
        data.update(risk_assessment)

        self.logger.info("Collecting market indices...")
        for ticker, name in self.indices.items():
            try:
                index_data = self.get_stock_data(ticker)
                if index_data:
                    safe_name = name.lower().replace(' ', '_').replace('-', '_')
                    data[f"{safe_name}_price"] = index_data['price']
                    if index_data['pe_ratio'] is not None:
                        data[f"{safe_name}_pe"] = index_data['pe_ratio']
            except Exception as e:
                self.logger.error(f"Error getting {name} data: {e}")

        self.logger.info(f"Collected {len(data)} metrics")
        return data

    def load_existing_dataset(self):
        """Load existing dataset if it exists"""
        if os.path.exists(self.master_file):
            try:
                df = pd.read_excel(self.master_file)
                self.logger.info(f"Loaded existing dataset with {len(df)} records")
                return df
            except Exception as e:
                self.logger.error(f"Error loading existing dataset: {e}")
                return pd.DataFrame()
        else:
            self.logger.info("No existing dataset found, creating new one")
            return pd.DataFrame()

    def update_dataset(self, new_data):
        """Update dataset with new daily data"""
        existing_df = self.load_existing_dataset()
        new_df = pd.DataFrame([new_data])

        if existing_df.empty:
            updated_df = new_df
        else:
            today = datetime.now().strftime('%Y-%m-%d')
            if 'date' in existing_df.columns:
                today_mask = existing_df['date'] == today
                if today_mask.any():
                    for col in new_df.columns:
                        if col in existing_df.columns:
                            existing_df.loc[today_mask, col] = new_df.iloc[0][col]
                        else:
                            existing_df[col] = None
                            existing_df.loc[today_mask, col] = new_df.iloc[0][col]
                    updated_df = existing_df
                    self.logger.info("Updated today's existing record")
                else:
                    updated_df = pd.concat([existing_df, new_df], ignore_index=True)
                    self.logger.info("Added new daily record")
            else:
                updated_df = pd.concat([existing_df, new_df], ignore_index=True)
                self.logger.info("Added new daily record (no date column found)")

        updated_df = updated_df.sort_values('date', ascending=False)
        updated_df = updated_df.reset_index(drop=True)
        return updated_df

    def get_risk_color(self, value, column_name, row_data):
        """Get risk-based color for a cell based on its value and context"""
        try:
            # Low risk (white/light)
            low_risk_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White
            very_low_risk_fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")  # Very light gray
            
            # Medium risk (orange/yellow)
            medium_risk_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")  # Light orange
            elevated_risk_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")  # Light yellow
            
            # High risk (red)
            high_risk_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")  # Light red
            very_high_risk_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Bright red
            
            # VIX levels
            if column_name == 'vix_level':
                if value < 12:
                    return very_low_risk_fill  # Very low VIX
                elif value < 20:
                    return low_risk_fill  # Low VIX
                elif value < 30:
                    return medium_risk_fill  # Medium VIX
                elif value < 40:
                    return high_risk_fill  # High VIX
                else:
                    return very_high_risk_fill  # Very high VIX
            
            # S&P 500 P/E ratio
            elif column_name == 'sp500_pe_estimate':
                if value < 15:
                    return very_low_risk_fill  # Very low P/E
                elif value < 20:
                    return low_risk_fill  # Low P/E
                elif value < 25:
                    return medium_risk_fill  # Medium P/E
                elif value < 30:
                    return high_risk_fill  # High P/E
                else:
                    return very_high_risk_fill  # Very high P/E
            
            # Market concentration (top 10 companies as % of S&P 500)
            elif column_name == 'concentration_ratio':
                if value < 20:
                    return very_low_risk_fill  # Very low concentration
                elif value < 25:
                    return low_risk_fill  # Low concentration
                elif value < 30:
                    return medium_risk_fill  # Medium concentration
                elif value < 35:
                    return high_risk_fill  # High concentration
                else:
                    return very_high_risk_fill  # Very high concentration
            
            # NVIDIA dominance
            elif column_name == 'nvidia_dominance_ratio':
                if value < 15:
                    return very_low_risk_fill  # Very low dominance
                elif value < 25:
                    return low_risk_fill  # Low dominance
                elif value < 35:
                    return medium_risk_fill  # Medium dominance
                elif value < 45:
                    return high_risk_fill  # High dominance
                else:
                    return very_high_risk_fill  # Very high dominance
            
            # 10-Year Treasury (inverse relationship - lower rates = higher risk)
            elif column_name == 'ten_year_treasury':
                if value > 4.5:
                    return very_low_risk_fill  # High rates = low risk
                elif value > 3.5:
                    return low_risk_fill  # Medium-high rates
                elif value > 2.5:
                    return medium_risk_fill  # Medium rates
                elif value > 1.5:
                    return high_risk_fill  # Low rates
                else:
                    return very_high_risk_fill  # Very low rates
            
            # Bubble risk score
            elif column_name == 'bubble_risk_score':
                if value < 1:
                    return very_low_risk_fill  # Very low risk
                elif value < 3:
                    return low_risk_fill  # Low risk
                elif value < 5:
                    return medium_risk_fill  # Medium risk
                elif value < 7:
                    return high_risk_fill  # High risk
                else:
                    return very_high_risk_fill  # Very high risk
            
            # AI stock P/E ratios
            elif 'pe' in column_name and any(stock.lower().replace(' ', '_') in column_name for stock in self.ai_stocks.values()):
                if value < 15:
                    return very_low_risk_fill  # Very low P/E
                elif value < 25:
                    return low_risk_fill  # Low P/E
                elif value < 35:
                    return medium_risk_fill  # Medium P/E
                elif value < 50:
                    return high_risk_fill  # High P/E
                else:
                    return very_high_risk_fill  # Very high P/E
            
            # Default to low risk for other columns
            return low_risk_fill
            
        except Exception as e:
            self.logger.error(f"Error getting risk color for {column_name}: {e}")
            return PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Default white

    def apply_conditional_formatting(self, filepath):
        """Apply intelligent risk-based conditional formatting to Excel file"""
        try:
            wb = load_workbook(filepath)
            
            # Apply formatting to Dataset sheet (sheet 2) - clean data only
            if 'Dataset' in wb.sheetnames:
                ws = wb['Dataset']
                
                # Simple header formatting only
                if ws.max_row > 0:
                    # Format header row only
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value is not None:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filepath)
                self.logger.info("Applied simple formatting to Dataset sheet")
                return
            
            # Apply dashboard formatting to Summary sheet
            if 'Summary' in wb.sheetnames:
                ws = wb['Summary']
                
                # Define dashboard colors
                header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF", size=14)
                subheader_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
                subheader_font = Font(bold=True, color="FFFFFF", size=12)
                data_font = Font(size=11)
                border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                              top=Side(style='thin'), bottom=Side(style='thin'))
                
                # Apply dashboard formatting to all cells
                for row in range(1, ws.max_row + 1):
                    for col in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row, column=col)
                        cell.border = border
                        
                        # Apply different formatting based on content
                        if cell.value and isinstance(cell.value, str):
                            cell_value = str(cell.value).upper()
                            
                            # Main headers
                            if any(keyword in cell_value for keyword in ['AI BUBBLE INDICATOR', 'KEY METRICS', 'CURRENT RISK FACTORS', 'AI STOCK PRICES', 'DEFINITIONS']):
                                cell.fill = header_fill
                                cell.font = header_font
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Subheaders
                            elif any(keyword in cell_value for keyword in ['COMPANY', 'PRICE', 'MARKET CAP', 'P/E RATIO']):
                                cell.fill = subheader_fill
                                cell.font = subheader_font
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Risk factors
                            elif cell_value.startswith('•'):
                                cell.font = Font(size=10, color="FF4444")  # Red for risk factors
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                            # Definitions
                            elif ':' in cell_value and not any(keyword in cell_value for keyword in ['CURRENT DATE', 'DAYS RECORDED', 'VIX LEVEL', 'BUBBLE RISK', 'MARKET CONCENTRATION', 'S&P 500', '10-YEAR TREASURY', 'NVIDIA DOMINANCE']):
                                cell.font = Font(bold=True, size=11)
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                            # Data labels
                            elif any(keyword in cell_value for keyword in ['CURRENT DATE', 'DAYS RECORDED', 'VIX LEVEL', 'BUBBLE RISK', 'MARKET CONCENTRATION', 'S&P 500', '10-YEAR TREASURY', 'NVIDIA DOMINANCE']):
                                cell.font = Font(bold=True, size=11, color="2F4F4F")
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                            # Regular data
                            else:
                                cell.font = data_font
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Format numeric values with risk-based colors
                        elif isinstance(cell.value, (int, float)) and cell.value != 0:
                            # Apply risk-based background colors
                            if cell.value > 50:  # High risk values
                                cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                            elif cell.value > 30:  # Medium risk values
                                cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                            elif cell.value > 15:  # Low risk values
                                cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                            else:  # Very low risk values
                                cell.fill = PatternFill(start_color="F8F8F8", end_color="F8F8F8", fill_type="solid")
                            
                            cell.font = data_font
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                
                wb.save(filepath)
                self.logger.info("Applied dashboard formatting to Summary sheet")
                return

        except Exception as e:
            self.logger.error(f"Error applying formatting: {e}")

    def save_dataset(self, df):
        """Save dataset with daily backups and improved functionality"""
        if df.empty:
            self.logger.warning("No data to save")
            return
        
        # Clean and format data for proper Excel number formatting
        df = self.clean_and_format_data(df)
        
        # Sort by date (most recent first)
        df = df.sort_values('date', ascending=False)
        df = df.reset_index(drop=True)
        
        saved_locations = []
        
        try:
            # Create timestamped filename for daily backup
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            daily_backup_file = os.path.join(self.daily_backups_dir, f"bubble_indicators_dataset_{timestamp}.xlsx")
            
            # Create workbook with multiple sheets
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create summary sheet (sheet 1)
            self.create_summary_sheet(df, wb)
            
            # Create main dataset sheet (sheet 2)
            ws_data = wb.create_sheet('Dataset', 1)
            
            # Add only the latest daily data to dataset sheet
            if not df.empty:
                latest_row = df.iloc[0]  # Get the most recent record
                
                # Define the proper column order for the dataset
                proper_columns = [
                    'date', 'time', 'timestamp', 'vix_level', 'vix_interpretation',
                    'sp500_price', 'sp500_pe_estimate', 'top_10_market_cap', 'sp500_total_market_cap',
                    'concentration_ratio', 'company_breakdown', 'ten_year_treasury', 'fed_funds_rate_approx',
                    'nvidia_price', 'nvidia_market_cap', 'nvidia_pe',
                    'microsoft_price', 'microsoft_market_cap', 'microsoft_pe',
                    'alphabet_price', 'alphabet_market_cap', 'alphabet_pe',
                    'apple_price', 'apple_market_cap', 'apple_pe',
                    'amazon_price', 'amazon_market_cap', 'amazon_pe',
                    'meta_price', 'meta_market_cap', 'meta_pe',
                    'tesla_price', 'tesla_market_cap', 'tesla_pe',
                    'total_ai_market_cap', 'nvidia_dominance_ratio',
                    'bubble_risk_score', 'bubble_risk_level', 'risk_factors',
                    's&p_500_price', 'nasdaq_price', 'vix_price', '10_year_treasury_price'
                ]
                
                # Add headers if this is the first row
                if ws_data.max_row == 1:  # Empty sheet
                    ws_data.append(proper_columns)
                
                # Add the latest data row in proper order
                data_row = [latest_row.get(col, '') for col in proper_columns]
                ws_data.append(data_row)
            
            # Save main file to OneDrive
            wb.save(self.master_file)
            self.logger.info(f"Bubble indicators saved to: {self.master_file}")
            saved_locations.append(self.master_file)
            
            # Apply formatting to OneDrive file
            self.apply_conditional_formatting(self.master_file)
            
            # Save daily timestamped backup
            wb.save(daily_backup_file)
            self.logger.info(f"Daily backup saved to: {daily_backup_file}")
            saved_locations.append(daily_backup_file)
            
            # Apply formatting to daily backup
            self.apply_conditional_formatting(daily_backup_file)
            
            # Note: Downloads file will be updated with combined historical data later
            # Skip saving daily-only data to Downloads to avoid overwriting historical data
            
            # Print comprehensive summary
            print(f"\n=== Bubble Indicators Summary ===")
            print(f"Total daily records: {len(df)}")
            if not df.empty:
                latest = df.iloc[0]
                print(f"Latest data: {latest['date']} at {latest['time']}")
                print(f"Current VIX: {latest.get('vix_level', 'N/A')} ({latest.get('vix_interpretation', 'N/A')})")
                print(f"Bubble Risk Level: {latest.get('bubble_risk_level', 'N/A')} (Score: {latest.get('bubble_risk_score', 'N/A')})")
                print(f"Market Concentration: {latest.get('concentration_ratio', 'N/A'):.1f}% (Top 10 vs S&P 500)")
                print(f"S&P 500 Price: ${latest.get('sp500_price', 'N/A'):,.2f}" if latest.get('sp500_price') else "S&P 500 Price: N/A")
                print(f"10-Year Treasury: {latest.get('ten_year_treasury', 'N/A')}%")
                print(f"NVIDIA Dominance: {latest.get('nvidia_dominance_ratio', 'N/A'):.1f}%")
                print(f"Top 10 Market Cap: ${latest.get('top_10_market_cap', 'N/A'):,.0f}" if latest.get('top_10_market_cap') else "Top 10 Market Cap: N/A")
                
                # Show risk factors
                risk_factors = latest.get('risk_factors', '')
                if risk_factors:
                    print(f"Risk Factors: {risk_factors}")
            
            print(f"\nSaved to {len(saved_locations)} location(s):")
            for location in saved_locations:
                print(f"  - {location}")
            
        except Exception as e:
            self.logger.error(f"Error saving dataset: {e}")

    def run(self):
        """Main execution method"""
        self.logger.info("Starting Bubble Indicator Scraper")
        start_time = datetime.now()
        
        try:
            # Collect all metrics
            new_data = self.collect_all_metrics()
            
            # Update dataset
            updated_df = self.update_dataset(new_data)
            
            # Save dataset
            self.save_dataset(updated_df)
            
            # Create combined historical + current dataset
            self.logger.info("Creating combined historical + current dataset...")
            combined_df = self.create_combined_dataset()
            
            end_time = datetime.now()
            duration = end_time - start_time
            self.logger.info(f"Bubble indicator update completed in {duration}")
            self.logger.info(f"Created combined dataset with {len(combined_df)} records")
            
        except Exception as e:
            self.logger.error(f"Error during bubble indicator update: {e}")

def main():
    """Main function"""
    scraper = BubbleIndicatorScraper()
    scraper.run()

if __name__ == "__main__":
    main()
