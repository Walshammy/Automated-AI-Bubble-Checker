import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def fix_excel_formatting():
    """Fix Excel file to include all valuation methods with proper formatting"""
    excel_file = r"C:\Users\james\Downloads\Stock Valuation\stock_valuation_dataset.xlsx"
    
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        print("=== FIXING EXCEL FORMATTING ===")
        
        # Remove existing sheets
        if 'My Portfolio' in wb.sheetnames:
            wb.remove(wb['My Portfolio'])
        if 'Prospects' in wb.sheetnames:
            wb.remove(wb['Prospects'])
        
        # Load the latest data
        df = pd.read_excel(excel_file, sheet_name='Valuation Data')
        latest_data = df.sort_values(by='timestamp').groupby('ticker').last().reset_index()
        
        print(f"Loaded {len(latest_data)} stocks")
        
        # Create new My Portfolio sheet with all 9 columns
        create_comprehensive_portfolio_sheet(wb, latest_data)
        
        # Create new Prospects sheet with all 9 columns
        create_comprehensive_prospects_sheet(wb, latest_data)
        
        # Save the updated workbook
        wb.save(excel_file)
        print(f"Successfully updated Excel file: {excel_file}")
        print("Now includes all Tier 1 & Tier 2 valuation methods:")
        print("   • Enhanced DCF with Scenario Analysis")
        print("   • Relative Valuation (EV/EBITDA, P/E, P/S, P/B)")
        print("   • Reverse DCF (What's Priced In?)")
        print("   • Earnings Power Value (EPV)")
        print("   • Residual Income Model (RIM)")
        
    except Exception as e:
        print(f"Error fixing Excel file: {e}")
        import traceback
        traceback.print_exc()

def safe_get_status(stock_row, key, default='N/A'):
    """Safely get status with None handling"""
    value = stock_row.get(key, default)
    return value if value is not None else default

def safe_get_delta(stock_row, key, default=0):
    """Safely get delta with None handling"""
    value = stock_row.get(key, default)
    return value if value is not None and not pd.isna(value) else default

def create_comprehensive_portfolio_sheet(wb, df):
    """Create comprehensive My Portfolio sheet with all valuation methods"""
    ws = wb.create_sheet('My Portfolio', 0)
    
    # Define styles
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    subheader_font = Font(bold=True, size=12)
    data_font = Font(size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                  top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Consistent conditional formatting fills
    strong_buy_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    buy_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
    hold_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    sell_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    strong_sell_fill = PatternFill(start_color="FFA0A0", end_color="FFA0A0", fill_type="solid")
    no_data_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Title
    ws['A1'] = "Stock Valuation Analysis Dashboard"
    ws['A1'].font = Font(bold=True, size=16, color="2F4F4F")
    ws.merge_cells('A1:I1')
    
    # Analysis info
    ws['A3'] = f"Analysis Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"Stocks Analyzed: {len(df)}"
    
    # Headers for all 9 columns
    row = 6
    headers = [
        "Company (Ticker)",
        "Peter Lynch", 
        "DCF Valuation",
        "Munger Farm",
        "Enhanced DCF",
        "Relative Valuation", 
        "Reverse DCF",
        "EPV/RIM",
        "Current Price"
    ]
    
    for i, header in enumerate(headers):
        col_letter = chr(65 + i)  # A, B, C, etc.
        ws[f'{col_letter}{row}'] = header
        cell = ws[f'{col_letter}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Add all stocks (simplified for now)
    for _, stock_row in df.iterrows():
        add_stock_row_safe(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 45
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 35

def create_comprehensive_prospects_sheet(wb, df):
    """Create comprehensive Prospects sheet with all valuation methods"""
    ws = wb.create_sheet('Prospects', 1)
    
    # Define styles
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    subheader_font = Font(bold=True, size=12)
    data_font = Font(size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                  top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Consistent conditional formatting fills
    strong_buy_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    buy_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
    hold_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    sell_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    strong_sell_fill = PatternFill(start_color="FFA0A0", end_color="FFA0A0", fill_type="solid")
    no_data_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Title
    ws['A1'] = "NZX PROSPECTS - RANKED BY UNDERVALUATION"
    ws['A1'].font = Font(bold=True, size=16, color="2F4F4F")
    ws.merge_cells('A1:I1')
    
    # Headers for all 9 columns
    row = 3
    headers = [
        "Company (Ticker)",
        "Peter Lynch", 
        "DCF Valuation",
        "Munger Farm",
        "Enhanced DCF",
        "Relative Valuation", 
        "Reverse DCF",
        "EPV/RIM",
        "Current Price"
    ]
    
    for i, header in enumerate(headers):
        col_letter = chr(65 + i)  # A, B, C, etc.
        ws[f'{col_letter}{row}'] = header
        cell = ws[f'{col_letter}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Add all stocks (simplified for now)
    for _, stock_row in df.iterrows():
        add_stock_row_safe(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill)
        row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 45
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 35

def add_stock_row_safe(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill):
    """Add a stock row with all 9 columns - safe version with None handling"""
    
    # Company name (Column A)
    company_name = stock_row.get('company_name', stock_row.get('ticker', 'Unknown'))
    ticker = stock_row.get('ticker', 'Unknown')
    ws[f'A{row}'] = f"{company_name} ({ticker})"
    ws[f'A{row}'].font = data_font
    ws[f'A{row}'].border = border
    
    # Peter Lynch (Column B)
    lynch_status = safe_get_status(stock_row, 'lynch_valuation_status')
    lynch_delta = safe_get_delta(stock_row, 'lynch_delta_percentage')
    if lynch_status != 'N/A':
        lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
        ws[f'B{row}'] = lynch_text
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].border = border
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'VERY UNDERVALUED' in lynch_status.upper() or lynch_delta > 20:
            ws[f'B{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in lynch_status.upper() or lynch_delta > 5:
            ws[f'B{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in lynch_status.upper() or abs(lynch_delta) <= 5:
            ws[f'B{row}'].fill = hold_fill
        elif 'OVERVALUED' in lynch_status.upper() or lynch_delta < -5:
            ws[f'B{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in lynch_status.upper() or lynch_delta < -20:
            ws[f'B{row}'].fill = strong_sell_fill
    else:
        ws[f'B{row}'] = "Insufficient Data"
        ws[f'B{row}'].font = Font(size=11, color="808080")
        ws[f'B{row}'].fill = no_data_fill
        ws[f'B{row}'].border = border
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # DCF Valuation (Column C)
    dcf_status = safe_get_status(stock_row, 'dcf_valuation_status')
    dcf_delta = safe_get_delta(stock_row, 'dcf_delta_percentage')
    if dcf_status != 'N/A':
        dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
        ws[f'C{row}'] = dcf_text
        ws[f'C{row}'].font = data_font
        ws[f'C{row}'].border = border
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
            ws[f'C{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
            ws[f'C{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
            ws[f'C{row}'].fill = hold_fill
        elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
            ws[f'C{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
            ws[f'C{row}'].fill = strong_sell_fill
    else:
        ws[f'C{row}'] = "Insufficient Data"
        ws[f'C{row}'].font = Font(size=11, color="808080")
        ws[f'C{row}'].fill = no_data_fill
        ws[f'C{row}'].border = border
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Munger Farm (Column D)
    munger_assessment = safe_get_status(stock_row, 'munger_7pct_assessment')
    munger_delta = safe_get_delta(stock_row, 'munger_7pct_delta_percentage')
    if munger_assessment != 'N/A':
        munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
        ws[f'D{row}'] = munger_text
        ws[f'D{row}'].font = data_font
        ws[f'D{row}'].border = border
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
            ws[f'D{row}'].fill = strong_buy_fill
        elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
            ws[f'D{row}'].fill = buy_fill
        elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
            ws[f'D{row}'].fill = hold_fill
        elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
            ws[f'D{row}'].fill = sell_fill
        elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
            ws[f'D{row}'].fill = strong_sell_fill
    else:
        ws[f'D{row}'] = "Insufficient Data"
        ws[f'D{row}'].font = Font(size=11, color="808080")
        ws[f'D{row}'].fill = no_data_fill
        ws[f'D{row}'].border = border
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Enhanced DCF (Column E)
    enhanced_dcf_status = safe_get_status(stock_row, 'enhanced_dcf_status')
    enhanced_dcf_delta = safe_get_delta(stock_row, 'enhanced_dcf_delta')
    if enhanced_dcf_status != 'N/A':
        enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
        ws[f'E{row}'] = enhanced_dcf_text
        ws[f'E{row}'].font = data_font
        ws[f'E{row}'].border = border
        ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
            ws[f'E{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
            ws[f'E{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
            ws[f'E{row}'].fill = hold_fill
        elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
            ws[f'E{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
            ws[f'E{row}'].fill = strong_sell_fill
    else:
        ws[f'E{row}'] = "Insufficient Data"
        ws[f'E{row}'].font = Font(size=11, color="808080")
        ws[f'E{row}'].fill = no_data_fill
        ws[f'E{row}'].border = border
        ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Relative Valuation (Column F)
    relative_status = safe_get_status(stock_row, 'relative_valuation_status')
    relative_delta = safe_get_delta(stock_row, 'relative_valuation_delta')
    if relative_status != 'N/A':
        relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
        ws[f'F{row}'] = relative_text
        ws[f'F{row}'].font = data_font
        ws[f'F{row}'].border = border
        ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
            ws[f'F{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
            ws[f'F{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
            ws[f'F{row}'].fill = hold_fill
        elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
            ws[f'F{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
            ws[f'F{row}'].fill = strong_sell_fill
    else:
        ws[f'F{row}'] = "Insufficient Data"
        ws[f'F{row}'].font = Font(size=11, color="808080")
        ws[f'F{row}'].fill = no_data_fill
        ws[f'F{row}'].border = border
        ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Reverse DCF (Column G)
    reverse_dcf_assessment = safe_get_status(stock_row, 'reverse_dcf_assessment')
    reverse_dcf_reasonable = stock_row.get('reverse_dcf_reasonable', None)
    if reverse_dcf_assessment != 'N/A':
        reasonable_text = 'Yes' if reverse_dcf_reasonable else 'No'
        reverse_dcf_text = f"{reverse_dcf_assessment}\nReasonable: {reasonable_text}"
        ws[f'G{row}'] = reverse_dcf_text
        ws[f'G{row}'].font = data_font
        ws[f'G{row}'].border = border
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'REASONABLE' in reverse_dcf_assessment.upper() or reverse_dcf_reasonable:
            ws[f'G{row}'].fill = hold_fill
        elif 'UNREASONABLE' in reverse_dcf_assessment.upper() or not reverse_dcf_reasonable:
            ws[f'G{row}'].fill = sell_fill
    else:
        ws[f'G{row}'] = "Insufficient Data"
        ws[f'G{row}'].font = Font(size=11, color="808080")
        ws[f'G{row}'].fill = no_data_fill
        ws[f'G{row}'].border = border
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # EPV/RIM (Column H)
    epv_assessment = safe_get_status(stock_row, 'epv_assessment')
    epv_delta = safe_get_delta(stock_row, 'epv_delta')
    rim_assessment = safe_get_status(stock_row, 'rim_assessment')
    rim_delta = safe_get_delta(stock_row, 'rim_delta')
    
    if epv_assessment != 'N/A':
        epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
        ws[f'H{row}'] = epv_rim_text
        ws[f'H{row}'].font = data_font
        ws[f'H{row}'].border = border
        ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
            ws[f'H{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
            ws[f'H{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
            ws[f'H{row}'].fill = hold_fill
        elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
            ws[f'H{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
            ws[f'H{row}'].fill = strong_sell_fill
    elif rim_assessment != 'N/A':
        epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
        ws[f'H{row}'] = epv_rim_text
        ws[f'H{row}'].font = data_font
        ws[f'H{row}'].border = border
        ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
            ws[f'H{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
            ws[f'H{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
            ws[f'H{row}'].fill = hold_fill
        elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
            ws[f'H{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
            ws[f'H{row}'].fill = strong_sell_fill
    else:
        ws[f'H{row}'] = "Insufficient Data"
        ws[f'H{row}'].font = Font(size=11, color="808080")
        ws[f'H{row}'].fill = no_data_fill
        ws[f'H{row}'].border = border
        ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Current Price (Column I)
    current_price = safe_get_delta(stock_row, 'current_price')
    if current_price and current_price > 0:
        ws[f'I{row}'] = f"${current_price:.2f}"
        ws[f'I{row}'].font = data_font
        ws[f'I{row}'].border = border
        ws[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
    else:
        ws[f'I{row}'] = "N/A"
        ws[f'I{row}'].font = Font(size=11, color="808080")
        ws[f'I{row}'].fill = no_data_fill
        ws[f'I{row}'].border = border
        ws[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')

if __name__ == "__main__":
    fix_excel_formatting()
