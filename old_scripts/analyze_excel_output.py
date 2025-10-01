import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

def analyze_excel_output(excel_file):
    """Analyze the Excel output to check validity of valuation values"""
    try:
        print("=== STOCK VALUATION EXCEL OUTPUT ANALYSIS ===")
        print(f"Analyzing file: {excel_file}")
        
        # Load the workbook to check sheets
        wb = load_workbook(excel_file)
        print(f"\nAvailable sheets: {wb.sheetnames}")
        
        # Read the Valuation Data sheet
        df = pd.read_excel(excel_file, sheet_name='Valuation Data')
        print(f"\nTotal records in Valuation Data: {len(df)}")
        
        # Get latest data for each ticker
        latest_data = df.sort_values(by='timestamp').groupby('ticker').last().reset_index()
        print(f"Unique tickers: {len(latest_data)}")
        
        print("\n=== VALUATION METHOD VALIDITY CHECK ===")
        
        # Check Peter Lynch valuations
        print("\n1. PETER LYNCH VALUATION VALIDITY:")
        lynch_data = latest_data[latest_data['lynch_valuation_status'].notna()]
        print(f"   Records with Lynch data: {len(lynch_data)}/{len(latest_data)} ({len(lynch_data)/len(latest_data)*100:.1f}%)")
        
        # Check for extreme values (>1000% delta)
        extreme_lynch = lynch_data[abs(lynch_data['lynch_delta_percentage']) > 1000]
        print(f"   Extreme valuations (>1000%): {len(extreme_lynch)}")
        if len(extreme_lynch) > 0:
            print("   Sample extreme cases:")
            for _, stock in extreme_lynch.head(5).iterrows():
                print(f"      {stock['ticker']}: {stock['lynch_delta_percentage']:+.1f}% (Lynch Ratio: {stock.get('lynch_ratio', 'N/A')})")
        
        # Check for reasonable valuations (between -50% and +200%)
        reasonable_lynch = lynch_data[(lynch_data['lynch_delta_percentage'] >= -50) & (lynch_data['lynch_delta_percentage'] <= 200)]
        print(f"   Reasonable valuations (-50% to +200%): {len(reasonable_lynch)}")
        
        # Check DCF valuations
        print("\n2. DCF VALUATION VALIDITY:")
        dcf_data = latest_data[latest_data['dcf_valuation_status'].notna()]
        print(f"   Records with DCF data: {len(dcf_data)}/{len(latest_data)} ({len(dcf_data)/len(latest_data)*100:.1f}%)")
        
        # Check for extreme values
        extreme_dcf = dcf_data[abs(dcf_data['dcf_delta_percentage']) > 500]
        print(f"   Extreme valuations (>500%): {len(extreme_dcf)}")
        if len(extreme_dcf) > 0:
            print("   Sample extreme cases:")
            for _, stock in extreme_dcf.head(5).iterrows():
                print(f"      {stock['ticker']}: {stock['dcf_delta_percentage']:+.1f}%")
        
        # Check Enhanced DCF valuations
        print("\n3. ENHANCED DCF VALUATION VALIDITY:")
        enhanced_dcf_data = latest_data[latest_data['enhanced_dcf_status'].notna()]
        print(f"   Records with Enhanced DCF data: {len(enhanced_dcf_data)}/{len(latest_data)} ({len(enhanced_dcf_data)/len(latest_data)*100:.1f}%)")
        
        if len(enhanced_dcf_data) > 0:
            extreme_enhanced_dcf = enhanced_dcf_data[abs(enhanced_dcf_data['enhanced_dcf_delta']) > 200]
            print(f"   Extreme valuations (>200%): {len(extreme_enhanced_dcf)}")
        
        # Check Relative Valuation
        print("\n4. RELATIVE VALUATION VALIDITY:")
        relative_data = latest_data[latest_data['relative_valuation_status'].notna()]
        print(f"   Records with Relative Valuation data: {len(relative_data)}/{len(latest_data)} ({len(relative_data)/len(latest_data)*100:.1f}%)")
        
        # Check Reverse DCF
        print("\n5. REVERSE DCF VALIDITY:")
        reverse_dcf_data = latest_data[latest_data['reverse_dcf_assessment'].notna()]
        print(f"   Records with Reverse DCF data: {len(reverse_dcf_data)}/{len(latest_data)} ({len(reverse_dcf_data)/len(latest_data)*100:.1f}%)")
        
        # Check EPV/RIM
        print("\n6. EPV/RIM VALIDITY:")
        epv_data = latest_data[latest_data['epv_assessment'].notna()]
        rim_data = latest_data[latest_data['rim_assessment'].notna()]
        print(f"   Records with EPV data: {len(epv_data)}/{len(latest_data)} ({len(epv_data)/len(latest_data)*100:.1f}%)")
        print(f"   Records with RIM data: {len(rim_data)}/{len(latest_data)} ({len(rim_data)/len(latest_data)*100:.1f}%)")
        
        print("\n=== AMERICAN STOCKS ANALYSIS ===")
        american_stocks = latest_data[latest_data['ticker'].apply(lambda x: not x.endswith('.NZ') and not x.endswith('.AX'))]
        print(f"Found {len(american_stocks)} American stocks")
        
        if len(american_stocks) > 0:
            print("\nAmerican stocks valuation status:")
            for _, stock in american_stocks.iterrows():
                print(f"\n--- {stock['ticker']} ({stock['company_name']}) ---")
                print(f"Peter Lynch: {stock.get('lynch_valuation_status', 'N/A')} ({stock.get('lynch_delta_percentage', 0):+.1f}%)")
                print(f"DCF: {stock.get('dcf_valuation_status', 'N/A')} ({stock.get('dcf_delta_percentage', 0):+.1f}%)")
                print(f"Munger Farm: {stock.get('munger_7pct_assessment', 'N/A')} ({stock.get('munger_7pct_delta_percentage', 0):+.1f}%)")
                print(f"Enhanced DCF: {stock.get('enhanced_dcf_status', 'N/A')} ({stock.get('enhanced_dcf_delta', 0):+.1f}%)")
                print(f"Relative Valuation: {stock.get('relative_valuation_status', 'N/A')} ({stock.get('relative_valuation_delta', 0):+.1f}%)")
                print(f"Reverse DCF: {stock.get('reverse_dcf_assessment', 'N/A')} ({stock.get('reverse_dcf_implied_growth', 0):+.1f}%)")
                print(f"EPV: {stock.get('epv_assessment', 'N/A')} ({stock.get('epv_delta', 0):+.1f}%)")
                print(f"RIM: {stock.get('rim_assessment', 'N/A')} ({stock.get('rim_delta', 0):+.1f}%)")
        
        print("\n=== EXCEL SHEET FORMATTING CHECK ===")
        
        # Check My Portfolio sheet
        if 'My Portfolio' in wb.sheetnames:
            ws_portfolio = wb['My Portfolio']
            print(f"\nMy Portfolio sheet:")
            print(f"   Max rows: {ws_portfolio.max_row}")
            print(f"   Max columns: {ws_portfolio.max_column}")
            
            # Check headers
            headers = [cell.value for cell in ws_portfolio[1]]
            print(f"   Headers: {headers}")
            
            # Check if all 9 columns are populated
            if ws_portfolio.max_column >= 9:
                print("   ✓ All 9 columns present")
            else:
                print(f"   ✗ Only {ws_portfolio.max_column} columns present (expected 9)")
        
        # Check Prospects sheet
        if 'Prospects' in wb.sheetnames:
            ws_prospects = wb['Prospects']
            print(f"\nProspects sheet:")
            print(f"   Max rows: {ws_prospects.max_row}")
            print(f"   Max columns: {ws_prospects.max_column}")
            
            # Check headers
            headers = [cell.value for cell in ws_prospects[1]]
            print(f"   Headers: {headers}")
            
            # Check if all 9 columns are populated
            if ws_prospects.max_column >= 9:
                print("   ✓ All 9 columns present")
            else:
                print(f"   ✗ Only {ws_prospects.max_column} columns present (expected 9)")
        
        print("\n=== SUMMARY ===")
        print(f"Total stocks analyzed: {len(latest_data)}")
        print(f"Peter Lynch extreme valuations: {len(extreme_lynch)}")
        print(f"DCF extreme valuations: {len(extreme_dcf)}")
        print(f"American stocks with data: {len(american_stocks)}")
        
        # Check for any NaN or invalid values
        nan_columns = latest_data.columns[latest_data.isnull().any()].tolist()
        if nan_columns:
            print(f"\nColumns with missing data: {nan_columns}")
        
        return True
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")
        return False

if __name__ == "__main__":
    excel_file_path = r"C:\Users\james\Downloads\Stock Valuation\stock_valuation_dataset.xlsx"
    analyze_excel_output(excel_file_path)
