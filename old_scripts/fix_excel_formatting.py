import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def fix_excel_formatting():
    """Fix Excel file to include all valuation methods with proper formatting"""
    excel_file = r"C:\Users\james\Downloads\Stock Valuation\stock_valuation_dataset.xlsx"
    
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        print("=== FIXING EXCEL FORMATTING ===")
        
        # Remove existing sheets
        if 'My Portfolio' in wb.sheetnames:
            wb.remove(wb['My Portfolio'])
        if 'Prospects' in wb.sheetnames:
            wb.remove(wb['Prospects'])
        
        # Load the latest data
        df = pd.read_excel(excel_file, sheet_name='Valuation Data')
        latest_data = df.sort_values(by='timestamp').groupby('ticker').last().reset_index()
        
        print(f"Loaded {len(latest_data)} stocks")
        
        # Create new My Portfolio sheet with all 9 columns
        create_comprehensive_portfolio_sheet(wb, latest_data)
        
        # Create new Prospects sheet with all 9 columns
        create_comprehensive_prospects_sheet(wb, latest_data)
        
        # Save the updated workbook
        wb.save(excel_file)
        print(f"Successfully updated Excel file: {excel_file}")
        print("Now includes all Tier 1 & Tier 2 valuation methods:")
        print("   • Enhanced DCF with Scenario Analysis")
        print("   • Relative Valuation (EV/EBITDA, P/E, P/S, P/B)")
        print("   • Reverse DCF (What's Priced In?)")
        print("   • Earnings Power Value (EPV)")
        print("   • Residual Income Model (RIM)")
        
    except Exception as e:
        print(f"Error fixing Excel file: {e}")

def create_comprehensive_portfolio_sheet(wb, df):
    """Create comprehensive My Portfolio sheet with all valuation methods"""
    ws = wb.create_sheet('My Portfolio', 0)
    
    # Define styles
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    subheader_font = Font(bold=True, size=12)
    data_font = Font(size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                  top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Consistent conditional formatting fills
    strong_buy_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    buy_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
    hold_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    sell_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    strong_sell_fill = PatternFill(start_color="FFA0A0", end_color="FFA0A0", fill_type="solid")
    no_data_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Title
    ws['A1'] = "Stock Valuation Analysis Dashboard"
    ws['A1'].font = Font(bold=True, size=16, color="2F4F4F")
    ws.merge_cells('A1:I1')
    
    # Analysis info
    ws['A3'] = f"Analysis Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A4'] = f"Stocks Analyzed: {len(df)}"
    
    # Headers for all 9 columns
    row = 6
    headers = [
        "Company (Ticker)",
        "Peter Lynch", 
        "DCF Valuation",
        "Munger Farm",
        "Enhanced DCF",
        "Relative Valuation", 
        "Reverse DCF",
        "EPV/RIM",
        "Current Price"
    ]
    
    for i, header in enumerate(headers):
        col_letter = chr(65 + i)  # A, B, C, etc.
        ws[f'{col_letter}{row}'] = header
        cell = ws[f'{col_letter}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Categorize stocks by completeness
    stocks_complete = []
    stocks_partial = []
    stocks_without_data = []
    
    for _, stock_row in df.iterrows():
        completeness_score = 0
        if stock_row.get('lynch_valuation_status', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('dcf_valuation_status', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('munger_7pct_assessment', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('enhanced_dcf_status', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('relative_valuation_status', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('reverse_dcf_assessment', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('epv_assessment', 'N/A') != 'N/A' or stock_row.get('rim_assessment', 'N/A') != 'N/A':
            completeness_score += 1
        
        if completeness_score >= 5:
            stocks_complete.append(stock_row)
        elif completeness_score >= 2:
            stocks_partial.append(stock_row)
        else:
            stocks_without_data.append(stock_row)
    
    # Sort complete stocks by undervaluation score
    def calculate_undervaluation_score(stock_row):
        score = 0
        if stock_row.get('lynch_delta_percentage', 0) > 0:
            score += stock_row.get('lynch_delta_percentage', 0) * 0.3
        if stock_row.get('dcf_delta_percentage', 0) > 0:
            score += stock_row.get('dcf_delta_percentage', 0) * 0.3
        if stock_row.get('munger_7pct_delta_percentage', 0) > 0:
            score += stock_row.get('munger_7pct_delta_percentage', 0) * 0.2
        if stock_row.get('enhanced_dcf_delta', 0) > 0:
            score += stock_row.get('enhanced_dcf_delta', 0) * 0.1
        if stock_row.get('relative_valuation_delta', 0) > 0:
            score += stock_row.get('relative_valuation_delta', 0) * 0.1
        return score
    
    stocks_complete.sort(key=calculate_undervaluation_score, reverse=True)
    stocks_partial.sort(key=calculate_undervaluation_score, reverse=True)
    
    # Add complete stocks
    for stock_row in stocks_complete:
        add_stock_row(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill)
        row += 1
    
    # Add partial stocks
    if stocks_partial:
        row += 1
        ws[f'A{row}'] = "STOCKS WITH PARTIAL DATA"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        row += 1
        
        for stock_row in stocks_partial:
            add_stock_row(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill)
            row += 1
    
    # Add stocks without data
    if stocks_without_data:
        row += 1
        ws[f'A{row}'] = "STOCKS WITH INSUFFICIENT DATA"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        row += 1
        
        for stock_row in stocks_without_data:
            add_stock_row(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill)
            row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 45
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 35

def create_comprehensive_prospects_sheet(wb, df):
    """Create comprehensive Prospects sheet with all valuation methods"""
    ws = wb.create_sheet('Prospects', 1)
    
    # Define styles
    header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    subheader_font = Font(bold=True, size=12)
    data_font = Font(size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                  top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Consistent conditional formatting fills
    strong_buy_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    buy_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
    hold_fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    sell_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    strong_sell_fill = PatternFill(start_color="FFA0A0", end_color="FFA0A0", fill_type="solid")
    no_data_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Title
    ws['A1'] = "NZX PROSPECTS - RANKED BY UNDERVALUATION"
    ws['A1'].font = Font(bold=True, size=16, color="2F4F4F")
    ws.merge_cells('A1:I1')
    
    # Headers for all 9 columns
    row = 3
    headers = [
        "Company (Ticker)",
        "Peter Lynch", 
        "DCF Valuation",
        "Munger Farm",
        "Enhanced DCF",
        "Relative Valuation", 
        "Reverse DCF",
        "EPV/RIM",
        "Current Price"
    ]
    
    for i, header in enumerate(headers):
        col_letter = chr(65 + i)  # A, B, C, etc.
        ws[f'{col_letter}{row}'] = header
        cell = ws[f'{col_letter}{row}']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row += 1
    
    # Calculate undervaluation scores and sort
    def calculate_undervaluation_score(stock_row):
        score = 0
        if stock_row.get('lynch_delta_percentage', 0) > 0:
            score += stock_row.get('lynch_delta_percentage', 0) * 0.3
        if stock_row.get('dcf_delta_percentage', 0) > 0:
            score += stock_row.get('dcf_delta_percentage', 0) * 0.3
        if stock_row.get('munger_7pct_delta_percentage', 0) > 0:
            score += stock_row.get('munger_7pct_delta_percentage', 0) * 0.2
        if stock_row.get('enhanced_dcf_delta', 0) > 0:
            score += stock_row.get('enhanced_dcf_delta', 0) * 0.1
        if stock_row.get('relative_valuation_delta', 0) > 0:
            score += stock_row.get('relative_valuation_delta', 0) * 0.1
        return score
    
    # Categorize stocks
    stocks_complete = []
    stocks_partial = []
    stocks_without_data = []
    
    for _, stock_row in df.iterrows():
        completeness_score = 0
        if stock_row.get('lynch_valuation_status', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('dcf_valuation_status', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('munger_7pct_assessment', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('enhanced_dcf_status', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('relative_valuation_status', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('reverse_dcf_assessment', 'N/A') != 'N/A':
            completeness_score += 1
        if stock_row.get('epv_assessment', 'N/A') != 'N/A' or stock_row.get('rim_assessment', 'N/A') != 'N/A':
            completeness_score += 1
        
        if completeness_score >= 5:
            stocks_complete.append(stock_row)
        elif completeness_score >= 2:
            stocks_partial.append(stock_row)
        else:
            stocks_without_data.append(stock_row)
    
    # Sort by undervaluation score
    stocks_complete.sort(key=calculate_undervaluation_score, reverse=True)
    stocks_partial.sort(key=calculate_undervaluation_score, reverse=True)
    
    # Add complete stocks
    for stock_row in stocks_complete:
        add_stock_row(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill)
        row += 1
    
    # Add partial stocks
    if stocks_partial:
        row += 1
        ws[f'A{row}'] = "STOCKS WITH PARTIAL DATA"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        row += 1
        
        for stock_row in stocks_partial:
            add_stock_row(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill)
            row += 1
    
    # Add stocks without data
    if stocks_without_data:
        row += 1
        ws[f'A{row}'] = "STOCKS WITH INSUFFICIENT DATA"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        row += 1
        
        for stock_row in stocks_without_data:
            add_stock_row(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill)
            row += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 45
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 35

def add_stock_row(ws, stock_row, row, data_font, border, strong_buy_fill, buy_fill, hold_fill, sell_fill, strong_sell_fill, no_data_fill):
    """Add a stock row with all 9 columns"""
    
    # Company name (Column A)
    company_name = stock_row.get('company_name', stock_row.get('ticker', 'Unknown'))
    ticker = stock_row.get('ticker', 'Unknown')
    ws[f'A{row}'] = f"{company_name} ({ticker})"
    ws[f'A{row}'].font = data_font
    ws[f'A{row}'].border = border
    
    # Peter Lynch (Column B)
    lynch_status = stock_row.get('lynch_valuation_status', 'N/A')
    lynch_delta = stock_row.get('lynch_delta_percentage', 0)
    if lynch_status and lynch_status != 'N/A':
        lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
        ws[f'B{row}'] = lynch_text
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].border = border
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if lynch_status and 'VERY UNDERVALUED' in lynch_status.upper() or lynch_delta > 20:
            ws[f'B{row}'].fill = strong_buy_fill
        elif lynch_status and 'UNDERVALUED' in lynch_status.upper() or lynch_delta > 5:
            ws[f'B{row}'].fill = buy_fill
        elif lynch_status and 'FAIRLY VALUED' in lynch_status.upper() or abs(lynch_delta) <= 5:
            ws[f'B{row}'].fill = hold_fill
        elif lynch_status and 'OVERVALUED' in lynch_status.upper() or lynch_delta < -5:
            ws[f'B{row}'].fill = sell_fill
        elif lynch_status and 'SIGNIFICANTLY OVERVALUED' in lynch_status.upper() or lynch_delta < -20:
            ws[f'B{row}'].fill = strong_sell_fill
    else:
        ws[f'B{row}'] = "Insufficient Data"
        ws[f'B{row}'].font = Font(size=11, color="808080")
        ws[f'B{row}'].fill = no_data_fill
        ws[f'B{row}'].border = border
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # DCF Valuation (Column C)
    dcf_status = stock_row.get('dcf_valuation_status', 'N/A')
    dcf_delta = stock_row.get('dcf_delta_percentage', 0)
    if dcf_status != 'N/A':
        dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
        ws[f'C{row}'] = dcf_text
        ws[f'C{row}'].font = data_font
        ws[f'C{row}'].border = border
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
            ws[f'C{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
            ws[f'C{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
            ws[f'C{row}'].fill = hold_fill
        elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
            ws[f'C{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
            ws[f'C{row}'].fill = strong_sell_fill
    else:
        ws[f'C{row}'] = "Insufficient Data"
        ws[f'C{row}'].font = Font(size=11, color="808080")
        ws[f'C{row}'].fill = no_data_fill
        ws[f'C{row}'].border = border
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Munger Farm (Column D)
    munger_assessment = stock_row.get('munger_7pct_assessment', 'N/A')
    munger_delta = stock_row.get('munger_7pct_delta_percentage', 0)
    if munger_assessment != 'N/A':
        munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
        ws[f'D{row}'] = munger_text
        ws[f'D{row}'].font = data_font
        ws[f'D{row}'].border = border
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
            ws[f'D{row}'].fill = strong_buy_fill
        elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
            ws[f'D{row}'].fill = buy_fill
        elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
            ws[f'D{row}'].fill = hold_fill
        elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
            ws[f'D{row}'].fill = sell_fill
        elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
            ws[f'D{row}'].fill = strong_sell_fill
    else:
        ws[f'D{row}'] = "Insufficient Data"
        ws[f'D{row}'].font = Font(size=11, color="808080")
        ws[f'D{row}'].fill = no_data_fill
        ws[f'D{row}'].border = border
        ws[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Enhanced DCF (Column E)
    enhanced_dcf_status = stock_row.get('enhanced_dcf_status', 'N/A')
    enhanced_dcf_delta = stock_row.get('enhanced_dcf_delta', 0)
    if enhanced_dcf_status != 'N/A':
        enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
        ws[f'E{row}'] = enhanced_dcf_text
        ws[f'E{row}'].font = data_font
        ws[f'E{row}'].border = border
        ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
            ws[f'E{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
            ws[f'E{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
            ws[f'E{row}'].fill = hold_fill
        elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
            ws[f'E{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
            ws[f'E{row}'].fill = strong_sell_fill
    else:
        ws[f'E{row}'] = "Insufficient Data"
        ws[f'E{row}'].font = Font(size=11, color="808080")
        ws[f'E{row}'].fill = no_data_fill
        ws[f'E{row}'].border = border
        ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Relative Valuation (Column F)
    relative_status = stock_row.get('relative_valuation_status', 'N/A')
    relative_delta = stock_row.get('relative_valuation_delta', 0)
    if relative_status != 'N/A':
        relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
        ws[f'F{row}'] = relative_text
        ws[f'F{row}'].font = data_font
        ws[f'F{row}'].border = border
        ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
            ws[f'F{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
            ws[f'F{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
            ws[f'F{row}'].fill = hold_fill
        elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
            ws[f'F{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
            ws[f'F{row}'].fill = strong_sell_fill
    else:
        ws[f'F{row}'] = "Insufficient Data"
        ws[f'F{row}'].font = Font(size=11, color="808080")
        ws[f'F{row}'].fill = no_data_fill
        ws[f'F{row}'].border = border
        ws[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Reverse DCF (Column G)
    reverse_dcf_assessment = stock_row.get('reverse_dcf_assessment', 'N/A')
    reverse_dcf_reasonable = stock_row.get('reverse_dcf_reasonable', None)
    if reverse_dcf_assessment != 'N/A':
        reverse_dcf_text = f"{reverse_dcf_assessment}\nReasonable: {'Yes' if reverse_dcf_reasonable else 'No'}"
        ws[f'G{row}'] = reverse_dcf_text
        ws[f'G{row}'].font = data_font
        ws[f'G{row}'].border = border
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'REASONABLE' in reverse_dcf_assessment.upper() or reverse_dcf_reasonable:
            ws[f'G{row}'].fill = hold_fill
        elif 'UNREASONABLE' in reverse_dcf_assessment.upper() or not reverse_dcf_reasonable:
            ws[f'G{row}'].fill = sell_fill
    else:
        ws[f'G{row}'] = "Insufficient Data"
        ws[f'G{row}'].font = Font(size=11, color="808080")
        ws[f'G{row}'].fill = no_data_fill
        ws[f'G{row}'].border = border
        ws[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # EPV/RIM (Column H)
    epv_assessment = stock_row.get('epv_assessment', 'N/A')
    epv_delta = stock_row.get('epv_delta', 0)
    rim_assessment = stock_row.get('rim_assessment', 'N/A')
    rim_delta = stock_row.get('rim_delta', 0)
    
    if epv_assessment != 'N/A':
        epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
        ws[f'H{row}'] = epv_rim_text
        ws[f'H{row}'].font = data_font
        ws[f'H{row}'].border = border
        ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
            ws[f'H{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
            ws[f'H{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
            ws[f'H{row}'].fill = hold_fill
        elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
            ws[f'H{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
            ws[f'H{row}'].fill = strong_sell_fill
    elif rim_assessment != 'N/A':
        epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
        ws[f'H{row}'] = epv_rim_text
        ws[f'H{row}'].font = data_font
        ws[f'H{row}'].border = border
        ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply conditional formatting
        if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
            ws[f'H{row}'].fill = strong_buy_fill
        elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
            ws[f'H{row}'].fill = buy_fill
        elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
            ws[f'H{row}'].fill = hold_fill
        elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
            ws[f'H{row}'].fill = sell_fill
        elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
            ws[f'H{row}'].fill = strong_sell_fill
    else:
        ws[f'H{row}'] = "Insufficient Data"
        ws[f'H{row}'].font = Font(size=11, color="808080")
        ws[f'H{row}'].fill = no_data_fill
        ws[f'H{row}'].border = border
        ws[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Current Price (Column I)
    current_price = stock_row.get('current_price', 0)
    if current_price and current_price > 0:
        ws[f'I{row}'] = f"${current_price:.2f}"
        ws[f'I{row}'].font = data_font
        ws[f'I{row}'].border = border
        ws[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
    else:
        ws[f'I{row}'] = "N/A"
        ws[f'I{row}'].font = Font(size=11, color="808080")
        ws[f'I{row}'].fill = no_data_fill
        ws[f'I{row}'].border = border
        ws[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')

if __name__ == "__main__":
    fix_excel_formatting()
