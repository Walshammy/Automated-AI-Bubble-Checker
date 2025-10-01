import pandas as pd
import numpy as np
import yfinance as yf
import requests
from bs4 import BeautifulSoup
import time
import os
from datetime import datetime, timedelta
import logging
import json
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import functools
from typing import Dict, Optional, Tuple, List
from tqdm import tqdm

warnings.filterwarnings('ignore')

# Configuration constants
VALUATION_CONFIG = {
    'discount_rate': 0.08,
    'perpetual_growth_rate': 0.02,  # Reduced from 0.025 to 0.02 (2%)
    'projection_years': 10,
    'default_fcf_growth': 0.08,  # More conservative than 0.15
    'max_fcf_growth': 0.15,     # Cap at 15%
    'min_fcf_growth': 0.02,     # Minimum 2%
    'lynch_thresholds': {
        'VERY_UNDERVALUED': 2.0,
        'UNDERVALUED': 1.5,
        'FAIRLY_VALUED': 1.0,
        'OVERVALUED': 0.0
    },
    'api_delay': 0.5,  # Increased from 0.2 to 0.5 (500ms)
    'batch_size': 20,   # Process in batches
    'batch_delay': 5.0,  # 5 second pause between batches
    'max_terminal_value_ratio': 0.8,  # Terminal value shouldn't exceed 80% of total
    
    # Historical data collection settings
    'historical_start_year': 2000,
    'historical_batch_delay': 10.0,  # 10 seconds between batches for historical
    'max_historical_period': '25y',  # Go back 25 years
    'quarterly_data_enabled': True,
    'historical_batch_size': 5,  # Smaller batches for historical data
    'historical_checkpoint_interval': 10  # Save progress every 10 stocks
}

def rate_limit(delay=VALUATION_CONFIG['api_delay']):
    """Rate limiting decorator for API calls"""
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            time.sleep(delay)
            return func(*args, **kwargs)
        return wrapper
    return decorator

class StockValuationScraper:
    def __init__(self):
        # Setup logging
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger(__name__)

        # OneDrive directory (primary storage location) - Use environment variable
        self.onedrive_dir = os.path.join(
            os.environ.get('ONEDRIVE', os.path.expanduser('~/OneDrive')),
            'StockValuation'
        )
        os.makedirs(self.onedrive_dir, exist_ok=True)

        # Second storage location - Downloads/Stock Valuation
        # Primary path
        primary_path = r"C:\Users\james\Downloads\Stock Valuation"
        if os.path.exists(os.path.dirname(primary_path)):
            self.downloads_dir = primary_path
        else:
            # Fallback: search for james.walsham in Users directory
            users_dir = r"C:\Users"
            if os.path.exists(users_dir):
                for item in os.listdir(users_dir):
                    if "james.walsham" in item.lower():
                        fallback_path = os.path.join(users_dir, item, "Downloads", "Stock Valuation")
                        if os.path.exists(os.path.dirname(fallback_path)):
                            self.downloads_dir = fallback_path
                            break
                else:
                    self.downloads_dir = None
            else:
                self.downloads_dir = None
        
        if self.downloads_dir:
            os.makedirs(self.downloads_dir, exist_ok=True)

        # Daily backups directory
        self.daily_backups_dir = os.path.join(self.onedrive_dir, "daily_backups")
        os.makedirs(self.daily_backups_dir, exist_ok=True)

        # Master dataset files
        self.master_file = os.path.join(self.onedrive_dir, "stock_valuation_dataset.xlsx")
        self.downloads_file = os.path.join(self.downloads_dir, "stock_valuation_dataset.xlsx") if self.downloads_dir else None
        
        # Historical dataset file
        self.historical_file = os.path.join(self.onedrive_dir, "stock_valuation_historical.xlsx")
        
        # Combined dataset file (historical + current)
        self.combined_file = os.path.join(self.onedrive_dir, "stock_valuation_combined.xlsx")
        
        # Historical progress tracking
        self.historical_progress_file = os.path.join(self.onedrive_dir, "valuation_historical_progress.json")

        # Market indices for context
        self.indices = {
            '^GSPC': 'S&P 500',
            '^IXIC': 'NASDAQ',
            '^VIX': 'VIX',
            '^TNX': '10-Year Treasury'
        }

        # Valuation parameters from configuration
        self.discount_rate = VALUATION_CONFIG['discount_rate']
        self.perpetual_growth_rate = VALUATION_CONFIG['perpetual_growth_rate']
        self.projection_years = VALUATION_CONFIG['projection_years']

        # Focus stocks for valuation analysis - Complete NZX + International
        self.focus_stocks = {
            # NZX Mega/Large Cap (>$5B) - 13 companies
            'WBC.NZ': 'Westpac Banking Corporation',
            'ANZ.NZ': 'ANZ Group Holdings Limited',
            'FPH.NZ': 'Fisher & Paykel Healthcare Corporation Limited',
            'MEL.NZ': 'Meridian Energy Limited',
            'AIA.NZ': 'Auckland International Airport Limited',
            'IFT.NZ': 'Infratil Limited',
            'AFI.NZ': 'Australian Foundation Investment Company Limited',
            'MCY.NZ': 'Mercury NZ Limited',
            'EBO.NZ': 'EBOS Group Limited',
            'FCG.NZ': 'Fonterra Co-operative Group Limited',
            'CEN.NZ': 'Contact Energy Limited',
            'MFT.NZ': 'Mainfreight Limited',
            'ATM.NZ': 'The a2 Milk Company Limited',
            
            # NZX Mid Cap ($1B - $5B) - 16 companies
            'POT.NZ': 'Port of Tauranga Limited',
            'SPK.NZ': 'Spark New Zealand Limited',
            'VNT.NZ': 'Ventia Services Group Limited',
            'VCT.NZ': 'Vector Limited',
            'CNU.NZ': 'Chorus Limited',
            'FBU.NZ': 'Fletcher Building Limited',
            'GMT.NZ': 'Goodman Property Trust',
            'SUM.NZ': 'Summerset Group Holdings Limited',
            'GNE.NZ': 'Genesis Energy Limited',
            'RYM.NZ': 'Ryman Healthcare Limited',
            'FRW.NZ': 'Freightways Group Limited',
            'MNW.NZ': 'Manawa Energy Limited',
            'PCT.NZ': 'Precinct Properties NZ Ltd',
            'AIR.NZ': 'Air New Zealand Limited',
            'KPG.NZ': 'Kiwi Property Group Limited',
            'GTK.NZ': 'Gentrack Group Limited',
            
            # NZX Small Cap ($300M - $1B) - 29 companies
            'VHP.NZ': 'Vital Healthcare Property Trust',
            'PFI.NZ': 'Property For Industry Limited',
            'BGP.NZ': 'Briscoe Group Limited',
            'ARG.NZ': 'Argosy Property Limited',
            'SKL.NZ': 'Skellerup Holdings Limited',
            'CHI.NZ': 'Channel Infrastructure NZ Limited',
            'VSL.NZ': 'Vulcan Steel Limited',
            'VGL.NZ': 'Vista Group International Limited',
            'HGH.NZ': 'Heartland Group Holdings Limited',
            'SKC.NZ': 'SkyCity Entertainment Group Limited',
            'SCL.NZ': 'Scales Corporation Limited',
            'SPG.NZ': 'Stride Property Group',
            'NPH.NZ': 'Napier Port Holdings Limited',
            'TRA.NZ': 'Turners Automotive Group Limited',
            'WIN.NZ': 'Winton Land Limited',
            'OCA.NZ': 'Oceania Healthcare Limited',
            'TWR.NZ': 'Tower Limited',
            'SAN.NZ': 'Sanford Limited',
            'HLG.NZ': 'Hallenstein Glasson Holdings Limited',
            'NZX.NZ': 'NZX Limited',
            'THL.NZ': 'Tourism Holdings Limited',
            'IPL.NZ': 'Investore Property Limited',
            'DGL.NZ': 'Delegat Group Limited',
            'MCK.NZ': 'Millennium & Copthorne Hotels NZ Limited',
            'SKT.NZ': 'SKY Network Television Limited',
            'SKO.NZ': 'Serko Limited',
            'RBD.NZ': 'Restaurant Brands New Zealand Limited',
            'SML.NZ': 'Synlait Milk Limited',
            
            # NZX Micro Cap ($100M - $300M) - 22 companies
            'WHS.NZ': 'The Warehouse Group Limited',
            'AFT.NZ': 'AFT Pharmaceuticals Limited',
            'SPY.NZ': 'Smartpay Holdings Limited',
            'ERD.NZ': 'EROAD Limited',
            'CDI.NZ': 'CDL Investments New Zealand Limited',
            'TGG.NZ': 'T&G Global Limited',
            'CMO.NZ': 'The Colonial Motor Company Limited',
            'NZM.NZ': 'NZME Limited',
            'MLN.NZ': 'Marlin Global Limited',
            'KMD.NZ': 'KMD Brands Limited',
            'SPN.NZ': 'South Port New Zealand Limited',
            'MHJ.NZ': 'Michael Hill International Limited',
            'SEK.NZ': 'Seeka Limited',
            'SCT.NZ': 'Scott Technology Limited',
            'PGW.NZ': 'PGG Wrightson Limited',
            'RAK.NZ': 'Rakon Limited',
            'IKE.NZ': 'ikeGPS Group Limited',
            'LIC.NZ': 'Livestock Improvement Corporation Limited',
            'NZL.NZ': 'New Zealand Rural Land Company Limited',
            'GXH.NZ': 'Green Cross Health Limited',
            'STU.NZ': 'Steel & Tube Holdings Limited',
            'NZK.NZ': 'New Zealand King Salmon Investments Limited',
            
            # NZX Nano Cap (<$100M) - 35 companies
            'RAD.NZ': 'Radius Residential Care Limited',
            'NWF.NZ': 'NZ Windfarms Limited',
            'PEB.NZ': 'Pacific Edge Limited',
            'BPG.NZ': 'Black Pearl Group Limited',
            'ARB.NZ': 'ArborGen Holdings Limited',
            'APL.NZ': 'Asset Plus Limited',
            'MFB.NZ': 'My Food Bag Group Limited',
            'FWL.NZ': 'Foley Wines Limited',
            'BRW.NZ': 'Bremworth Limited',
            'AOF.NZ': 'AoFrio Limited',
            'NTL.NZ': 'New Talisman Gold Mines Limited',
            'TAH.NZ': 'Third Age Health Services Limited',
            'CVT.NZ': 'Comvita Limited',
            '2CC.NZ': '2 Cheap Cars Group Limited',
            'PYS.NZ': 'PaySauce Limited',
            'TWL.NZ': 'Trade Window Holdings Limited',
            'GEN.NZ': 'General Capital Limited',
            'BLT.NZ': 'BLIS Technologies Limited',
            'MOV.NZ': 'MOVE Logistics Group Limited',
            'ALF.NZ': 'Allied Farmers Limited',
            'PHL.NZ': 'Promisia Healthcare Limited',
            'WCO.NZ': 'WasteCo Group Limited',
            'VTL.NZ': 'Vital Limited',
            'CCC.NZ': 'Cooks Coffee Company Limited',
            'TRU.NZ': 'TruScreen Group Limited',
            'ENS.NZ': 'Enprise Group Limited',
            'SVR.NZ': 'Savor Limited',
            'BFG.NZ': 'Burger Fuel Group Limited',
            'BAI.NZ': 'Being AI Limited',
            'SDL.NZ': 'Solution Dynamics Limited',
            'MPG.NZ': 'Metro Performance Glass Limited',
            'AGL.NZ': 'Accordant Group Limited',
            'CRP.NZ': 'Chatham Rock Phosphate Limited',
            'RUA.NZ': 'Rua Bioscience Limited',
            'ME.NZ': 'Me Today Limited',
            'RTO.NZ': 'RTO Limited',
            
            # International Stocks
            'BRK-B': 'Berkshire Hathaway Class B',
            'IWM': 'iShares Russell 2000 ETF',
            'MSFT': 'Microsoft Corporation',
            'META': 'Meta Platforms',
            'AAPL': 'Apple Inc.',
            'NVDA': 'NVIDIA Corporation',
            'SNOW': 'Snowflake Inc.',
            'AMZN': 'Amazon.com Inc.',
            'LMT': 'Lockheed Martin Corporation',
            'TSM': 'Taiwan Semiconductor Manufacturing',
            'INTC': 'Intel Corporation',
            'GOOGL': 'Alphabet Inc.',
            'AMD': 'Advanced Micro Devices',
            'RKLB': 'Rocket Lab USA',
            'AMAT': 'Applied Materials',
            'NVO': 'Novo Nordisk',
            'NOC': 'Northrop Grumman Corporation',
            'BRK-A': 'Berkshire Hathaway Class A',
            'SMI.AX': 'Santos Limited'
        }

    def validate_financial_data(self, data: Dict, required_fields: List[str]) -> bool:
        """Centralized validation for financial data"""
        for field in required_fields:
            value = data.get(field, 0)
            if value is None or (isinstance(value, (int, float)) and value <= 0):
                return False
        return True

    def is_market_open(self) -> bool:
        """Check if NZX/ASX/NYSE is open"""
        now = datetime.now()
        # Simple implementation - can be enhanced with actual market hours
        # For now, assume markets are open during business hours
        weekday = now.weekday()  # 0 = Monday, 6 = Sunday
        hour = now.hour
        
        # Basic check: Monday-Friday, 9 AM - 5 PM (can be enhanced)
        return weekday < 5 and 9 <= hour <= 17

    def get_data_staleness(self, timestamp: str) -> str:
        """Get data staleness indicator"""
        try:
            age = datetime.now() - pd.to_datetime(timestamp)
            if age.total_seconds() < 300:  # 5 min
                return "🟢 Live"
            elif age.total_seconds() < 3600:  # 1 hour
                return "🟡 Recent"
            else:
                return "🔴 Stale"
        except:
            return "🔴 Unknown"

    def check_valuation_consensus(self, lynch_delta, dcf_delta, munger_delta):
        """Log when valuation methods disagree significantly"""
        deltas = [d for d in [lynch_delta, dcf_delta, munger_delta] if d is not None]
        if len(deltas) >= 2:
            spread = max(deltas) - min(deltas)
            if spread > 50:  # 50% disagreement
                self.logger.warning(f"Large valuation spread: {spread:.1f}%")


        # Market indices for context
        self.indices = {
            '^GSPC': 'S&P 500',
            '^IXIC': 'NASDAQ',
            '^VIX': 'VIX',
            '^TNX': '10-Year Treasury'
        }

        # Valuation parameters from configuration
        self.discount_rate = VALUATION_CONFIG['discount_rate']
        self.perpetual_growth_rate = VALUATION_CONFIG['perpetual_growth_rate']
        self.projection_years = VALUATION_CONFIG['projection_years']

        # Focus stocks for valuation analysis - Complete NZX + International
        self.focus_stocks = {
            # NZX Mega/Large Cap (>$5B) - 13 companies
            'WBC.NZ': 'Westpac Banking Corporation',
            'ANZ.NZ': 'ANZ Group Holdings Limited',
            'FPH.NZ': 'Fisher & Paykel Healthcare Corporation Limited',
            'MEL.NZ': 'Meridian Energy Limited',
            'AIA.NZ': 'Auckland International Airport Limited',
            'IFT.NZ': 'Infratil Limited',
            'AFI.NZ': 'Australian Foundation Investment Company Limited',
            'MCY.NZ': 'Mercury NZ Limited',
            'EBO.NZ': 'EBOS Group Limited',
            'FCG.NZ': 'Fonterra Co-operative Group Limited',
            'CEN.NZ': 'Contact Energy Limited',
            'MFT.NZ': 'Mainfreight Limited',
            'ATM.NZ': 'The a2 Milk Company Limited',
            
            # NZX Mid Cap ($1B - $5B) - 16 companies
            'POT.NZ': 'Port of Tauranga Limited',
            'SPK.NZ': 'Spark New Zealand Limited',
            'VNT.NZ': 'Ventia Services Group Limited',
            'VCT.NZ': 'Vector Limited',
            'CNU.NZ': 'Chorus Limited',
            'FBU.NZ': 'Fletcher Building Limited',
            'GMT.NZ': 'Goodman Property Trust',
            'SUM.NZ': 'Summerset Group Holdings Limited',
            'GNE.NZ': 'Genesis Energy Limited',
            'RYM.NZ': 'Ryman Healthcare Limited',
            'FRW.NZ': 'Freightways Group Limited',
            'MNW.NZ': 'Manawa Energy Limited',
            'PCT.NZ': 'Precinct Properties NZ Ltd',
            'AIR.NZ': 'Air New Zealand Limited',
            'KPG.NZ': 'Kiwi Property Group Limited',
            'GTK.NZ': 'Gentrack Group Limited',
            
            # NZX Small Cap ($300M - $1B) - 29 companies
            'VHP.NZ': 'Vital Healthcare Property Trust',
            'PFI.NZ': 'Property For Industry Limited',
            'BGP.NZ': 'Briscoe Group Limited',
            'ARG.NZ': 'Argosy Property Limited',
            'SKL.NZ': 'Skellerup Holdings Limited',
            'CHI.NZ': 'Channel Infrastructure NZ Limited',
            'VSL.NZ': 'Vulcan Steel Limited',
            'VGL.NZ': 'Vista Group International Limited',
            'HGH.NZ': 'Heartland Group Holdings Limited',
            'SKC.NZ': 'SkyCity Entertainment Group Limited',
            'SCL.NZ': 'Scales Corporation Limited',
            'SPG.NZ': 'Stride Property Group',
            'NPH.NZ': 'Napier Port Holdings Limited',
            'TRA.NZ': 'Turners Automotive Group Limited',
            'WIN.NZ': 'Winton Land Limited',
            'OCA.NZ': 'Oceania Healthcare Limited',
            'TWR.NZ': 'Tower Limited',
            'SAN.NZ': 'Sanford Limited',
            'HLG.NZ': 'Hallenstein Glasson Holdings Limited',
            'NZX.NZ': 'NZX Limited',
            'THL.NZ': 'Tourism Holdings Limited',
            'IPL.NZ': 'Investore Property Limited',
            'DGL.NZ': 'Delegat Group Limited',
            'MCK.NZ': 'Millennium & Copthorne Hotels NZ Limited',
            'SKT.NZ': 'SKY Network Television Limited',
            'SKO.NZ': 'Serko Limited',
            'RBD.NZ': 'Restaurant Brands New Zealand Limited',
            'SML.NZ': 'Synlait Milk Limited',
            
            # NZX Micro Cap ($100M - $300M) - 22 companies
            'WHS.NZ': 'The Warehouse Group Limited',
            'AFT.NZ': 'AFT Pharmaceuticals Limited',
            'SPY.NZ': 'Smartpay Holdings Limited',
            'ERD.NZ': 'EROAD Limited',
            'CDI.NZ': 'CDL Investments New Zealand Limited',
            'TGG.NZ': 'T&G Global Limited',
            'CMO.NZ': 'The Colonial Motor Company Limited',
            'NZM.NZ': 'NZME Limited',
            'MLN.NZ': 'Marlin Global Limited',
            'KMD.NZ': 'KMD Brands Limited',
            'SPN.NZ': 'South Port New Zealand Limited',
            'MHJ.NZ': 'Michael Hill International Limited',
            'SEK.NZ': 'Seeka Limited',
            'SCT.NZ': 'Scott Technology Limited',
            'PGW.NZ': 'PGG Wrightson Limited',
            'RAK.NZ': 'Rakon Limited',
            'IKE.NZ': 'ikeGPS Group Limited',
            'LIC.NZ': 'Livestock Improvement Corporation Limited',
            'NZL.NZ': 'New Zealand Rural Land Company Limited',
            'GXH.NZ': 'Green Cross Health Limited',
            'STU.NZ': 'Steel & Tube Holdings Limited',
            'NZK.NZ': 'New Zealand King Salmon Investments Limited',
            
            # NZX Nano Cap (<$100M) - 35 companies
            'RAD.NZ': 'Radius Residential Care Limited',
            'NWF.NZ': 'NZ Windfarms Limited',
            'PEB.NZ': 'Pacific Edge Limited',
            'BPG.NZ': 'Black Pearl Group Limited',
            'ARB.NZ': 'ArborGen Holdings Limited',
            'APL.NZ': 'Asset Plus Limited',
            'MFB.NZ': 'My Food Bag Group Limited',
            'FWL.NZ': 'Foley Wines Limited',
            'RYM.NZ': 'Ryman Healthcare Limited',
            'RUA.NZ': 'Rua Bioscience Limited',
            'ME.NZ': 'Me Today Limited',
            'RTO.NZ': 'RTO Limited',
            
            # International Stocks
            'BRK-B': 'Berkshire Hathaway Class B',
            'IWM': 'iShares Russell 2000 ETF',
            'MSFT': 'Microsoft Corporation',
            'META': 'Meta Platforms',
            'AAPL': 'Apple Inc.',
            'NVDA': 'NVIDIA Corporation',
            'SNOW': 'Snowflake Inc.',
            'AMZN': 'Amazon.com Inc.',
            'LMT': 'Lockheed Martin Corporation',
            'TSM': 'Taiwan Semiconductor Manufacturing',
            'INTC': 'Intel Corporation',
            'GOOGL': 'Alphabet Inc.',
            'AMD': 'Advanced Micro Devices',
            'RKLB': 'Rocket Lab USA',
            'AMAT': 'Applied Materials',
            'NVO': 'Novo Nordisk',
            'NOC': 'Northrop Grumman Corporation',
            'BRK-A': 'Berkshire Hathaway Class A',
            'SMI.AX': 'Santos Limited'
        }

        # Market indices for context
        self.indices = {
            '^GSPC': 'S&P 500',
            '^IXIC': 'NASDAQ',
            '^VIX': 'VIX',
            '^TNX': '10-Year Treasury'
        }

    def _get_downloads_directory(self):
        """Get the Downloads/Stock Valuation directory with fallback to james.walsham search"""
        # Primary path
        primary_path = r"C:\Users\james\Downloads\Stock Valuation"
        if os.path.exists(os.path.dirname(primary_path)):
            return primary_path
        
        # Fallback: search for james.walsham in Users directory
        users_dir = r"C:\Users"
        if os.path.exists(users_dir):
            for item in os.listdir(users_dir):
                if "james.walsham" in item.lower():
                    fallback_path = os.path.join(users_dir, item, "Downloads", "Stock Valuation")
                    if os.path.exists(os.path.dirname(fallback_path)):
                        return fallback_path
        
        # If neither found, return None
        self.logger.warning("Could not find Downloads directory for james.walsham")
        return None

    def try_ticker_variations(self, ticker: str) -> Optional[str]:
        """Try different ticker variations to find valid data"""
        variations = [ticker, f"{ticker}.NZ", ticker.replace('.NZ', ''), ticker.replace('.AX', '')]
        
        for variant in variations:
            try:
                stock = yf.Ticker(variant)
                info = stock.info
                if info.get('regularMarketPrice') or info.get('currentPrice'):
                    self.logger.info(f"Found valid ticker: {variant} for {ticker}")
                    return variant
            except Exception as e:
                self.logger.debug(f"Ticker {variant} failed: {e}")
                continue
        
        self.logger.warning(f"No valid ticker found for {ticker}")
        return None

    def is_data_stale(self, stock_data: Dict) -> bool:
        """Check if market data is stale"""
        try:
            # For simplicity, assume data is stale if older than 15 minutes during market hours
            # In production, you'd check actual market hours
            data_age = datetime.now() - datetime.now()  # Placeholder - would need actual timestamp
            return False  # Placeholder implementation
        except Exception:
            return False

    @rate_limit()
    def get_stock_data(self, ticker: str, period: str = "1d") -> Optional[Dict]:
        """Get comprehensive stock data using yfinance with improved error handling"""
        try:
            # Try ticker variations first
            valid_ticker = self.try_ticker_variations(ticker)
            if not valid_ticker:
                raise ValueError(f"No valid ticker found for {ticker}")
            
            stock = yf.Ticker(valid_ticker)
            hist = stock.history(period=period)
            info = stock.info

            if hist.empty:
                raise ValueError(f"No historical data available for {valid_ticker}")

            current_price = hist['Close'].iloc[-1]
            
            # Validate essential data
            if not current_price or current_price <= 0:
                raise ValueError(f"Invalid current price for {valid_ticker}: {current_price}")
            
            # Get comprehensive financial data
            financial_data = {
                'ticker': ticker,
                'valid_ticker': valid_ticker,
                'current_price': current_price,
                'market_cap': info.get('marketCap', 0),
                'shares_outstanding': info.get('sharesOutstanding', 0),
                'trailing_pe': info.get('trailingPE', None),
                'forward_pe': info.get('forwardPE', None),
                'peg_ratio': info.get('pegRatio', None),
                'price_to_sales': info.get('priceToSalesTrailing12Months', None),
                'price_to_book': info.get('priceToBook', None),
                'dividend_yield': info.get('dividendYield', 0) if info.get('dividendYield') else 0,  # Already in percentage
                'dividend_rate': info.get('dividendRate', 0),
                'trailing_eps': info.get('trailingEps', None),
                'forward_eps': info.get('forwardEps', None),
                'revenue_growth': info.get('revenueGrowth', None),
                'earnings_growth': info.get('earningsGrowth', None),
                'free_cashflow': info.get('freeCashflow', None),
                'total_cash': info.get('totalCash', 0),
                'total_debt': info.get('totalDebt', 0),
                'book_value': info.get('bookValue', None),
                'return_on_equity': info.get('returnOnEquity', None),
                'return_on_assets': info.get('returnOnAssets', None),
                'debt_to_equity': info.get('debtToEquity', None),
                'current_ratio': info.get('currentRatio', None),
                'volume': hist['Volume'].iloc[-1] if 'Volume' in hist.columns else None,
                'avg_volume': info.get('averageVolume', None),
                'beta': info.get('beta', None),
                '52_week_high': info.get('fiftyTwoWeekHigh', None),
                '52_week_low': info.get('fiftyTwoWeekLow', None),
                'enterprise_value': info.get('enterpriseValue', None),
                'net_income': info.get('netIncomeToCommon', None),
                'revenue': info.get('totalRevenue', None),
                'ebitda': info.get('ebitda', None),
                'sector': info.get('sector', 'Unknown')
            }
            
            return financial_data
        except Exception as e:
            self.logger.error(f"Error getting data for {ticker}: {e}")
            return None

    @rate_limit()
    def get_historical_financial_data(self, ticker: str, period: str = "5y") -> Optional[Dict]:
        """Get historical financial data for growth rate calculations"""
        try:
            # Use the valid ticker if available
            valid_ticker = ticker
            if '.' in ticker:  # Try variations for NZ/AU stocks
                variations = [ticker, f"{ticker}.NZ", ticker.replace('.NZ', ''), ticker.replace('.AX', '')]
                for variant in variations:
                    try:
                        stock = yf.Ticker(variant)
                        hist = stock.history(period=period)
                        if not hist.empty:
                            valid_ticker = variant
                            break
                    except:
                        continue
            
            stock = yf.Ticker(valid_ticker)
            hist = stock.history(period=period)
            
            if hist.empty:
                return None
            
            # Calculate historical growth rates
            hist_data = {
                'price_history': hist['Close'].tolist(),
                'dates': hist.index.tolist(),
                'volume_history': hist['Volume'].tolist() if 'Volume' in hist.columns else []
            }
            
            # Calculate price growth rates
            if len(hist_data['price_history']) > 1:
                price_growth_rates = []
                for i in range(1, len(hist_data['price_history'])):
                    if hist_data['price_history'][i-1] > 0:  # Avoid division by zero
                        growth_rate = (hist_data['price_history'][i] - hist_data['price_history'][i-1]) / hist_data['price_history'][i-1]
                        price_growth_rates.append(growth_rate)
                
                hist_data['price_growth_rates'] = price_growth_rates
                hist_data['avg_price_growth_rate'] = np.mean(price_growth_rates) if price_growth_rates else 0
                
                # Calculate 5-year growth rate with division by zero protection
                if len(hist_data['price_history']) >= 5 and hist_data['price_history'][0] > 0:
                    hist_data['5_year_price_growth_rate'] = ((hist_data['price_history'][-1] / hist_data['price_history'][0]) ** (1/5)) - 1
                else:
                    hist_data['5_year_price_growth_rate'] = 0
            
            return hist_data
        except Exception as e:
            self.logger.error(f"Error getting historical data for {ticker}: {e}")
            return None

    def calculate_peter_lynch_valuation(self, stock_data, historical_data):
        """Calculate Peter Lynch valuation model"""
        try:
            if not stock_data:
                return None
            
            # Get required inputs
            current_price = stock_data.get('current_price', 0)
            trailing_pe = stock_data.get('trailing_pe', 0)
            forward_pe = stock_data.get('forward_pe', 0)
            dividend_yield = stock_data.get('dividend_yield', 0)
            earnings_growth = stock_data.get('earnings_growth', 0)
            sector = stock_data.get('sector', 'Unknown')
            
            # Use forward P/E if available, otherwise trailing P/E
            pe_ratio = forward_pe if forward_pe and forward_pe > 0 else trailing_pe
            
            if not pe_ratio or pe_ratio <= 0:
                self.logger.warning(f"No valid P/E ratio for Lynch calculation")
                return {
                    'lynch_ratio': None,
                    'valuation_status': 'N/A',
                    'valuation_color': 'gray',
                    'intrinsic_value': None,
                    'delta_percentage': None,
                    'eps_growth_rate': None,
                    'pe_ratio': None,
                    'dividend_yield': dividend_yield,
                    'weighted_growth_rate': None,
                    'advanced_lynch_ratio': None,
                    'advanced_intrinsic_value': None,
                    'advanced_delta_percentage': None
                }
            
            # Calculate earnings growth rate - standardize on decimals internally
            if earnings_growth:
                # Always store as decimals (0.10 = 10%)
                eps_growth_rate = earnings_growth if earnings_growth <= 1 else earnings_growth / 100
            elif historical_data and historical_data.get('5_year_price_growth_rate'):
                # Use historical price growth as proxy for earnings growth
                historical_growth = historical_data['5_year_price_growth_rate']
                eps_growth_rate = historical_growth if historical_growth <= 1 else historical_growth / 100
            else:
                # Conservative estimate
                eps_growth_rate = 0.10  # 10% growth as decimal
            
            # Keep as decimal for calculations, convert to percentage only for display
            eps_growth_rate_decimal = eps_growth_rate
            eps_growth_rate_percentage = eps_growth_rate * 100
            
            # Convert dividend yield from percentage to decimal for calculation
            # Handle both percentage (>1) and decimal (<1) formats
            dividend_yield_decimal = (dividend_yield / 100) if dividend_yield > 1 else dividend_yield
            
            # Basic Lynch Ratio Calculation with division by zero protection
            if pe_ratio and pe_ratio > 0:
                lynch_ratio = (eps_growth_rate_decimal + dividend_yield_decimal) / pe_ratio
                # Cap extreme Lynch ratios to prevent unrealistic valuations
                lynch_ratio = min(lynch_ratio, 5.0)  # Cap at 5x (500% upside)
            else:
                # If no P/E ratio or negative earnings, use sector-specific default
                sector_defaults = {
                    'Technology': 35, 'Healthcare': 22, 'Financial Services': 12,
                    'Energy': 15, 'Utilities': 15, 'Consumer Staples': 18,
                    'Real Estate': 15, 'Unknown': 15
                }
                default_pe = sector_defaults.get(sector, 15)
                lynch_ratio = (eps_growth_rate_decimal + dividend_yield_decimal) / default_pe
                lynch_ratio = min(lynch_ratio, 3.0)  # More conservative cap for default P/E
            
            # Interpretation using configuration thresholds
            thresholds = VALUATION_CONFIG['lynch_thresholds']
            if lynch_ratio >= thresholds['VERY_UNDERVALUED']:
                valuation_status = "VERY UNDERVALUED"
                valuation_color = "green"
            elif lynch_ratio >= thresholds['UNDERVALUED']:
                valuation_status = "UNDERVALUED"
                valuation_color = "light_green"
            elif lynch_ratio >= thresholds['FAIRLY_VALUED']:
                valuation_status = "FAIRLY VALUED"
                valuation_color = "yellow"
            else:
                valuation_status = "OVERVALUED"
                valuation_color = "red"
            
            # Calculate intrinsic value per share with division by zero protection
            if current_price and current_price > 0 and lynch_ratio > 0:
                intrinsic_value = current_price * lynch_ratio
                delta = (intrinsic_value / current_price) - 1
            else:
                intrinsic_value = 0
                delta = 0
            
            # Advanced weighted growth rate calculation (keep as decimal)
            forward_eps_growth = eps_growth_rate_decimal
            historical_eps_growth = historical_data.get('5_year_price_growth_rate', 0) if historical_data else 0
            
            weighted_growth_rate = ((forward_eps_growth * 2) + historical_eps_growth) / 3
            
            # Advanced Lynch Ratio with weighted growth
            if pe_ratio and pe_ratio > 0:
                advanced_lynch_ratio = (weighted_growth_rate + dividend_yield_decimal) / pe_ratio
                # Cap extreme ratios
                advanced_lynch_ratio = min(advanced_lynch_ratio, 5.0)
            else:
                # Use sector default for advanced calculation too
                sector_defaults = {
                    'Technology': 35, 'Healthcare': 22, 'Financial Services': 12,
                    'Energy': 15, 'Utilities': 15, 'Consumer Staples': 18,
                    'Real Estate': 15, 'Unknown': 15
                }
                default_pe = sector_defaults.get(sector, 15)
                advanced_lynch_ratio = (weighted_growth_rate + dividend_yield_decimal) / default_pe
                advanced_lynch_ratio = min(advanced_lynch_ratio, 3.0)
            
            advanced_intrinsic_value = current_price * advanced_lynch_ratio
            advanced_delta = (advanced_intrinsic_value / current_price) - 1
            
            return {
                'lynch_ratio': lynch_ratio,
                'valuation_status': valuation_status,
                'valuation_color': valuation_color,
                'intrinsic_value': intrinsic_value,
                'delta_percentage': delta * 100,
                'eps_growth_rate': eps_growth_rate_percentage,  # Return percentage for display
                'pe_ratio': pe_ratio,
                'dividend_yield': dividend_yield,  # Already in percentage
                'weighted_growth_rate': weighted_growth_rate * 100,  # Convert to percentage for display
                'advanced_lynch_ratio': advanced_lynch_ratio,
                'advanced_intrinsic_value': advanced_intrinsic_value,
                'advanced_delta_percentage': advanced_delta * 100
            }
        except Exception as e:
            self.logger.error(f"Error calculating Peter Lynch valuation: {e}")
            return None

    def calculate_dcf_valuation(self, stock_data: Dict, historical_data: Optional[Dict]) -> Optional[Dict]:
        """Calculate Discounted Cash Flow (DCF) valuation with improved assumptions"""
        try:
            if not stock_data:
                return None
            
            # Get required inputs
            current_price = stock_data.get('current_price', 0)
            free_cashflow = stock_data.get('free_cashflow', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            total_cash = stock_data.get('total_cash', 0)
            total_debt = stock_data.get('total_debt', 0)
            
            if not free_cashflow or free_cashflow <= 0:
                self.logger.warning(f"No valid free cash flow for DCF calculation")
                return {
                    'fcf_growth_rate': None,
                    'discount_rate': self.discount_rate,
                    'perpetual_growth_rate': self.perpetual_growth_rate,
                    'projection_years': self.projection_years,
                    'terminal_value': None,
                    'enterprise_value': None,
                    'equity_value': None,
                    'intrinsic_value_per_share': None,
                    'delta_percentage': None,
                    'valuation_status': 'N/A',
                    'valuation_color': 'gray',
                    'sum_pv_fcfs': None,
                    'present_value_terminal': None,
                    'future_fcfs': [],
                    'present_value_fcfs': []
                }
            
            # Calculate FCF growth rate more conservatively
            fcf_growth_rate = VALUATION_CONFIG['default_fcf_growth']  # Start with 8%
            
            if historical_data and historical_data.get('avg_price_growth_rate'):
                # Use historical growth as proxy but cap it
                historical_growth = historical_data['avg_price_growth_rate']
                fcf_growth_rate = max(
                    VALUATION_CONFIG['min_fcf_growth'],
                    min(VALUATION_CONFIG['max_fcf_growth'], historical_growth * 0.7)  # 70% of price growth
                )
            
            # Additional conservative adjustments
            if stock_data.get('beta', 1) > 1.5:  # High beta stocks
                fcf_growth_rate *= 0.8  # Reduce growth by 20%
            elif stock_data.get('beta', 1) < 0.8:  # Low beta stocks
                fcf_growth_rate *= 1.1  # Increase growth by 10%
            
            # Cap the growth rate
            fcf_growth_rate = max(
                VALUATION_CONFIG['min_fcf_growth'],
                min(VALUATION_CONFIG['max_fcf_growth'], fcf_growth_rate)
            )
            
            # Project future free cash flows
            future_fcfs = []
            present_value_fcfs = []
            
            current_fcf = free_cashflow
            
            for year in range(1, self.projection_years + 1):
                # Project future FCF
                future_fcf = current_fcf * ((1 + fcf_growth_rate) ** year)
                future_fcfs.append(future_fcf)
                
                # Calculate present value
                present_value = future_fcf / ((1 + self.discount_rate) ** year)
                present_value_fcfs.append(present_value)
            
            # Calculate terminal value with sanity check
            last_year_fcf = future_fcfs[-1]
            terminal_value = (last_year_fcf * (1 + self.perpetual_growth_rate)) / (self.discount_rate - self.perpetual_growth_rate)
            present_value_terminal = terminal_value / ((1 + self.discount_rate) ** self.projection_years)
            
            # Calculate enterprise value
            sum_pv_fcfs = sum(present_value_fcfs)
            enterprise_value = sum_pv_fcfs + present_value_terminal
            
            # Sanity check: terminal value shouldn't exceed 80% of total value
            terminal_ratio = present_value_terminal / enterprise_value if enterprise_value > 0 else 0
            if terminal_ratio > VALUATION_CONFIG['max_terminal_value_ratio']:
                self.logger.warning(f"Terminal value ratio too high: {terminal_ratio:.2%}, capping at {VALUATION_CONFIG['max_terminal_value_ratio']:.2%}")
                # Recalculate with capped terminal value
                max_terminal_value = sum_pv_fcfs * VALUATION_CONFIG['max_terminal_value_ratio'] / (1 - VALUATION_CONFIG['max_terminal_value_ratio'])
                enterprise_value = sum_pv_fcfs + max_terminal_value
                present_value_terminal = max_terminal_value
            
            # Calculate equity value
            equity_value = enterprise_value + total_cash - total_debt
            
            # Calculate intrinsic value per share with division by zero protection
            if shares_outstanding and shares_outstanding > 0:
                intrinsic_value_per_share = equity_value / shares_outstanding
            else:
                intrinsic_value_per_share = 0
            
            # Calculate delta with division by zero protection
            if current_price and current_price > 0:
                delta = (intrinsic_value_per_share / current_price) - 1
            else:
                delta = 0
            
            # Valuation assessment
            if delta > 0.2:
                valuation_status = "SIGNIFICANTLY UNDERVALUED"
                valuation_color = "green"
            elif delta > 0.05:
                valuation_status = "UNDERVALUED"
                valuation_color = "light_green"
            elif delta > -0.05:
                valuation_status = "FAIRLY VALUED"
                valuation_color = "yellow"
            elif delta > -0.2:
                valuation_status = "OVERVALUED"
                valuation_color = "orange"
            else:
                valuation_status = "SIGNIFICANTLY OVERVALUED"
                valuation_color = "red"
            
            return {
                'fcf_growth_rate': fcf_growth_rate,
                'discount_rate': self.discount_rate,
                'perpetual_growth_rate': self.perpetual_growth_rate,
                'projection_years': self.projection_years,
                'terminal_value': terminal_value,
                'enterprise_value': enterprise_value,
                'equity_value': equity_value,
                'intrinsic_value_per_share': intrinsic_value_per_share,
                'delta_percentage': delta * 100,
                'valuation_status': valuation_status,
                'valuation_color': valuation_color,
                'sum_pv_fcfs': sum_pv_fcfs,
                'present_value_terminal': present_value_terminal,
                'future_fcfs': future_fcfs,
                'present_value_fcfs': present_value_fcfs,
                'terminal_ratio': terminal_ratio
            }
        except Exception as e:
            self.logger.error(f"Error calculating DCF valuation: {e}")
            return None

    def calculate_enhanced_dcf_valuation(self, stock_data):
        """Calculate enhanced DCF with multi-scenario analysis and industry-specific adjustments"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            free_cashflow = stock_data.get('free_cashflow', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            sector = stock_data.get('sector', 'Unknown')
            beta = stock_data.get('beta', 1.0)
            
            if not free_cashflow or free_cashflow <= 0 or not shares_outstanding:
                return {
                    'enhanced_dcf_intrinsic_value': None,
                    'enhanced_dcf_status': 'N/A',
                    'enhanced_dcf_delta': 0,
                    'enhanced_dcf_scenarios': None
                }
            
            # Industry-specific WACC adjustments
            base_discount_rate = self.discount_rate
            industry_adjustments = {
                'Technology': 0.5, 'Healthcare': 0.3, 'Financial Services': 0.4,
                'Energy': 0.6, 'Utilities': -0.5, 'Consumer Staples': -0.3,
                'Real Estate': 0.2
            }
            discount_rate = base_discount_rate + industry_adjustments.get(sector, 0)
            
            # Multi-scenario analysis
            scenarios = {
                'bear': {'probability': 0.25, 'growth_multiplier': 0.5, 'terminal_multiplier': 0.8},
                'base': {'probability': 0.5, 'growth_multiplier': 1.0, 'terminal_multiplier': 1.0},
                'bull': {'probability': 0.25, 'growth_multiplier': 1.5, 'terminal_multiplier': 1.2}
            }
            
            # Calculate base growth rate with fade periods
            # Use conservative default growth rate since historical data not available in this function
            base_growth_rate = 0.05  # 5% as decimal
            
            scenario_values = {}
            
            for scenario_name, scenario_data in scenarios.items():
                # Calculate growth rates with fade periods
                growth_rates = []
                scenario_growth = base_growth_rate * scenario_data['growth_multiplier']
                
                # Fade period: growth gradually declines to perpetual rate
                fade_years = 5
                for year in range(1, self.projection_years + 1):
                    if year <= fade_years:
                        fade_factor = (fade_years - year + 1) / fade_years
                        year_growth = self.perpetual_growth_rate + (scenario_growth - self.perpetual_growth_rate) * fade_factor
                    else:
                        year_growth = self.perpetual_growth_rate
                    growth_rates.append(year_growth)
                
                # Calculate projected FCF with fade periods
                projected_fcf = []
                current_fcf = free_cashflow
                
                for year in range(self.projection_years):
                    current_fcf *= (1 + growth_rates[year])  # growth_rates is now decimal
                    projected_fcf.append(current_fcf)
                
                # Calculate terminal value with scenario adjustment
                last_year_fcf = projected_fcf[-1]
                terminal_growth = self.perpetual_growth_rate * scenario_data['terminal_multiplier']
                terminal_value = (last_year_fcf * (1 + terminal_growth)) / (discount_rate / 100 - terminal_growth)
                
                # Calculate present value of projected FCF
                pv_fcf = []
                for i, fcf in enumerate(projected_fcf):
                    pv = fcf / ((1 + discount_rate / 100) ** (i + 1))
                    pv_fcf.append(pv)
                
                # Calculate present value of terminal value
                pv_terminal = terminal_value / ((1 + discount_rate / 100) ** self.projection_years)
                
                # Total enterprise value for this scenario
                enterprise_value = sum(pv_fcf) + pv_terminal
                intrinsic_value_per_share = enterprise_value / shares_outstanding
                
                scenario_values[scenario_name] = {
                    'intrinsic_value': intrinsic_value_per_share,
                    'probability': scenario_data['probability'],
                    'growth_rate': scenario_growth,
                    'terminal_value': terminal_value
                }
            
            # Calculate probability-weighted intrinsic value
            weighted_value = sum(scenario_values[s]['intrinsic_value'] * scenario_values[s]['probability'] 
                                for s in scenario_values)
            
            # Determine valuation status
            if weighted_value > current_price * 1.2:
                valuation_status = "SIGNIFICANTLY UNDERVALUED"
            elif weighted_value > current_price * 1.05:
                valuation_status = "UNDERVALUED"
            elif weighted_value > current_price * 0.95:
                valuation_status = "FAIRLY VALUED"
            elif weighted_value > current_price * 0.8:
                valuation_status = "OVERVALUED"
            else:
                valuation_status = "SIGNIFICANTLY OVERVALUED"
            
            delta_percentage = ((weighted_value - current_price) / current_price) * 100
            
            return {
                'enhanced_dcf_intrinsic_value': weighted_value,
                'enhanced_dcf_status': valuation_status,
                'enhanced_dcf_delta': delta_percentage,
                'enhanced_dcf_scenarios': scenario_values,
                'enhanced_dcf_discount_rate': discount_rate,
                'enhanced_dcf_base_growth_rate': base_growth_rate
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating enhanced DCF valuation: {e}")
            return {
                'enhanced_dcf_intrinsic_value': None,
                'enhanced_dcf_status': 'N/A',
                'enhanced_dcf_delta': 0,
                'enhanced_dcf_scenarios': None
            }

    def calculate_relative_valuation(self, stock_data):
        """Calculate relative valuation using comparable company multiples"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            sector = stock_data.get('sector', 'Unknown')
            
            # Get financial metrics
            enterprise_value = stock_data.get('enterprise_value', 0)
            ebitda = stock_data.get('ebitda', 0)
            net_income = stock_data.get('net_income', 0)
            revenue = stock_data.get('revenue', 0)
            book_value = stock_data.get('book_value', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            
            if not shares_outstanding or shares_outstanding <= 0:
                return {
                    'relative_valuation_status': 'N/A',
                    'relative_valuation_delta': 0,
                    'ev_ebitda_multiple': None,
                    'pe_ratio': None,
                    'ps_ratio': None,
                    'pb_ratio': None
                }
            
            # Calculate multiples
            ev_ebitda_multiple = None
            pe_ratio = None
            ps_ratio = None
            pb_ratio = None
            
            if ebitda and ebitda > 0:
                ev_ebitda_multiple = enterprise_value / ebitda
            
            if net_income and net_income > 0 and shares_outstanding and shares_outstanding > 0:
                pe_ratio = (current_price * shares_outstanding) / net_income
            
            if revenue and revenue > 0 and shares_outstanding and shares_outstanding > 0:
                ps_ratio = (current_price * shares_outstanding) / revenue
            
            if book_value and book_value > 0 and shares_outstanding and shares_outstanding > 0:
                pb_ratio = (current_price * shares_outstanding) / book_value
            
            # Sector-specific median multiples (simplified - in practice, you'd use real sector data)
            sector_multiples = {
                'Technology': {'ev_ebitda': 15, 'pe': 25, 'ps': 5, 'pb': 3},
                'Healthcare': {'ev_ebitda': 12, 'pe': 20, 'ps': 4, 'pb': 2.5},
                'Financial Services': {'ev_ebitda': 8, 'pe': 12, 'ps': 2, 'pb': 1.2},
                'Energy': {'ev_ebitda': 6, 'pe': 15, 'ps': 1.5, 'pb': 1.5},
                'Utilities': {'ev_ebitda': 10, 'pe': 18, 'ps': 2.5, 'pb': 1.8},
                'Consumer Staples': {'ev_ebitda': 12, 'pe': 18, 'ps': 2, 'pb': 2},
                'Real Estate': {'ev_ebitda': 15, 'pe': 20, 'ps': 8, 'pb': 1.5}
            }
            
            sector_medians = sector_multiples.get(sector, {'ev_ebitda': 12, 'pe': 18, 'ps': 3, 'pb': 2})
            
            # Calculate valuation scores
            valuation_scores = []
            
            if ev_ebitda_multiple and sector_medians['ev_ebitda']:
                ev_ebitda_score = (sector_medians['ev_ebitda'] - ev_ebitda_multiple) / sector_medians['ev_ebitda']
                valuation_scores.append(ev_ebitda_score)
            
            if pe_ratio and sector_medians['pe']:
                pe_score = (sector_medians['pe'] - pe_ratio) / sector_medians['pe']
                valuation_scores.append(pe_score)
            
            if ps_ratio and sector_medians['ps']:
                ps_score = (sector_medians['ps'] - ps_ratio) / sector_medians['ps']
                valuation_scores.append(ps_score)
            
            if pb_ratio and sector_medians['pb']:
                pb_score = (sector_medians['pb'] - pb_ratio) / sector_medians['pb']
                valuation_scores.append(pb_score)
            
            # Calculate average valuation score
            if valuation_scores:
                avg_score = sum(valuation_scores) / len(valuation_scores)
                
                if avg_score > 0.2:
                    valuation_status = "SIGNIFICANTLY UNDERVALUED"
                elif avg_score > 0.05:
                    valuation_status = "UNDERVALUED"
                elif avg_score > -0.05:
                    valuation_status = "FAIRLY VALUED"
                elif avg_score > -0.2:
                    valuation_status = "OVERVALUED"
                else:
                    valuation_status = "SIGNIFICANTLY OVERVALUED"
                
                delta_percentage = avg_score * 100
            else:
                valuation_status = "N/A"
                delta_percentage = 0
            
            return {
                'relative_valuation_status': valuation_status,
                'relative_valuation_delta': delta_percentage,
                'ev_ebitda_multiple': ev_ebitda_multiple,
                'pe_ratio': pe_ratio,
                'ps_ratio': ps_ratio,
                'pb_ratio': pb_ratio,
                'sector_medians': sector_medians,
                'valuation_scores': valuation_scores
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating relative valuation: {e}")
            return {
                'relative_valuation_status': 'N/A',
                'relative_valuation_delta': 0,
                'ev_ebitda_multiple': None,
                'pe_ratio': None,
                'ps_ratio': None,
                'pb_ratio': None
            }

    def calculate_munger_farm_valuation(self, stock_data):
        """Calculate Charlie Munger's farm valuation concept"""
        try:
            if not stock_data:
                return None
            
            # Get required inputs
            current_price = stock_data.get('current_price', 0)
            trailing_eps = stock_data.get('trailing_eps', 0)
            market_cap = stock_data.get('market_cap', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            
            if not trailing_eps or trailing_eps <= 0:
                self.logger.warning(f"No valid EPS for Munger farm valuation")
                return {
                    'annual_profit_per_share': None,
                    'farm_valuations': {},
                    'total_company_valuations': {},
                    'current_price': current_price,
                    'market_cap': market_cap
                }
            
            # Calculate annual profit per share (EPS)
            annual_profit_per_share = trailing_eps
            
            # Different required return scenarios
            required_returns = [0.05, 0.07, 0.10, 0.12, 0.15]  # 5%, 7%, 10%, 12%, 15%
            
            farm_valuations = {}
            
            for required_return in required_returns:
                # Calculate intrinsic value based on required return
                intrinsic_value = annual_profit_per_share / required_return
                
                # Calculate delta
                delta = (intrinsic_value / current_price) - 1 if current_price > 0 else 0
                
                # Assessment
                if delta > 0.2:
                    assessment = "STRONG BUY"
                elif delta > 0.05:
                    assessment = "BUY"
                elif delta > -0.05:
                    assessment = "HOLD"
                elif delta > -0.2:
                    assessment = "SELL"
                else:
                    assessment = "STRONG SELL"
                
                farm_valuations[f"{int(required_return*100)}%_return"] = {
                    'required_return': required_return,
                    'intrinsic_value': intrinsic_value,
                    'delta_percentage': delta * 100,
                    'assessment': assessment
                }
            
            # Calculate total company valuation scenarios
            total_company_valuations = {}
            for required_return in required_returns:
                total_intrinsic_value = annual_profit_per_share * shares_outstanding / required_return
                total_delta = (total_intrinsic_value / market_cap) - 1 if market_cap > 0 else 0
                
                total_company_valuations[f"{int(required_return*100)}%_return"] = {
                    'required_return': required_return,
                    'total_intrinsic_value': total_intrinsic_value,
                    'total_delta_percentage': total_delta * 100
                }
            
            return {
                'annual_profit_per_share': annual_profit_per_share,
                'farm_valuations': farm_valuations,
                'total_company_valuations': total_company_valuations,
                'current_price': current_price,
                'market_cap': market_cap
            }
        except Exception as e:
            self.logger.error(f"Error calculating Munger farm valuation: {e}")
            return None

    def calculate_reverse_dcf(self, stock_data):
        """Calculate reverse DCF to determine what growth rate is priced in by the market"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            free_cashflow = stock_data.get('free_cashflow', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            sector = stock_data.get('sector', 'Unknown')
            
            if not free_cashflow or free_cashflow <= 0 or not shares_outstanding or not current_price:
                return {
                    'reverse_dcf_implied_growth': None,
                    'reverse_dcf_assessment': 'N/A',
                    'reverse_dcf_reasonable': None
                }
            
            # Industry-specific WACC adjustments
            base_discount_rate = self.discount_rate
            industry_adjustments = {
                'Technology': 0.5, 'Healthcare': 0.3, 'Financial Services': 0.4,
                'Energy': 0.6, 'Utilities': -0.5, 'Consumer Staples': -0.3,
                'Real Estate': 0.2
            }
            discount_rate = base_discount_rate + industry_adjustments.get(sector, 0)
            
            # Calculate implied growth rate using iterative approach
            market_cap = current_price * shares_outstanding
            target_enterprise_value = market_cap  # Simplified
            
            # Use binary search to find implied growth rate
            low_growth = 0.0
            high_growth = 0.5  # Cap at 50% growth
            relative_tolerance = 0.01  # 1% relative accuracy (more reasonable)
            max_iterations = 50
            
            implied_growth = None
            
            for iteration in range(max_iterations):
                test_growth = (low_growth + high_growth) / 2
                
                # Calculate DCF with test growth rate
                projected_fcf = []
                current_fcf = free_cashflow
                
                for year in range(1, self.projection_years + 1):
                    current_fcf *= (1 + test_growth)
                    projected_fcf.append(current_fcf)
                
                # Calculate terminal value
                last_year_fcf = projected_fcf[-1]
                terminal_value = (last_year_fcf * (1 + self.perpetual_growth_rate)) / (discount_rate - self.perpetual_growth_rate)
                
                # Calculate present value of projected FCF
                pv_fcf = []
                for i, fcf in enumerate(projected_fcf):
                    pv = fcf / ((1 + discount_rate) ** (i + 1))
                    pv_fcf.append(pv)
                
                # Calculate present value of terminal value
                pv_terminal = terminal_value / ((1 + discount_rate) ** self.projection_years)
                
                # Total enterprise value
                calculated_enterprise_value = sum(pv_fcf) + pv_terminal
                
                # Check if we're close enough (use relative tolerance)
                if target_enterprise_value > 0 and abs(calculated_enterprise_value - target_enterprise_value) / target_enterprise_value < relative_tolerance:
                    implied_growth = test_growth
                    break
                
                # Adjust search range
                if calculated_enterprise_value < target_enterprise_value:
                    low_growth = test_growth
                else:
                    high_growth = test_growth
            
            if implied_growth is None:
                implied_growth = (low_growth + high_growth) / 2
            
            # Convert to percentage
            implied_growth_percentage = implied_growth * 100
            
            # Assess reasonableness
            # Use conservative assessment since historical data not available in this function
            assessment = "NO HISTORICAL GROWTH DATA AVAILABLE"
            reasonable = True
            
            return {
                'reverse_dcf_implied_growth': implied_growth_percentage,
                'reverse_dcf_assessment': assessment,
                'reverse_dcf_reasonable': reasonable,
                'reverse_dcf_historical_growth': None,  # Not available in this function
                'reverse_dcf_growth_ratio': None,  # Not available in this function
                'reverse_dcf_discount_rate': discount_rate
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating reverse DCF: {e}")
            return {
                'reverse_dcf_implied_growth': None,
                'reverse_dcf_assessment': 'N/A',
                'reverse_dcf_reasonable': None
            }

    def calculate_earnings_power_value(self, stock_data):
        """Calculate Bruce Greenwald's Earnings Power Value"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            net_income = stock_data.get('net_income', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            
            if not net_income or net_income <= 0 or not shares_outstanding:
                return {
                    'epv_intrinsic_value': None,
                    'epv_assessment': 'N/A',
                    'epv_delta': 0
                }
            
            # Normalize earnings (adjust for cyclicality)
            normalized_earnings = net_income  # Simplified - use multi-year average in practice
            
            # Capitalize at appropriate rate (conservative)
            epv_discount_rate = self.discount_rate + 0.02  # Add 2% for no-growth assumption
            
            # Calculate EPV
            epv_enterprise_value = normalized_earnings / epv_discount_rate
            epv_intrinsic_value = epv_enterprise_value / shares_outstanding
            
            # Determine valuation status
            if epv_intrinsic_value > current_price * 1.2:
                assessment = "SIGNIFICANTLY UNDERVALUED"
            elif epv_intrinsic_value > current_price * 1.05:
                assessment = "UNDERVALUED"
            elif epv_intrinsic_value > current_price * 0.95:
                assessment = "FAIRLY VALUED"
            elif epv_intrinsic_value > current_price * 0.8:
                assessment = "OVERVALUED"
            else:
                assessment = "SIGNIFICANTLY OVERVALUED"
            
            delta_percentage = ((epv_intrinsic_value - current_price) / current_price) * 100
            
            return {
                'epv_intrinsic_value': epv_intrinsic_value,
                'epv_assessment': assessment,
                'epv_delta': delta_percentage,
                'epv_normalized_earnings': normalized_earnings,
                'epv_discount_rate': epv_discount_rate
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating Earnings Power Value: {e}")
            return {
                'epv_intrinsic_value': None,
                'epv_assessment': 'N/A',
                'epv_delta': 0
            }

    def calculate_residual_income_model(self, stock_data):
        """Calculate Residual Income Model - particularly good for financial stocks"""
        try:
            if not stock_data:
                return None
            
            current_price = stock_data.get('current_price', 0)
            book_value = stock_data.get('book_value', 0)
            net_income = stock_data.get('net_income', 0)
            shares_outstanding = stock_data.get('shares_outstanding', 0)
            sector = stock_data.get('sector', 'Unknown')
            roe = stock_data.get('return_on_equity', 0)
            
            if not book_value or book_value <= 0 or not shares_outstanding:
                return {
                    'rim_intrinsic_value': None,
                    'rim_assessment': 'N/A - Missing Book Value',
                    'rim_delta': 0
                }
            
            # Check if company is profitable (RIM doesn't work for unprofitable companies)
            if net_income <= 0:
                return {
                    'rim_intrinsic_value': None,
                    'rim_assessment': 'N/A - Company Unprofitable',
                    'rim_delta': 0
                }
            
            # Calculate book value per share
            book_value_per_share = book_value / shares_outstanding
            
            # Determine cost of equity based on sector
            sector_cost_of_equity = {
                'Financial Services': 0.10,  # 10% - higher risk
                'Technology': 0.12,          # 12% - high growth, high risk
                'Healthcare': 0.09,         # 9% - moderate risk
                'Energy': 0.11,             # 11% - cyclical risk
                'Utilities': 0.07,          # 7% - lower risk
                'Consumer Staples': 0.08,   # 8% - stable
                'Real Estate': 0.09         # 9% - moderate risk
            }
            
            cost_of_equity = sector_cost_of_equity.get(sector, 0.10)  # Default 10%
            
            # Calculate current ROE if not provided
            if not roe and net_income and book_value:
                roe = net_income / book_value
            
            # Project residual income for next 5 years
            projection_years = 5
            residual_income_projection = []
            current_book_value = book_value_per_share
            
            # Use current ROE or sector average if not available
            if not roe or roe <= 0:
                sector_roe_averages = {
                    'Financial Services': 0.12,
                    'Technology': 0.15,
                    'Healthcare': 0.10,
                    'Energy': 0.08,
                    'Utilities': 0.08,
                    'Consumer Staples': 0.12,
                    'Real Estate': 0.10
                }
                roe = sector_roe_averages.get(sector, 0.10)
            
            # Apply fade period - ROE gradually declines to cost of equity
            fade_years = 3
            terminal_roe = cost_of_equity
            
            for year in range(1, projection_years + 1):
                # Calculate ROE for this year (with fade)
                if year <= fade_years:
                    fade_factor = (fade_years - year + 1) / fade_years
                    year_roe = terminal_roe + (roe - terminal_roe) * fade_factor
                else:
                    year_roe = terminal_roe
                
                # Calculate expected earnings
                expected_earnings = current_book_value * year_roe
                
                # Calculate residual income
                residual_income = expected_earnings - (current_book_value * cost_of_equity)
                
                # Calculate present value of residual income
                pv_residual_income = residual_income / ((1 + cost_of_equity) ** year)
                
                residual_income_projection.append({
                    'year': year,
                    'roe': year_roe,
                    'expected_earnings': expected_earnings,
                    'residual_income': residual_income,
                    'pv_residual_income': pv_residual_income
                })
                
                # Update book value for next year
                current_book_value += expected_earnings
            
            # Calculate terminal value (assuming ROE = cost of equity)
            terminal_residual_income = residual_income_projection[-1]['residual_income']
            terminal_value = terminal_residual_income / cost_of_equity
            pv_terminal_value = terminal_value / ((1 + cost_of_equity) ** projection_years)
            
            # Calculate intrinsic value
            sum_pv_residual_income = sum(ri['pv_residual_income'] for ri in residual_income_projection)
            rim_intrinsic_value = book_value_per_share + sum_pv_residual_income + pv_terminal_value
            
            # Determine valuation status
            if rim_intrinsic_value > current_price * 1.2:
                assessment = "SIGNIFICANTLY UNDERVALUED"
            elif rim_intrinsic_value > current_price * 1.05:
                assessment = "UNDERVALUED"
            elif rim_intrinsic_value > current_price * 0.95:
                assessment = "FAIRLY VALUED"
            elif rim_intrinsic_value > current_price * 0.8:
                assessment = "OVERVALUED"
            else:
                assessment = "SIGNIFICANTLY OVERVALUED"
            
            delta_percentage = ((rim_intrinsic_value - current_price) / current_price) * 100
            
            return {
                'rim_intrinsic_value': rim_intrinsic_value,
                'rim_assessment': assessment,
                'rim_delta': delta_percentage,
                'rim_book_value_per_share': book_value_per_share,
                'rim_cost_of_equity': cost_of_equity,
                'rim_current_roe': roe,
                'rim_terminal_value': terminal_value,
                'rim_residual_income_projection': residual_income_projection,
                'rim_sum_pv_residual_income': sum_pv_residual_income
            }
            
        except Exception as e:
            self.logger.error(f"Error calculating Residual Income Model: {e}")
            return {
                'rim_intrinsic_value': None,
                'rim_assessment': 'N/A',
                'rim_delta': 0
            }

    def collect_valuation_metrics(self, ticker):
        """Collect all valuation metrics for a specific ticker"""
        self.logger.info(f"Collecting valuation metrics for {ticker}")
        
        # Get current stock data
        stock_data = self.get_stock_data(ticker)
        if not stock_data:
            self.logger.error(f"Failed to get stock data for {ticker}")
            return None
        
        # Get historical data
        historical_data = self.get_historical_financial_data(ticker)
        
        # Calculate all valuation methods
        lynch_valuation = self.calculate_peter_lynch_valuation(stock_data, historical_data)
        dcf_valuation = self.calculate_dcf_valuation(stock_data, historical_data)
        munger_valuation = self.calculate_munger_farm_valuation(stock_data)
        
        # Calculate new Tier 1 and Tier 2 valuation methods
        enhanced_dcf_valuation = self.calculate_enhanced_dcf_valuation(stock_data)
        relative_valuation = self.calculate_relative_valuation(stock_data)
        reverse_dcf_valuation = self.calculate_reverse_dcf(stock_data)
        epv_valuation = self.calculate_earnings_power_value(stock_data)
        rim_valuation = self.calculate_residual_income_model(stock_data)
        
        # Combine all data
        valuation_data = {
            'date': datetime.now().strftime('%Y-%m-%d'),
            'time': datetime.now().strftime('%H:%M:%S'),
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'ticker': ticker,
            'company_name': self.focus_stocks.get(ticker, ticker)
        }
        
        # Add stock data
        valuation_data.update(stock_data)
        
        # Add historical data summary
        if historical_data:
            valuation_data.update({
                'avg_price_growth_rate': historical_data.get('avg_price_growth_rate', 0),
                '5_year_price_growth_rate': historical_data.get('5_year_price_growth_rate', 0),
                'price_history_length': len(historical_data.get('price_history', []))
            })
        
        # Add Lynch valuation
        if lynch_valuation:
            valuation_data.update({
                'lynch_ratio': lynch_valuation.get('lynch_ratio', 0),
                'lynch_valuation_status': lynch_valuation.get('valuation_status', 'N/A'),
                'lynch_intrinsic_value': lynch_valuation.get('intrinsic_value', 0),
                'lynch_delta_percentage': lynch_valuation.get('delta_percentage', 0),
                'lynch_eps_growth_rate': lynch_valuation.get('eps_growth_rate', 0),
                'advanced_lynch_ratio': lynch_valuation.get('advanced_lynch_ratio', 0),
                'advanced_lynch_intrinsic_value': lynch_valuation.get('advanced_intrinsic_value', 0),
                'advanced_lynch_delta_percentage': lynch_valuation.get('advanced_delta_percentage', 0)
            })
        
        # Add DCF valuation
        if dcf_valuation:
            valuation_data.update({
                'dcf_fcf_growth_rate': dcf_valuation.get('fcf_growth_rate', 0),
                'dcf_discount_rate': dcf_valuation.get('discount_rate', 0),
                'dcf_perpetual_growth_rate': dcf_valuation.get('perpetual_growth_rate', 0),
                'dcf_terminal_value': dcf_valuation.get('terminal_value', 0),
                'dcf_enterprise_value': dcf_valuation.get('enterprise_value', 0),
                'dcf_equity_value': dcf_valuation.get('equity_value', 0),
                'dcf_intrinsic_value_per_share': dcf_valuation.get('intrinsic_value_per_share', 0),
                'dcf_delta_percentage': dcf_valuation.get('delta_percentage', 0),
                'dcf_valuation_status': dcf_valuation.get('valuation_status', 'N/A')
            })
        
        # Add Munger valuation summary
        if munger_valuation:
            valuation_data.update({
                'munger_annual_profit_per_share': munger_valuation.get('annual_profit_per_share', 0),
                'munger_7pct_intrinsic_value': munger_valuation.get('farm_valuations', {}).get('7%_return', {}).get('intrinsic_value', 0),
                'munger_7pct_delta_percentage': munger_valuation.get('farm_valuations', {}).get('7%_return', {}).get('delta_percentage', 0),
                'munger_7pct_assessment': munger_valuation.get('farm_valuations', {}).get('7%_return', {}).get('assessment', 'N/A'),
                'munger_10pct_intrinsic_value': munger_valuation.get('farm_valuations', {}).get('10%_return', {}).get('intrinsic_value', 0),
                'munger_10pct_delta_percentage': munger_valuation.get('farm_valuations', {}).get('10%_return', {}).get('delta_percentage', 0),
                'munger_10pct_assessment': munger_valuation.get('farm_valuations', {}).get('10%_return', {}).get('assessment', 'N/A')
            })
        
        # Add Enhanced DCF valuation (Tier 1)
        if enhanced_dcf_valuation:
            valuation_data.update({
                'enhanced_dcf_intrinsic_value': enhanced_dcf_valuation.get('enhanced_dcf_intrinsic_value', 0),
                'enhanced_dcf_status': enhanced_dcf_valuation.get('enhanced_dcf_status', 'N/A'),
                'enhanced_dcf_delta': enhanced_dcf_valuation.get('enhanced_dcf_delta', 0),
                'enhanced_dcf_discount_rate': enhanced_dcf_valuation.get('enhanced_dcf_discount_rate', 0),
                'enhanced_dcf_base_growth_rate': enhanced_dcf_valuation.get('enhanced_dcf_base_growth_rate', 0)
            })
        
        # Add Relative Valuation (Tier 1)
        if relative_valuation:
            valuation_data.update({
                'relative_valuation_status': relative_valuation.get('relative_valuation_status', 'N/A'),
                'relative_valuation_delta': relative_valuation.get('relative_valuation_delta', 0),
                'ev_ebitda_multiple': relative_valuation.get('ev_ebitda_multiple', 0),
                'pe_ratio': relative_valuation.get('pe_ratio', 0),
                'ps_ratio': relative_valuation.get('ps_ratio', 0),
                'pb_ratio': relative_valuation.get('pb_ratio', 0)
            })
        
        # Add Reverse DCF (Tier 2)
        if reverse_dcf_valuation:
            valuation_data.update({
                'reverse_dcf_implied_growth': reverse_dcf_valuation.get('reverse_dcf_implied_growth', 0),
                'reverse_dcf_assessment': reverse_dcf_valuation.get('reverse_dcf_assessment', 'N/A'),
                'reverse_dcf_reasonable': reverse_dcf_valuation.get('reverse_dcf_reasonable', None),
                'reverse_dcf_historical_growth': reverse_dcf_valuation.get('reverse_dcf_historical_growth', 0),
                'reverse_dcf_growth_ratio': reverse_dcf_valuation.get('reverse_dcf_growth_ratio', 0)
            })
        
        # Add Earnings Power Value (Tier 2)
        if epv_valuation:
            valuation_data.update({
                'epv_intrinsic_value': epv_valuation.get('epv_intrinsic_value', 0),
                'epv_assessment': epv_valuation.get('epv_assessment', 'N/A'),
                'epv_delta': epv_valuation.get('epv_delta', 0),
                'epv_normalized_earnings': epv_valuation.get('epv_normalized_earnings', 0),
                'epv_discount_rate': epv_valuation.get('epv_discount_rate', 0)
            })
        
        # Add Residual Income Model (Tier 2)
        if rim_valuation:
            valuation_data.update({
                'rim_intrinsic_value': rim_valuation.get('rim_intrinsic_value', 0),
                'rim_assessment': rim_valuation.get('rim_assessment', 'N/A'),
                'rim_delta': rim_valuation.get('rim_delta', 0),
                'rim_book_value_per_share': rim_valuation.get('rim_book_value_per_share', 0),
                'rim_cost_of_equity': rim_valuation.get('rim_cost_of_equity', 0),
                'rim_current_roe': rim_valuation.get('rim_current_roe', 0)
            })
        
        # Check for valuation consensus/disagreement
        lynch_delta = valuation_data.get('lynch_delta_percentage', 0)
        dcf_delta = valuation_data.get('dcf_delta_percentage', 0)
        munger_delta = valuation_data.get('munger_7pct_delta_percentage', 0)
        self.check_valuation_consensus(lynch_delta, dcf_delta, munger_delta)
        
        self.logger.info(f"Collected {len(valuation_data)} valuation metrics for {ticker}")
        return valuation_data

    def load_existing_data(self):
        """Load existing valuation data if available and recent"""
        try:
            if not os.path.exists(self.master_file):
                return None
            
            # Load existing data
            df_existing = pd.read_excel(self.master_file, sheet_name='Valuation Data')
            
            # Check if data has required new fields
            required_fields = ['net_income', 'revenue', 'ebitda', 'sector']
            missing_fields = [field for field in required_fields if field not in df_existing.columns]
            
            if missing_fields:
                self.logger.info(f"Existing data missing required fields: {missing_fields}, will fetch fresh data")
                return None
            
            # Check if data is recent (within last 1 hour for fresh calculations)
            if 'analysis_timestamp' in df_existing.columns:
                latest_timestamp = pd.to_datetime(df_existing['analysis_timestamp']).max()
                time_diff = datetime.now() - latest_timestamp
                
                if time_diff.total_seconds() < 1 * 3600:  # 1 hour for fresh data
                    self.logger.info(f"Using existing data from {latest_timestamp} (age: {time_diff})")
                    return df_existing
            
            self.logger.info("Existing data is stale, will fetch fresh data")
            return None
            
        except Exception as e:
            self.logger.warning(f"Could not load existing data: {e}")
            return None

    def run(self):
        """Main execution method"""
        self.logger.info("Starting Stock Valuation Scraper")
        start_time = datetime.now()
        
        try:
            # Try to load existing data first
            existing_df = self.load_existing_data()
            
            if existing_df is not None:
                # Use existing data
                df = existing_df
                self.logger.info(f"Using existing data with {len(df)} stocks")
            else:
                # Collect fresh valuation metrics for focus stocks
                all_valuation_data = []
                
                # Process stocks in batches with progress indicator
                stock_items = list(self.focus_stocks.items())
                batch_size = VALUATION_CONFIG['batch_size']
                
                for i in tqdm(range(0, len(stock_items), batch_size), desc="Processing stock batches"):
                    batch = stock_items[i:i + batch_size]
                    
                    for ticker, company_name in batch:
                        self.logger.info(f"Analyzing {company_name} ({ticker})")
                        valuation_data = self.collect_valuation_metrics(ticker)
                        if valuation_data:
                            all_valuation_data.append(valuation_data)
                    
                    # Add delay between batches
                    if i + batch_size < len(stock_items):
                        time.sleep(VALUATION_CONFIG['batch_delay'])
                
                if all_valuation_data:
                    # Create DataFrame
                    df = pd.DataFrame(all_valuation_data)
                    # Add timestamp
                    df['analysis_timestamp'] = datetime.now()
                else:
                    self.logger.error("No valuation data collected")
                    return
            
            # Save dataset
            self.save_dataset(df)
            
            # Print summary
            self.print_valuation_summary(df)
            
            end_time = datetime.now()
            duration = end_time - start_time
            self.logger.info(f"Stock valuation analysis completed in {duration}")
            
        except Exception as e:
            self.logger.error(f"Error during stock valuation analysis: {e}")

    def print_valuation_summary(self, df):
        """Print comprehensive valuation summary"""
        print(f"\n=== Stock Valuation Analysis Summary ===")
        print(f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Stocks Analyzed: {len(df)}")
        
        for _, row in df.iterrows():
            ticker = row['ticker']
            company_name = row['company_name']
            current_price = row['current_price']
            
            print(f"\n--- {company_name} ({ticker}) ---")
            print(f"Current Price: ${current_price:,.2f}")
            
            # Lynch Valuation Summary
            if row.get('lynch_valuation_status') != 'N/A':
                print(f"Peter Lynch Valuation:")
                print(f"  Status: {row['lynch_valuation_status']}")
                print(f"  Lynch Ratio: {row.get('lynch_ratio', 0):.2f}")
                print(f"  Intrinsic Value: ${row.get('lynch_intrinsic_value', 0):,.2f}")
                print(f"  Delta: {row.get('lynch_delta_percentage', 0):+.1f}%")
            
            # DCF Valuation Summary
            if row.get('dcf_valuation_status') != 'N/A':
                print(f"DCF Valuation:")
                print(f"  Status: {row['dcf_valuation_status']}")
                print(f"  Intrinsic Value: ${row.get('dcf_intrinsic_value_per_share', 0):,.2f}")
                print(f"  Delta: {row.get('dcf_delta_percentage', 0):+.1f}%")
            
            # Munger Farm Valuation Summary
            if row.get('munger_7pct_assessment') != 'N/A':
                print(f"Munger Farm Valuation (7% return):")
                print(f"  Assessment: {row['munger_7pct_assessment']}")
                print(f"  Intrinsic Value: ${row.get('munger_7pct_intrinsic_value', 0):,.2f}")
                print(f"  Delta: {row.get('munger_7pct_delta_percentage', 0):+.1f}%")

    def save_dataset(self, df):
        """Save valuation dataset to Excel with formatting"""
        if df.empty:
            self.logger.warning("No data to save")
            return
        
        try:
            # Create timestamped filename for daily backup
            timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
            daily_backup_file = os.path.join(self.daily_backups_dir, f"stock_valuation_dataset_{timestamp}.xlsx")
            
            # Try to load existing workbook first
            try:
                wb = load_workbook(self.master_file)
                self.logger.info("Loaded existing workbook, adding new data")
            except FileNotFoundError:
                # Create new workbook if file doesn't exist
                wb = Workbook()
                wb.remove(wb.active)
                self.logger.info("Created new workbook")
            
            # Create or update summary sheet
            if 'My Portfolio' in wb.sheetnames:
                wb.remove(wb['My Portfolio'])
            self.create_valuation_summary_sheet(df, wb)
            
            # Create or update prospects sheet
            if 'Prospects' in wb.sheetnames:
                wb.remove(wb['Prospects'])
            self.create_prospects_sheet(df, wb)
            
            # Create or update detailed data sheet
            if 'Valuation Data' in wb.sheetnames:
                # Check if we need to update headers for new columns
                ws_data = wb['Valuation Data']
                existing_headers = [cell.value for cell in ws_data[1]]
                new_headers = list(df.columns)
                
                # If headers don't match, recreate the sheet with new headers
                if existing_headers != new_headers:
                    self.logger.info("Updating Valuation Data sheet with new columns")
                    wb.remove(ws_data)
                    ws_data = wb.create_sheet('Valuation Data', 2)
                    # Add new headers
                    for col, header in enumerate(new_headers, 1):
                        ws_data.cell(row=1, column=col, value=header)
                    next_row = 2
                else:
                    # Find the next empty row
                    next_row = ws_data.max_row + 1
            else:
                ws_data = wb.create_sheet('Valuation Data', 2)
                next_row = 1
                # Add headers if new sheet
                headers = list(df.columns)
                for col, header in enumerate(headers, 1):
                    ws_data.cell(row=1, column=col, value=header)
                next_row = 2
            
            # Add new data to worksheet using chunked writing for better memory management
            chunk_size = 100  # Process 100 rows at a time
            for chunk_start in range(0, len(df), chunk_size):
                chunk_end = min(chunk_start + chunk_size, len(df))
                chunk_df = df.iloc[chunk_start:chunk_end]
                
                for _, row in chunk_df.iterrows():
                    for col, value in enumerate(row, 1):
                        ws_data.cell(row=next_row, column=col, value=value)
                    next_row += 1
            
            # Auto-adjust column widths for Valuation Data sheet
            for column in ws_data.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 35)  # Cap at 35 for better readability
                ws_data.column_dimensions[column_letter].width = adjusted_width
            
            # Save files
            wb.save(self.master_file)
            wb.save(daily_backup_file)
            
            if self.downloads_file:
                wb.save(self.downloads_file)
            
            self.logger.info(f"Valuation dataset saved to: {self.master_file}")
            self.logger.info(f"Daily backup saved to: {daily_backup_file}")
            
        except Exception as e:
            self.logger.error(f"Error saving valuation dataset: {e}")

    def create_valuation_summary_sheet(self, df, wb):
        """Create comprehensive valuation summary sheet with all 9 valuation methods"""
        try:
            ws_summary = wb.create_sheet('My Portfolio', 0)
            
            # Define styles
            header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=14)
            subheader_font = Font(bold=True, size=12)
            data_font = Font(size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Improved conditional formatting fills with better contrast
            strong_buy_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Forest Green
            buy_fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")  # Lime Green
            hold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
            sell_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")  # Tomato
            strong_sell_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Crimson
            no_data_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray
            
            row = 1
            
            # Title
            ws_summary[f'A{row}'] = "My Portfolio - All Holdings Analysis"
            ws_summary[f'A{row}'].font = Font(bold=True, size=16)
            ws_summary[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws_summary.merge_cells(f'A{row}:I{row}')
            row += 1
            
            # Analysis date
            ws_summary[f'A{row}'] = f"Analysis Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws_summary[f'A{row}'].font = Font(italic=True, size=10)
            row += 1
            
            ws_summary[f'A{row}'] = f"Stocks Analyzed: {len(df)}"
            ws_summary[f'A{row}'].font = Font(italic=True, size=10)
            row += 1
            
            # Historical data reference
            ws_summary[f'A{row}'] = f"Historical Data: See 'stock_valuation_historical.xlsx' for comprehensive daily trends, P/E ratios, and financial history"
            ws_summary[f'A{row}'].font = Font(italic=True, size=10, color="366092")
            ws_summary[f'A{row}'].alignment = Alignment(horizontal='left')
            ws_summary.merge_cells(f'A{row}:I{row}')
            row += 2
            
            # Headers for all 9 valuation methods
            ws_summary[f'A{row}'] = "Company (Ticker)"
            ws_summary[f'B{row}'] = "Peter Lynch"
            ws_summary[f'C{row}'] = "DCF Valuation"
            ws_summary[f'D{row}'] = "Munger Farm"
            ws_summary[f'E{row}'] = "Enhanced DCF"
            ws_summary[f'F{row}'] = "Relative Valuation"
            ws_summary[f'G{row}'] = "Reverse DCF"
            ws_summary[f'H{row}'] = "EPV/RIM"
            ws_summary[f'I{row}'] = "Current Price"
            
            # Format headers
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                cell = ws_summary[f'{col}{row}']
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            row += 1
            
            # Get latest data for each ticker to avoid duplicates
            latest_data = df.sort_values('timestamp').groupby('ticker').last().reset_index()
            
            # Separate stocks by completeness - most complete at top
            stocks_complete = []      # All 3 valuation methods
            stocks_partial = []       # 1-2 valuation methods
            stocks_without_data = []  # No valuation methods
            
            for _, stock_row in latest_data.iterrows():
                lynch_status = stock_row['lynch_valuation_status'] if 'lynch_valuation_status' in stock_row and pd.notna(stock_row['lynch_valuation_status']) else 'N/A'
                dcf_status = stock_row['dcf_valuation_status'] if 'dcf_valuation_status' in stock_row and pd.notna(stock_row['dcf_valuation_status']) else 'N/A'
                munger_assessment = stock_row['munger_7pct_assessment'] if 'munger_7pct_assessment' in stock_row and pd.notna(stock_row['munger_7pct_assessment']) else 'N/A'
                
                # Count how many valuation methods are available
                available_methods = 0
                if lynch_status != 'N/A':
                    available_methods += 1
                if dcf_status != 'N/A':
                    available_methods += 1
                if munger_assessment != 'N/A':
                    available_methods += 1
                
                if available_methods == 3:
                    stocks_complete.append(stock_row)
                elif available_methods > 0:
                    stocks_partial.append(stock_row)
                else:
                    stocks_without_data.append(stock_row)
            
            # Add stocks with complete data first (all 3 valuation methods)
            for stock_row in stocks_complete:
                ticker = stock_row['ticker']
                company_name = stock_row['company_name']
                current_price = stock_row['current_price']
                
                # Company name and ticker
                ws_summary[f'A{row}'] = f"{company_name} ({ticker})"
                ws_summary[f'A{row}'].font = data_font
                ws_summary[f'A{row}'].border = border
                ws_summary[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Current price
                ws_summary[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_summary[f'E{row}'].font = data_font
                ws_summary[f'E{row}'].border = border
                ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Peter Lynch Valuation (Column B)
                lynch_status = stock_row['lynch_valuation_status'] if 'lynch_valuation_status' in stock_row and pd.notna(stock_row['lynch_valuation_status']) else 'N/A'
                lynch_delta = stock_row['lynch_delta_percentage'] if 'lynch_delta_percentage' in stock_row and pd.notna(stock_row['lynch_delta_percentage']) else 0
                
                if lynch_status != 'N/A':
                    lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
                    ws_summary[f'B{row}'] = lynch_text
                    ws_summary[f'B{row}'].font = data_font
                    ws_summary[f'B{row}'].border = border
                    ws_summary[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in lynch_status.upper() or lynch_delta > 20:
                        ws_summary[f'B{row}'].fill = strong_buy_fill
                    elif 'BUY' in lynch_status.upper() or lynch_delta > 5:
                        ws_summary[f'B{row}'].fill = buy_fill
                    elif 'HOLD' in lynch_status.upper() or abs(lynch_delta) <= 5:
                        ws_summary[f'B{row}'].fill = hold_fill
                    elif 'SELL' in lynch_status.upper() or lynch_delta < -5:
                        ws_summary[f'B{row}'].fill = sell_fill
                    elif 'STRONG SELL' in lynch_status.upper() or lynch_delta < -20:
                        ws_summary[f'B{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'B{row}'] = "Insufficient Data"
                    ws_summary[f'B{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'B{row}'].border = border
                    ws_summary[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'B{row}'].fill = no_data_fill
                
                # DCF Valuation (Column C)
                dcf_status = stock_row['dcf_valuation_status'] if 'dcf_valuation_status' in stock_row and pd.notna(stock_row['dcf_valuation_status']) else 'N/A'
                dcf_delta = stock_row['dcf_delta_percentage'] if 'dcf_delta_percentage' in stock_row and pd.notna(stock_row['dcf_delta_percentage']) else 0
                
                if dcf_status != 'N/A':
                    dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
                    ws_summary[f'C{row}'] = dcf_text
                    ws_summary[f'C{row}'].font = data_font
                    ws_summary[f'C{row}'].border = border
                    ws_summary[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
                        ws_summary[f'C{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
                        ws_summary[f'C{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
                        ws_summary[f'C{row}'].fill = hold_fill
                    elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
                        ws_summary[f'C{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
                        ws_summary[f'C{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'C{row}'] = "Insufficient Data"
                    ws_summary[f'C{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'C{row}'].border = border
                    ws_summary[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'C{row}'].fill = no_data_fill
                
                # Munger Farm Valuation (Column D)
                munger_assessment = stock_row['munger_7pct_assessment'] if 'munger_7pct_assessment' in stock_row and pd.notna(stock_row['munger_7pct_assessment']) else 'N/A'
                munger_delta = stock_row['munger_7pct_delta_percentage'] if 'munger_7pct_delta_percentage' in stock_row and pd.notna(stock_row['munger_7pct_delta_percentage']) else 0
                
                if munger_assessment != 'N/A':
                    munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
                    ws_summary[f'D{row}'] = munger_text
                    ws_summary[f'D{row}'].font = data_font
                    ws_summary[f'D{row}'].border = border
                    ws_summary[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
                        ws_summary[f'D{row}'].fill = strong_buy_fill
                    elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
                        ws_summary[f'D{row}'].fill = buy_fill
                    elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
                        ws_summary[f'D{row}'].fill = hold_fill
                    elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
                        ws_summary[f'D{row}'].fill = sell_fill
                    elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
                        ws_summary[f'D{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'D{row}'] = "Insufficient Data"
                    ws_summary[f'D{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'D{row}'].border = border
                    ws_summary[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'D{row}'].fill = no_data_fill
                
                # Enhanced DCF Valuation (Column E)
                enhanced_dcf_status = stock_row['enhanced_dcf_status'] if 'enhanced_dcf_status' in stock_row and pd.notna(stock_row['enhanced_dcf_status']) else 'N/A'
                enhanced_dcf_delta = stock_row['enhanced_dcf_delta'] if 'enhanced_dcf_delta' in stock_row and pd.notna(stock_row['enhanced_dcf_delta']) else 0
                
                if enhanced_dcf_status != 'N/A':
                    enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
                    ws_summary[f'E{row}'] = enhanced_dcf_text
                    ws_summary[f'E{row}'].font = data_font
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
                        ws_summary[f'E{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
                        ws_summary[f'E{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
                        ws_summary[f'E{row}'].fill = hold_fill
                    elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
                        ws_summary[f'E{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
                        ws_summary[f'E{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'E{row}'] = "Insufficient Data"
                    ws_summary[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'E{row}'].fill = no_data_fill
                
                # Relative Valuation (Column F)
                relative_status = stock_row['relative_valuation_status'] if 'relative_valuation_status' in stock_row and pd.notna(stock_row['relative_valuation_status']) else 'N/A'
                relative_delta = stock_row['relative_valuation_delta'] if 'relative_valuation_delta' in stock_row and pd.notna(stock_row['relative_valuation_delta']) else 0
                
                if relative_status != 'N/A':
                    relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
                    ws_summary[f'F{row}'] = relative_text
                    ws_summary[f'F{row}'].font = data_font
                    ws_summary[f'F{row}'].border = border
                    ws_summary[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
                        ws_summary[f'F{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
                        ws_summary[f'F{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
                        ws_summary[f'F{row}'].fill = hold_fill
                    elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
                        ws_summary[f'F{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
                        ws_summary[f'F{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'F{row}'] = "Insufficient Data"
                    ws_summary[f'F{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'F{row}'].border = border
                    ws_summary[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'F{row}'].fill = no_data_fill
                
                # Reverse DCF Valuation (Column G)
                reverse_dcf_assessment = stock_row['reverse_dcf_assessment'] if 'reverse_dcf_assessment' in stock_row and pd.notna(stock_row['reverse_dcf_assessment']) else 'N/A'
                reverse_dcf_growth = stock_row['reverse_dcf_implied_growth'] if 'reverse_dcf_implied_growth' in stock_row and pd.notna(stock_row['reverse_dcf_implied_growth']) else 0
                
                if reverse_dcf_assessment != 'N/A':
                    reverse_dcf_text = f"{reverse_dcf_assessment}\nGrowth: {reverse_dcf_growth:+.1f}%"
                    ws_summary[f'G{row}'] = reverse_dcf_text
                    ws_summary[f'G{row}'].font = data_font
                    ws_summary[f'G{row}'].border = border
                    ws_summary[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'REASONABLE' in reverse_dcf_assessment.upper():
                        ws_summary[f'G{row}'].fill = hold_fill
                    elif 'UNREASONABLE' in reverse_dcf_assessment.upper():
                        ws_summary[f'G{row}'].fill = sell_fill
                    else:
                        ws_summary[f'G{row}'].fill = no_data_fill
                else:
                    ws_summary[f'G{row}'] = "Insufficient Data"
                    ws_summary[f'G{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'G{row}'].border = border
                    ws_summary[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'G{row}'].fill = no_data_fill
                
                # EPV/RIM Valuation (Column H)
                epv_assessment = stock_row['epv_assessment'] if 'epv_assessment' in stock_row and pd.notna(stock_row['epv_assessment']) else 'N/A'
                epv_delta = stock_row['epv_delta'] if 'epv_delta' in stock_row and pd.notna(stock_row['epv_delta']) else 0
                rim_assessment = stock_row['rim_assessment'] if 'rim_assessment' in stock_row and pd.notna(stock_row['rim_assessment']) else 'N/A'
                rim_delta = stock_row['rim_delta'] if 'rim_delta' in stock_row and pd.notna(stock_row['rim_delta']) else 0
                
                # Use EPV if available, otherwise RIM
                if epv_assessment != 'N/A':
                    epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
                    ws_summary[f'H{row}'] = epv_rim_text
                    ws_summary[f'H{row}'].font = data_font
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
                        ws_summary[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
                        ws_summary[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
                        ws_summary[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
                        ws_summary[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
                        ws_summary[f'H{row}'].fill = strong_sell_fill
                elif rim_assessment != 'N/A':
                    epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
                    ws_summary[f'H{row}'] = epv_rim_text
                    ws_summary[f'H{row}'].font = data_font
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
                        ws_summary[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
                        ws_summary[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
                        ws_summary[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
                        ws_summary[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
                        ws_summary[f'H{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'H{row}'] = "Insufficient Data"
                    ws_summary[f'H{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'H{row}'].fill = no_data_fill
                
                # Current Price (Column I)
                ws_summary[f'I{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_summary[f'I{row}'].font = data_font
                ws_summary[f'I{row}'].border = border
                ws_summary[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Add stocks with partial data (1-2 valuation methods)
            for stock_row in stocks_partial:
                ticker = stock_row['ticker']
                company_name = stock_row['company_name']
                current_price = stock_row['current_price']
                
                # Company name and ticker
                ws_summary[f'A{row}'] = f"{company_name} ({ticker})"
                ws_summary[f'A{row}'].font = data_font
                ws_summary[f'A{row}'].border = border
                ws_summary[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Current price
                ws_summary[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_summary[f'E{row}'].font = data_font
                ws_summary[f'E{row}'].border = border
                ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Peter Lynch Valuation (Column B)
                lynch_status = stock_row['lynch_valuation_status'] if 'lynch_valuation_status' in stock_row and pd.notna(stock_row['lynch_valuation_status']) else 'N/A'
                lynch_delta = stock_row['lynch_delta_percentage'] if 'lynch_delta_percentage' in stock_row and pd.notna(stock_row['lynch_delta_percentage']) else 0
                
                if lynch_status != 'N/A':
                    lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
                    ws_summary[f'B{row}'] = lynch_text
                    ws_summary[f'B{row}'].font = data_font
                    ws_summary[f'B{row}'].border = border
                    ws_summary[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in lynch_status.upper() or lynch_delta > 20:
                        ws_summary[f'B{row}'].fill = strong_buy_fill
                    elif 'BUY' in lynch_status.upper() or lynch_delta > 5:
                        ws_summary[f'B{row}'].fill = buy_fill
                    elif 'HOLD' in lynch_status.upper() or abs(lynch_delta) <= 5:
                        ws_summary[f'B{row}'].fill = hold_fill
                    elif 'SELL' in lynch_status.upper() or lynch_delta < -5:
                        ws_summary[f'B{row}'].fill = sell_fill
                    elif 'STRONG SELL' in lynch_status.upper() or lynch_delta < -20:
                        ws_summary[f'B{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'B{row}'] = "Insufficient Data"
                    ws_summary[f'B{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'B{row}'].border = border
                    ws_summary[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'B{row}'].fill = no_data_fill
                
                # DCF Valuation (Column C)
                dcf_status = stock_row['dcf_valuation_status'] if 'dcf_valuation_status' in stock_row and pd.notna(stock_row['dcf_valuation_status']) else 'N/A'
                dcf_delta = stock_row['dcf_delta_percentage'] if 'dcf_delta_percentage' in stock_row and pd.notna(stock_row['dcf_delta_percentage']) else 0
                
                if dcf_status != 'N/A':
                    dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
                    ws_summary[f'C{row}'] = dcf_text
                    ws_summary[f'C{row}'].font = data_font
                    ws_summary[f'C{row}'].border = border
                    ws_summary[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
                        ws_summary[f'C{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
                        ws_summary[f'C{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
                        ws_summary[f'C{row}'].fill = hold_fill
                    elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
                        ws_summary[f'C{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
                        ws_summary[f'C{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'C{row}'] = "Insufficient Data"
                    ws_summary[f'C{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'C{row}'].border = border
                    ws_summary[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'C{row}'].fill = no_data_fill
                
                # Munger Farm Valuation (Column D)
                munger_assessment = stock_row['munger_7pct_assessment'] if 'munger_7pct_assessment' in stock_row and pd.notna(stock_row['munger_7pct_assessment']) else 'N/A'
                munger_delta = stock_row['munger_7pct_delta_percentage'] if 'munger_7pct_delta_percentage' in stock_row and pd.notna(stock_row['munger_7pct_delta_percentage']) else 0
                
                if munger_assessment != 'N/A':
                    munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
                    ws_summary[f'D{row}'] = munger_text
                    ws_summary[f'D{row}'].font = data_font
                    ws_summary[f'D{row}'].border = border
                    ws_summary[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
                        ws_summary[f'D{row}'].fill = strong_buy_fill
                    elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
                        ws_summary[f'D{row}'].fill = buy_fill
                    elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
                        ws_summary[f'D{row}'].fill = hold_fill
                    elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
                        ws_summary[f'D{row}'].fill = sell_fill
                    elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
                        ws_summary[f'D{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'D{row}'] = "Insufficient Data"
                    ws_summary[f'D{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'D{row}'].border = border
                    ws_summary[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'D{row}'].fill = no_data_fill
                
                # Enhanced DCF Valuation (Column E)
                enhanced_dcf_status = stock_row['enhanced_dcf_status'] if 'enhanced_dcf_status' in stock_row and pd.notna(stock_row['enhanced_dcf_status']) else 'N/A'
                enhanced_dcf_delta = stock_row['enhanced_dcf_delta'] if 'enhanced_dcf_delta' in stock_row and pd.notna(stock_row['enhanced_dcf_delta']) else 0
                
                if enhanced_dcf_status != 'N/A':
                    enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
                    ws_summary[f'E{row}'] = enhanced_dcf_text
                    ws_summary[f'E{row}'].font = data_font
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
                        ws_summary[f'E{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
                        ws_summary[f'E{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
                        ws_summary[f'E{row}'].fill = hold_fill
                    elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
                        ws_summary[f'E{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
                        ws_summary[f'E{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'E{row}'] = "Insufficient Data"
                    ws_summary[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'E{row}'].fill = no_data_fill
                
                # Relative Valuation (Column F)
                relative_status = stock_row['relative_valuation_status'] if 'relative_valuation_status' in stock_row and pd.notna(stock_row['relative_valuation_status']) else 'N/A'
                relative_delta = stock_row['relative_valuation_delta'] if 'relative_valuation_delta' in stock_row and pd.notna(stock_row['relative_valuation_delta']) else 0
                
                if relative_status != 'N/A':
                    relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
                    ws_summary[f'F{row}'] = relative_text
                    ws_summary[f'F{row}'].font = data_font
                    ws_summary[f'F{row}'].border = border
                    ws_summary[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
                        ws_summary[f'F{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
                        ws_summary[f'F{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
                        ws_summary[f'F{row}'].fill = hold_fill
                    elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
                        ws_summary[f'F{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
                        ws_summary[f'F{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'F{row}'] = "Insufficient Data"
                    ws_summary[f'F{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'F{row}'].border = border
                    ws_summary[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'F{row}'].fill = no_data_fill
                
                # Reverse DCF Valuation (Column G)
                reverse_dcf_assessment = stock_row['reverse_dcf_assessment'] if 'reverse_dcf_assessment' in stock_row and pd.notna(stock_row['reverse_dcf_assessment']) else 'N/A'
                reverse_dcf_growth = stock_row['reverse_dcf_implied_growth'] if 'reverse_dcf_implied_growth' in stock_row and pd.notna(stock_row['reverse_dcf_implied_growth']) else 0
                
                if reverse_dcf_assessment != 'N/A':
                    reverse_dcf_text = f"{reverse_dcf_assessment}\nGrowth: {reverse_dcf_growth:+.1f}%"
                    ws_summary[f'G{row}'] = reverse_dcf_text
                    ws_summary[f'G{row}'].font = data_font
                    ws_summary[f'G{row}'].border = border
                    ws_summary[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'REASONABLE' in reverse_dcf_assessment.upper():
                        ws_summary[f'G{row}'].fill = hold_fill
                    elif 'UNREASONABLE' in reverse_dcf_assessment.upper():
                        ws_summary[f'G{row}'].fill = sell_fill
                    else:
                        ws_summary[f'G{row}'].fill = no_data_fill
                else:
                    ws_summary[f'G{row}'] = "Insufficient Data"
                    ws_summary[f'G{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'G{row}'].border = border
                    ws_summary[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'G{row}'].fill = no_data_fill
                
                # EPV/RIM Valuation (Column H)
                epv_assessment = stock_row['epv_assessment'] if 'epv_assessment' in stock_row and pd.notna(stock_row['epv_assessment']) else 'N/A'
                epv_delta = stock_row['epv_delta'] if 'epv_delta' in stock_row and pd.notna(stock_row['epv_delta']) else 0
                rim_assessment = stock_row['rim_assessment'] if 'rim_assessment' in stock_row and pd.notna(stock_row['rim_assessment']) else 'N/A'
                rim_delta = stock_row['rim_delta'] if 'rim_delta' in stock_row and pd.notna(stock_row['rim_delta']) else 0
                
                # Use EPV if available, otherwise RIM
                if epv_assessment != 'N/A':
                    epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
                    ws_summary[f'H{row}'] = epv_rim_text
                    ws_summary[f'H{row}'].font = data_font
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
                        ws_summary[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
                        ws_summary[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
                        ws_summary[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
                        ws_summary[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
                        ws_summary[f'H{row}'].fill = strong_sell_fill
                elif rim_assessment != 'N/A':
                    epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
                    ws_summary[f'H{row}'] = epv_rim_text
                    ws_summary[f'H{row}'].font = data_font
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
                        ws_summary[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
                        ws_summary[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
                        ws_summary[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
                        ws_summary[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
                        ws_summary[f'H{row}'].fill = strong_sell_fill
                else:
                    ws_summary[f'H{row}'] = "Insufficient Data"
                    ws_summary[f'H{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'H{row}'].border = border
                    ws_summary[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_summary[f'H{row}'].fill = no_data_fill
                
                # Current Price (Column I)
                ws_summary[f'I{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_summary[f'I{row}'].font = data_font
                ws_summary[f'I{row}'].border = border
                ws_summary[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Add stocks with insufficient data at the bottom
            if stocks_without_data:
                row += 2
                ws_summary[f'A{row}'] = "STOCKS WITH INSUFFICIENT DATA"
                ws_summary[f'A{row}'].font = Font(bold=True, size=12, color="808080")
                ws_summary[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                ws_summary[f'A{row}'].border = border
                ws_summary[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for stock_row in stocks_without_data:
                    ticker = stock_row['ticker']
                    company_name = stock_row['company_name']
                    current_price = stock_row['current_price']
                    
                    # Company name and ticker
                    ws_summary[f'A{row}'] = f"{company_name} ({ticker})"
                    ws_summary[f'A{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'A{row}'].border = border
                    ws_summary[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Current price
                    ws_summary[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                    ws_summary[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_summary[f'E{row}'].border = border
                    ws_summary[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # All valuation columns show "Insufficient Data"
                    for col in ['B', 'C', 'D']:
                        ws_summary[f'{col}{row}'] = "Insufficient Data"
                        ws_summary[f'{col}{row}'].font = Font(size=11, color="808080")  # Grey text
                        ws_summary[f'{col}{row}'].border = border
                        ws_summary[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                        ws_summary[f'{col}{row}'].fill = no_data_fill
                    
                    row += 1
            
            # Add legend
            row += 1
            ws_summary[f'A{row}'] = "Legend:"
            ws_summary[f'A{row}'].font = Font(bold=True, size=12)
            ws_summary[f'A{row}'].border = border
            row += 1
            
            legend_items = [
                ("Strong Buy", strong_buy_fill),
                ("Buy", buy_fill),
                ("Hold", hold_fill),
                ("Sell", sell_fill),
                ("Strong Sell", strong_sell_fill),
                ("Insufficient Data", no_data_fill)
            ]
            
            for item_text, item_fill in legend_items:
                ws_summary[f'A{row}'] = item_text
                ws_summary[f'A{row}'].font = data_font
                ws_summary[f'A{row}'].fill = item_fill
                ws_summary[f'A{row}'].border = border
                ws_summary[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                row += 1
            
            # Set optimal column widths for readability
            ws_summary.column_dimensions['A'].width = 45  # Company name and ticker
            ws_summary.column_dimensions['B'].width = 35  # Peter Lynch valuation
            ws_summary.column_dimensions['C'].width = 35  # DCF valuation
            ws_summary.column_dimensions['D'].width = 35  # Munger Farm valuation
            ws_summary.column_dimensions['E'].width = 35  # Enhanced DCF valuation
            ws_summary.column_dimensions['F'].width = 35  # Relative Valuation
            ws_summary.column_dimensions['G'].width = 35  # Reverse DCF valuation
            ws_summary.column_dimensions['H'].width = 35  # EPV/RIM valuation
            ws_summary.column_dimensions['I'].width = 20  # Current price
            
            self.logger.info("Created valuation summary sheet with conditional formatting")
            
        except Exception as e:
            self.logger.error(f"Error creating valuation summary sheet: {e}")

    def create_prospects_sheet(self, df, wb):
        """Create prospects sheet ranked by undervaluation"""
        try:
            ws_prospects = wb.create_sheet('Prospects', 1)
            
            # Define consistent color scheme across all sheets
            header_fill = PatternFill(start_color="2F4F4F", end_color="2F4F4F", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=14)
            subheader_font = Font(bold=True, size=12)
            data_font = Font(size=11)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Improved conditional formatting fills with better contrast
            strong_buy_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")  # Forest Green
            buy_fill = PatternFill(start_color="32CD32", end_color="32CD32", fill_type="solid")  # Lime Green
            hold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Gold
            sell_fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")  # Tomato
            strong_sell_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")  # Crimson
            no_data_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Light Gray
            
            # Title
            ws_prospects['A1'] = "NZX PROSPECTS - UNDERVALUED STOCKS ONLY (Δ > 10%)"
            ws_prospects['A1'].font = header_font
            ws_prospects['A1'].fill = header_fill
            ws_prospects['A1'].border = border
            ws_prospects['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws_prospects.merge_cells('A1:I1')
            
            # Headers for all 9 valuation methods
            headers = ['Company (Ticker)', 'Peter Lynch', 'DCF Valuation', 'Munger Farm', 'Enhanced DCF', 'Relative Valuation', 'Reverse DCF', 'EPV/RIM', 'Current Price']
            row = 3
            for col, header in enumerate(headers, 1):
                cell = ws_prospects.cell(row=row, column=col)
                cell.value = header
                cell.font = subheader_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            row += 1
            
            # Calculate undervaluation scores for ranking
            def calculate_undervaluation_score(stock_row):
                """Calculate composite undervaluation score"""
                score = 0
                count = 0
                
                # Peter Lynch score
                lynch_delta = stock_row['lynch_delta_percentage'] if 'lynch_delta_percentage' in stock_row and pd.notna(stock_row['lynch_delta_percentage']) else 0
                if lynch_delta != 0:
                    score += lynch_delta
                    count += 1
                
                # DCF score
                dcf_delta = stock_row['dcf_delta_percentage'] if 'dcf_delta_percentage' in stock_row and pd.notna(stock_row['dcf_delta_percentage']) else 0
                if dcf_delta != 0:
                    score += dcf_delta
                    count += 1
                
                # Munger score
                munger_delta = stock_row['munger_7pct_delta_percentage'] if 'munger_7pct_delta_percentage' in stock_row and pd.notna(stock_row['munger_7pct_delta_percentage']) else 0
                if munger_delta != 0:
                    score += munger_delta
                    count += 1
                
                return score / count if count > 0 else -999  # -999 for insufficient data
            
            # Get latest data for each ticker to avoid duplicates
            latest_data = df.sort_values('timestamp').groupby('ticker').last().reset_index()
            
            # Separate stocks by completeness and rank by undervaluation
            stocks_complete = []
            stocks_partial = []
            stocks_without_data = []
            
            for _, stock_row in latest_data.iterrows():
                lynch_status = stock_row['lynch_valuation_status'] if 'lynch_valuation_status' in stock_row and pd.notna(stock_row['lynch_valuation_status']) else 'N/A'
                dcf_status = stock_row['dcf_valuation_status'] if 'dcf_valuation_status' in stock_row and pd.notna(stock_row['dcf_valuation_status']) else 'N/A'
                munger_assessment = stock_row['munger_7pct_assessment'] if 'munger_7pct_assessment' in stock_row and pd.notna(stock_row['munger_7pct_assessment']) else 'N/A'
                
                # Count how many valuation methods are available
                available_methods = 0
                if lynch_status != 'N/A':
                    available_methods += 1
                if dcf_status != 'N/A':
                    available_methods += 1
                if munger_assessment != 'N/A':
                    available_methods += 1
                
                # Calculate undervaluation score
                undervaluation_score = calculate_undervaluation_score(stock_row)
                stock_row['undervaluation_score'] = undervaluation_score
                
                if available_methods == 3:
                    stocks_complete.append(stock_row)
                elif available_methods > 0:
                    stocks_partial.append(stock_row)
                else:
                    stocks_without_data.append(stock_row)
            
            # Sort by undervaluation score (highest first)
            stocks_complete.sort(key=lambda x: x['undervaluation_score'], reverse=True)
            stocks_partial.sort(key=lambda x: x['undervaluation_score'], reverse=True)
            
            # Add stocks with complete data first (ranked by undervaluation) - Only show undervalued stocks (delta > 10%)
            for stock_row in stocks_complete:
                # Only include stocks with significant undervaluation
                if stock_row['undervaluation_score'] <= 10:  # Skip stocks with <= 10% undervaluation
                    continue
                ticker = stock_row['ticker']
                company_name = stock_row['company_name']
                current_price = stock_row['current_price']
                
                # Company name and ticker
                ws_prospects[f'A{row}'] = f"{company_name} ({ticker})"
                ws_prospects[f'A{row}'].font = data_font
                ws_prospects[f'A{row}'].border = border
                ws_prospects[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Current price
                ws_prospects[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_prospects[f'E{row}'].font = data_font
                ws_prospects[f'E{row}'].border = border
                ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Peter Lynch Valuation (Column B)
                lynch_status = stock_row['lynch_valuation_status'] if 'lynch_valuation_status' in stock_row and pd.notna(stock_row['lynch_valuation_status']) else 'N/A'
                lynch_delta = stock_row['lynch_delta_percentage'] if 'lynch_delta_percentage' in stock_row and pd.notna(stock_row['lynch_delta_percentage']) else 0
                
                if lynch_status != 'N/A':
                    lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
                    ws_prospects[f'B{row}'] = lynch_text
                    ws_prospects[f'B{row}'].font = data_font
                    ws_prospects[f'B{row}'].border = border
                    ws_prospects[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in lynch_status.upper() or lynch_delta > 20:
                        ws_prospects[f'B{row}'].fill = strong_buy_fill
                    elif 'BUY' in lynch_status.upper() or lynch_delta > 5:
                        ws_prospects[f'B{row}'].fill = buy_fill
                    elif 'HOLD' in lynch_status.upper() or abs(lynch_delta) <= 5:
                        ws_prospects[f'B{row}'].fill = hold_fill
                    elif 'SELL' in lynch_status.upper() or lynch_delta < -5:
                        ws_prospects[f'B{row}'].fill = sell_fill
                    elif 'STRONG SELL' in lynch_status.upper() or lynch_delta < -20:
                        ws_prospects[f'B{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'B{row}'] = "Insufficient Data"
                    ws_prospects[f'B{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'B{row}'].border = border
                    ws_prospects[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'B{row}'].fill = no_data_fill
                
                # DCF Valuation (Column C)
                dcf_status = stock_row['dcf_valuation_status'] if 'dcf_valuation_status' in stock_row and pd.notna(stock_row['dcf_valuation_status']) else 'N/A'
                dcf_delta = stock_row['dcf_delta_percentage'] if 'dcf_delta_percentage' in stock_row and pd.notna(stock_row['dcf_delta_percentage']) else 0
                
                if dcf_status != 'N/A':
                    dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
                    ws_prospects[f'C{row}'] = dcf_text
                    ws_prospects[f'C{row}'].font = data_font
                    ws_prospects[f'C{row}'].border = border
                    ws_prospects[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
                        ws_prospects[f'C{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
                        ws_prospects[f'C{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
                        ws_prospects[f'C{row}'].fill = hold_fill
                    elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
                        ws_prospects[f'C{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
                        ws_prospects[f'C{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'C{row}'] = "Insufficient Data"
                    ws_prospects[f'C{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'C{row}'].border = border
                    ws_prospects[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'C{row}'].fill = no_data_fill
                
                # Munger Farm Valuation (Column D)
                munger_assessment = stock_row['munger_7pct_assessment'] if 'munger_7pct_assessment' in stock_row and pd.notna(stock_row['munger_7pct_assessment']) else 'N/A'
                munger_delta = stock_row['munger_7pct_delta_percentage'] if 'munger_7pct_delta_percentage' in stock_row and pd.notna(stock_row['munger_7pct_delta_percentage']) else 0
                
                if munger_assessment != 'N/A':
                    munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
                    ws_prospects[f'D{row}'] = munger_text
                    ws_prospects[f'D{row}'].font = data_font
                    ws_prospects[f'D{row}'].border = border
                    ws_prospects[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
                        ws_prospects[f'D{row}'].fill = strong_buy_fill
                    elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
                        ws_prospects[f'D{row}'].fill = buy_fill
                    elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
                        ws_prospects[f'D{row}'].fill = hold_fill
                    elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
                        ws_prospects[f'D{row}'].fill = sell_fill
                    elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
                        ws_prospects[f'D{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'D{row}'] = "Insufficient Data"
                    ws_prospects[f'D{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'D{row}'].border = border
                    ws_prospects[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'D{row}'].fill = no_data_fill
                
                # Enhanced DCF Valuation (Column E)
                enhanced_dcf_status = stock_row['enhanced_dcf_status'] if 'enhanced_dcf_status' in stock_row and pd.notna(stock_row['enhanced_dcf_status']) else 'N/A'
                enhanced_dcf_delta = stock_row['enhanced_dcf_delta'] if 'enhanced_dcf_delta' in stock_row and pd.notna(stock_row['enhanced_dcf_delta']) else 0
                
                if enhanced_dcf_status != 'N/A':
                    enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
                    ws_prospects[f'E{row}'] = enhanced_dcf_text
                    ws_prospects[f'E{row}'].font = data_font
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
                        ws_prospects[f'E{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
                        ws_prospects[f'E{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
                        ws_prospects[f'E{row}'].fill = hold_fill
                    elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
                        ws_prospects[f'E{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
                        ws_prospects[f'E{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'E{row}'] = "Insufficient Data"
                    ws_prospects[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'E{row}'].fill = no_data_fill
                
                # Relative Valuation (Column F)
                relative_status = stock_row['relative_valuation_status'] if 'relative_valuation_status' in stock_row and pd.notna(stock_row['relative_valuation_status']) else 'N/A'
                relative_delta = stock_row['relative_valuation_delta'] if 'relative_valuation_delta' in stock_row and pd.notna(stock_row['relative_valuation_delta']) else 0
                
                if relative_status != 'N/A':
                    relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
                    ws_prospects[f'F{row}'] = relative_text
                    ws_prospects[f'F{row}'].font = data_font
                    ws_prospects[f'F{row}'].border = border
                    ws_prospects[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
                        ws_prospects[f'F{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
                        ws_prospects[f'F{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
                        ws_prospects[f'F{row}'].fill = hold_fill
                    elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
                        ws_prospects[f'F{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
                        ws_prospects[f'F{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'F{row}'] = "Insufficient Data"
                    ws_prospects[f'F{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'F{row}'].border = border
                    ws_prospects[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'F{row}'].fill = no_data_fill
                
                # Reverse DCF Valuation (Column G)
                reverse_dcf_assessment = stock_row['reverse_dcf_assessment'] if 'reverse_dcf_assessment' in stock_row and pd.notna(stock_row['reverse_dcf_assessment']) else 'N/A'
                reverse_dcf_growth = stock_row['reverse_dcf_implied_growth'] if 'reverse_dcf_implied_growth' in stock_row and pd.notna(stock_row['reverse_dcf_implied_growth']) else 0
                
                if reverse_dcf_assessment != 'N/A':
                    reverse_dcf_text = f"{reverse_dcf_assessment}\nGrowth: {reverse_dcf_growth:+.1f}%"
                    ws_prospects[f'G{row}'] = reverse_dcf_text
                    ws_prospects[f'G{row}'].font = data_font
                    ws_prospects[f'G{row}'].border = border
                    ws_prospects[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'REASONABLE' in reverse_dcf_assessment.upper():
                        ws_prospects[f'G{row}'].fill = hold_fill
                    elif 'UNREASONABLE' in reverse_dcf_assessment.upper():
                        ws_prospects[f'G{row}'].fill = sell_fill
                    else:
                        ws_prospects[f'G{row}'].fill = no_data_fill
                else:
                    ws_prospects[f'G{row}'] = "Insufficient Data"
                    ws_prospects[f'G{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'G{row}'].border = border
                    ws_prospects[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'G{row}'].fill = no_data_fill
                
                # EPV/RIM Valuation (Column H)
                epv_assessment = stock_row['epv_assessment'] if 'epv_assessment' in stock_row and pd.notna(stock_row['epv_assessment']) else 'N/A'
                epv_delta = stock_row['epv_delta'] if 'epv_delta' in stock_row and pd.notna(stock_row['epv_delta']) else 0
                rim_assessment = stock_row['rim_assessment'] if 'rim_assessment' in stock_row and pd.notna(stock_row['rim_assessment']) else 'N/A'
                rim_delta = stock_row['rim_delta'] if 'rim_delta' in stock_row and pd.notna(stock_row['rim_delta']) else 0
                
                # Use EPV if available, otherwise RIM
                if epv_assessment != 'N/A':
                    epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
                    ws_prospects[f'H{row}'] = epv_rim_text
                    ws_prospects[f'H{row}'].font = data_font
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
                        ws_prospects[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
                        ws_prospects[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
                        ws_prospects[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
                        ws_prospects[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
                        ws_prospects[f'H{row}'].fill = strong_sell_fill
                elif rim_assessment != 'N/A':
                    epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
                    ws_prospects[f'H{row}'] = epv_rim_text
                    ws_prospects[f'H{row}'].font = data_font
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
                        ws_prospects[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
                        ws_prospects[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
                        ws_prospects[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
                        ws_prospects[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
                        ws_prospects[f'H{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'H{row}'] = "Insufficient Data"
                    ws_prospects[f'H{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'H{row}'].fill = no_data_fill
                
                # Current Price (Column I)
                ws_prospects[f'I{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_prospects[f'I{row}'].font = data_font
                ws_prospects[f'I{row}'].border = border
                ws_prospects[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Add stocks with partial data (ranked by undervaluation)
            for stock_row in stocks_partial:
                ticker = stock_row['ticker']
                company_name = stock_row['company_name']
                current_price = stock_row['current_price']
                
                # Company name and ticker
                ws_prospects[f'A{row}'] = f"{company_name} ({ticker})"
                ws_prospects[f'A{row}'].font = data_font
                ws_prospects[f'A{row}'].border = border
                ws_prospects[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                # Current price
                ws_prospects[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_prospects[f'E{row}'].font = data_font
                ws_prospects[f'E{row}'].border = border
                ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Peter Lynch Valuation (Column B)
                lynch_status = stock_row['lynch_valuation_status'] if 'lynch_valuation_status' in stock_row and pd.notna(stock_row['lynch_valuation_status']) else 'N/A'
                lynch_delta = stock_row['lynch_delta_percentage'] if 'lynch_delta_percentage' in stock_row and pd.notna(stock_row['lynch_delta_percentage']) else 0
                
                if lynch_status != 'N/A':
                    lynch_text = f"{lynch_status}\nDelta: {lynch_delta:+.1f}%"
                    ws_prospects[f'B{row}'] = lynch_text
                    ws_prospects[f'B{row}'].font = data_font
                    ws_prospects[f'B{row}'].border = border
                    ws_prospects[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in lynch_status.upper() or lynch_delta > 20:
                        ws_prospects[f'B{row}'].fill = strong_buy_fill
                    elif 'BUY' in lynch_status.upper() or lynch_delta > 5:
                        ws_prospects[f'B{row}'].fill = buy_fill
                    elif 'HOLD' in lynch_status.upper() or abs(lynch_delta) <= 5:
                        ws_prospects[f'B{row}'].fill = hold_fill
                    elif 'SELL' in lynch_status.upper() or lynch_delta < -5:
                        ws_prospects[f'B{row}'].fill = sell_fill
                    elif 'STRONG SELL' in lynch_status.upper() or lynch_delta < -20:
                        ws_prospects[f'B{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'B{row}'] = "Insufficient Data"
                    ws_prospects[f'B{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'B{row}'].border = border
                    ws_prospects[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'B{row}'].fill = no_data_fill
                
                # DCF Valuation (Column C)
                dcf_status = stock_row['dcf_valuation_status'] if 'dcf_valuation_status' in stock_row and pd.notna(stock_row['dcf_valuation_status']) else 'N/A'
                dcf_delta = stock_row['dcf_delta_percentage'] if 'dcf_delta_percentage' in stock_row and pd.notna(stock_row['dcf_delta_percentage']) else 0
                
                if dcf_status != 'N/A':
                    dcf_text = f"{dcf_status}\nDelta: {dcf_delta:+.1f}%"
                    ws_prospects[f'C{row}'] = dcf_text
                    ws_prospects[f'C{row}'].font = data_font
                    ws_prospects[f'C{row}'].border = border
                    ws_prospects[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in dcf_status.upper() or dcf_delta > 20:
                        ws_prospects[f'C{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in dcf_status.upper() or dcf_delta > 5:
                        ws_prospects[f'C{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in dcf_status.upper() or abs(dcf_delta) <= 5:
                        ws_prospects[f'C{row}'].fill = hold_fill
                    elif 'OVERVALUED' in dcf_status.upper() or dcf_delta < -5:
                        ws_prospects[f'C{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in dcf_status.upper() or dcf_delta < -20:
                        ws_prospects[f'C{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'C{row}'] = "Insufficient Data"
                    ws_prospects[f'C{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'C{row}'].border = border
                    ws_prospects[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'C{row}'].fill = no_data_fill
                
                # Munger Farm Valuation (Column D)
                munger_assessment = stock_row['munger_7pct_assessment'] if 'munger_7pct_assessment' in stock_row and pd.notna(stock_row['munger_7pct_assessment']) else 'N/A'
                munger_delta = stock_row['munger_7pct_delta_percentage'] if 'munger_7pct_delta_percentage' in stock_row and pd.notna(stock_row['munger_7pct_delta_percentage']) else 0
                
                if munger_assessment != 'N/A':
                    munger_text = f"{munger_assessment}\nDelta: {munger_delta:+.1f}%"
                    ws_prospects[f'D{row}'] = munger_text
                    ws_prospects[f'D{row}'].font = data_font
                    ws_prospects[f'D{row}'].border = border
                    ws_prospects[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'STRONG BUY' in munger_assessment.upper() or munger_delta > 20:
                        ws_prospects[f'D{row}'].fill = strong_buy_fill
                    elif 'BUY' in munger_assessment.upper() or munger_delta > 5:
                        ws_prospects[f'D{row}'].fill = buy_fill
                    elif 'HOLD' in munger_assessment.upper() or abs(munger_delta) <= 5:
                        ws_prospects[f'D{row}'].fill = hold_fill
                    elif 'SELL' in munger_assessment.upper() or munger_delta < -5:
                        ws_prospects[f'D{row}'].fill = sell_fill
                    elif 'STRONG SELL' in munger_assessment.upper() or munger_delta < -20:
                        ws_prospects[f'D{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'D{row}'] = "Insufficient Data"
                    ws_prospects[f'D{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'D{row}'].border = border
                    ws_prospects[f'D{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'D{row}'].fill = no_data_fill
                
                # Enhanced DCF Valuation (Column E)
                enhanced_dcf_status = stock_row['enhanced_dcf_status'] if 'enhanced_dcf_status' in stock_row and pd.notna(stock_row['enhanced_dcf_status']) else 'N/A'
                enhanced_dcf_delta = stock_row['enhanced_dcf_delta'] if 'enhanced_dcf_delta' in stock_row and pd.notna(stock_row['enhanced_dcf_delta']) else 0
                
                if enhanced_dcf_status != 'N/A':
                    enhanced_dcf_text = f"{enhanced_dcf_status}\nDelta: {enhanced_dcf_delta:+.1f}%"
                    ws_prospects[f'E{row}'] = enhanced_dcf_text
                    ws_prospects[f'E{row}'].font = data_font
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 20:
                        ws_prospects[f'E{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta > 5:
                        ws_prospects[f'E{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in enhanced_dcf_status.upper() or abs(enhanced_dcf_delta) <= 5:
                        ws_prospects[f'E{row}'].fill = hold_fill
                    elif 'OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -5:
                        ws_prospects[f'E{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in enhanced_dcf_status.upper() or enhanced_dcf_delta < -20:
                        ws_prospects[f'E{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'E{row}'] = "Insufficient Data"
                    ws_prospects[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'E{row}'].fill = no_data_fill
                
                # Relative Valuation (Column F)
                relative_status = stock_row['relative_valuation_status'] if 'relative_valuation_status' in stock_row and pd.notna(stock_row['relative_valuation_status']) else 'N/A'
                relative_delta = stock_row['relative_valuation_delta'] if 'relative_valuation_delta' in stock_row and pd.notna(stock_row['relative_valuation_delta']) else 0
                
                if relative_status != 'N/A':
                    relative_text = f"{relative_status}\nDelta: {relative_delta:+.1f}%"
                    ws_prospects[f'F{row}'] = relative_text
                    ws_prospects[f'F{row}'].font = data_font
                    ws_prospects[f'F{row}'].border = border
                    ws_prospects[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in relative_status.upper() or relative_delta > 20:
                        ws_prospects[f'F{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in relative_status.upper() or relative_delta > 5:
                        ws_prospects[f'F{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in relative_status.upper() or abs(relative_delta) <= 5:
                        ws_prospects[f'F{row}'].fill = hold_fill
                    elif 'OVERVALUED' in relative_status.upper() or relative_delta < -5:
                        ws_prospects[f'F{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in relative_status.upper() or relative_delta < -20:
                        ws_prospects[f'F{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'F{row}'] = "Insufficient Data"
                    ws_prospects[f'F{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'F{row}'].border = border
                    ws_prospects[f'F{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'F{row}'].fill = no_data_fill
                
                # Reverse DCF Valuation (Column G)
                reverse_dcf_assessment = stock_row['reverse_dcf_assessment'] if 'reverse_dcf_assessment' in stock_row and pd.notna(stock_row['reverse_dcf_assessment']) else 'N/A'
                reverse_dcf_growth = stock_row['reverse_dcf_implied_growth'] if 'reverse_dcf_implied_growth' in stock_row and pd.notna(stock_row['reverse_dcf_implied_growth']) else 0
                
                if reverse_dcf_assessment != 'N/A':
                    reverse_dcf_text = f"{reverse_dcf_assessment}\nGrowth: {reverse_dcf_growth:+.1f}%"
                    ws_prospects[f'G{row}'] = reverse_dcf_text
                    ws_prospects[f'G{row}'].font = data_font
                    ws_prospects[f'G{row}'].border = border
                    ws_prospects[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'REASONABLE' in reverse_dcf_assessment.upper():
                        ws_prospects[f'G{row}'].fill = hold_fill
                    elif 'UNREASONABLE' in reverse_dcf_assessment.upper():
                        ws_prospects[f'G{row}'].fill = sell_fill
                    else:
                        ws_prospects[f'G{row}'].fill = no_data_fill
                else:
                    ws_prospects[f'G{row}'] = "Insufficient Data"
                    ws_prospects[f'G{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'G{row}'].border = border
                    ws_prospects[f'G{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'G{row}'].fill = no_data_fill
                
                # EPV/RIM Valuation (Column H)
                epv_assessment = stock_row['epv_assessment'] if 'epv_assessment' in stock_row and pd.notna(stock_row['epv_assessment']) else 'N/A'
                epv_delta = stock_row['epv_delta'] if 'epv_delta' in stock_row and pd.notna(stock_row['epv_delta']) else 0
                rim_assessment = stock_row['rim_assessment'] if 'rim_assessment' in stock_row and pd.notna(stock_row['rim_assessment']) else 'N/A'
                rim_delta = stock_row['rim_delta'] if 'rim_delta' in stock_row and pd.notna(stock_row['rim_delta']) else 0
                
                # Use EPV if available, otherwise RIM
                if epv_assessment != 'N/A':
                    epv_rim_text = f"EPV: {epv_assessment}\nDelta: {epv_delta:+.1f}%"
                    ws_prospects[f'H{row}'] = epv_rim_text
                    ws_prospects[f'H{row}'].font = data_font
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in epv_assessment.upper() or epv_delta > 20:
                        ws_prospects[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in epv_assessment.upper() or epv_delta > 5:
                        ws_prospects[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in epv_assessment.upper() or abs(epv_delta) <= 5:
                        ws_prospects[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in epv_assessment.upper() or epv_delta < -5:
                        ws_prospects[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in epv_assessment.upper() or epv_delta < -20:
                        ws_prospects[f'H{row}'].fill = strong_sell_fill
                elif rim_assessment != 'N/A':
                    epv_rim_text = f"RIM: {rim_assessment}\nDelta: {rim_delta:+.1f}%"
                    ws_prospects[f'H{row}'] = epv_rim_text
                    ws_prospects[f'H{row}'].font = data_font
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Apply conditional formatting
                    if 'SIGNIFICANTLY UNDERVALUED' in rim_assessment.upper() or rim_delta > 20:
                        ws_prospects[f'H{row}'].fill = strong_buy_fill
                    elif 'UNDERVALUED' in rim_assessment.upper() or rim_delta > 5:
                        ws_prospects[f'H{row}'].fill = buy_fill
                    elif 'FAIRLY VALUED' in rim_assessment.upper() or abs(rim_delta) <= 5:
                        ws_prospects[f'H{row}'].fill = hold_fill
                    elif 'OVERVALUED' in rim_assessment.upper() or rim_delta < -5:
                        ws_prospects[f'H{row}'].fill = sell_fill
                    elif 'SIGNIFICANTLY OVERVALUED' in rim_assessment.upper() or rim_delta < -20:
                        ws_prospects[f'H{row}'].fill = strong_sell_fill
                else:
                    ws_prospects[f'H{row}'] = "Insufficient Data"
                    ws_prospects[f'H{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'H{row}'].border = border
                    ws_prospects[f'H{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    ws_prospects[f'H{row}'].fill = no_data_fill
                
                # Current Price (Column I)
                ws_prospects[f'I{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                ws_prospects[f'I{row}'].font = data_font
                ws_prospects[f'I{row}'].border = border
                ws_prospects[f'I{row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                row += 1
            
            # Add stocks with insufficient data at the bottom
            if stocks_without_data:
                row += 2
                ws_prospects[f'A{row}'] = "STOCKS WITH INSUFFICIENT DATA"
                ws_prospects[f'A{row}'].font = Font(bold=True, size=12, color="808080")
                ws_prospects[f'A{row}'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                ws_prospects[f'A{row}'].border = border
                ws_prospects[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
                row += 1
                
                for stock_row in stocks_without_data:
                    ticker = stock_row['ticker']
                    company_name = stock_row['company_name']
                    current_price = stock_row['current_price']
                    
                    # Company name and ticker
                    ws_prospects[f'A{row}'] = f"{company_name} ({ticker})"
                    ws_prospects[f'A{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'A{row}'].border = border
                    ws_prospects[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Current price
                    ws_prospects[f'E{row}'] = f"${current_price:,.2f}" if current_price else "N/A"
                    ws_prospects[f'E{row}'].font = Font(size=11, color="808080")  # Grey text
                    ws_prospects[f'E{row}'].border = border
                    ws_prospects[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # All valuation columns show "Insufficient Data"
                    for col in ['B', 'C', 'D']:
                        ws_prospects[f'{col}{row}'] = "Insufficient Data"
                        ws_prospects[f'{col}{row}'].font = Font(size=11, color="808080")  # Grey text
                        ws_prospects[f'{col}{row}'].border = border
                        ws_prospects[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
                        ws_prospects[f'{col}{row}'].fill = no_data_fill
                    
                    row += 1
            
            # Add legend
            row += 1
            ws_prospects[f'A{row}'] = "Legend:"
            ws_prospects[f'A{row}'].font = Font(bold=True, size=12)
            ws_prospects[f'A{row}'].border = border
            row += 1
            
            legend_items = [
                ("Strong Buy", strong_buy_fill),
                ("Buy", buy_fill),
                ("Hold", hold_fill),
                ("Sell", sell_fill),
                ("Strong Sell", strong_sell_fill),
                ("Insufficient Data", no_data_fill)
            ]
            
            for i, (label, fill) in enumerate(legend_items):
                col = i + 1
                ws_prospects.cell(row=row, column=col, value=label)
                ws_prospects.cell(row=row, column=col).font = Font(size=10)
                ws_prospects.cell(row=row, column=col).fill = fill
                ws_prospects.cell(row=row, column=col).border = border
                ws_prospects.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
            
            # Set optimal column widths for readability
            ws_prospects.column_dimensions['A'].width = 45  # Company name and ticker
            ws_prospects.column_dimensions['B'].width = 35  # Peter Lynch valuation
            ws_prospects.column_dimensions['C'].width = 35  # DCF valuation
            ws_prospects.column_dimensions['D'].width = 35  # Munger Farm valuation
            ws_prospects.column_dimensions['E'].width = 35  # Enhanced DCF valuation
            ws_prospects.column_dimensions['F'].width = 35  # Relative Valuation
            ws_prospects.column_dimensions['G'].width = 35  # Reverse DCF valuation
            ws_prospects.column_dimensions['H'].width = 35  # EPV/RIM valuation
            ws_prospects.column_dimensions['I'].width = 20  # Current price
            
            self.logger.info("Created prospects sheet with undervaluation ranking")
            
        except Exception as e:
            self.logger.error(f"Error creating prospects sheet: {e}")

    @rate_limit(delay=VALUATION_CONFIG['historical_batch_delay'])
    def collect_comprehensive_historical_data(self, ticker: str) -> Optional[Dict]:
        """
        Collect extensive historical data for trend analysis
        Returns: Dictionary with price history, financial history, valuation metrics over time
        """
        try:
            valid_ticker = self.try_ticker_variations(ticker)
            if not valid_ticker:
                self.logger.warning(f"Could not find valid ticker for {ticker}")
                return None
            
            self.logger.info(f"Collecting historical data for {ticker} ({valid_ticker})")
            stock = yf.Ticker(valid_ticker)
            
            # 1. Price History (daily back to 2000 or max available)
            max_period = VALUATION_CONFIG['max_historical_period']
            price_history = stock.history(period=max_period)
            
            if price_history.empty:
                self.logger.warning(f"No historical price data for {ticker}")
                return None
            
            # 2. Quarterly Financials (all available)
            quarterly_financials = stock.quarterly_financials
            quarterly_balance_sheet = stock.quarterly_balance_sheet
            quarterly_cashflow = stock.quarterly_cashflow
            
            # 3. Annual Financials (all available)
            annual_financials = stock.financials
            annual_balance_sheet = stock.balance_sheet
            annual_cashflow = stock.cashflow
            
            # 4. Calculate historical metrics
            historical_data = {
                'ticker': ticker,
                'valid_ticker': valid_ticker,
                'company_name': self.focus_stocks.get(ticker, ticker),
                'data_start_date': price_history.index[0].strftime('%Y-%m-%d'),
                'data_end_date': price_history.index[-1].strftime('%Y-%m-%d'),
                'total_days': len(price_history),
                'collection_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                
                # Price metrics over time
                'daily_prices': price_history['Close'].tolist(),
                'daily_volumes': price_history['Volume'].tolist() if 'Volume' in price_history.columns else [],
                'daily_highs': price_history['High'].tolist() if 'High' in price_history.columns else [],
                'daily_lows': price_history['Low'].tolist() if 'Low' in price_history.columns else [],
                'daily_opens': price_history['Open'].tolist() if 'Open' in price_history.columns else [],
                'dates': [d.strftime('%Y-%m-%d') for d in price_history.index],
                
                # Calculate rolling metrics
                'rolling_52w_high': price_history['Close'].rolling(252).max().tolist(),
                'rolling_52w_low': price_history['Close'].rolling(252).min().tolist(),
                'rolling_200d_ma': price_history['Close'].rolling(200).mean().tolist(),
                'rolling_50d_ma': price_history['Close'].rolling(50).mean().tolist(),
                'rolling_20d_ma': price_history['Close'].rolling(20).mean().tolist(),
                
                # Volatility metrics
                'rolling_volatility': price_history['Close'].pct_change().rolling(30).std().tolist(),
                'rolling_volatility_annualized': (price_history['Close'].pct_change().rolling(30).std() * np.sqrt(252)).tolist(),
                
                # Annual returns
                'annual_returns': self._calculate_annual_returns(price_history),
                
                # Quarterly financial data
                'quarterly_data': self._extract_quarterly_metrics(
                    quarterly_financials, 
                    quarterly_balance_sheet, 
                    quarterly_cashflow
                ),
                
                # Annual financial data
                'annual_data': self._extract_annual_metrics(
                    annual_financials, 
                    annual_balance_sheet, 
                    annual_cashflow
                ),
                
                # Current metrics for comparison
                'current_price': price_history['Close'].iloc[-1] if not price_history.empty else None,
                'current_volume': price_history['Volume'].iloc[-1] if 'Volume' in price_history.columns and not price_history.empty else None,
                'price_change_1d': price_history['Close'].pct_change().iloc[-1] * 100 if len(price_history) > 1 else None,
                'price_change_1w': ((price_history['Close'].iloc[-1] / price_history['Close'].iloc[-5]) - 1) * 100 if len(price_history) >= 5 else None,
                'price_change_1m': ((price_history['Close'].iloc[-1] / price_history['Close'].iloc[-20]) - 1) * 100 if len(price_history) >= 20 else None,
                'price_change_1y': ((price_history['Close'].iloc[-1] / price_history['Close'].iloc[-252]) - 1) * 100 if len(price_history) >= 252 else None,
            }
            
            # Calculate additional trend metrics
            historical_data.update(self._calculate_trend_metrics(price_history))
            
            # Calculate historical P/E ratios
            historical_data['historical_pe_ratios'] = self._calculate_historical_pe_ratios(
                price_history, quarterly_financials, annual_financials, quarterly_balance_sheet, annual_balance_sheet
            )
            
            self.logger.info(f"Successfully collected {len(price_history)} days of data for {ticker}")
            return historical_data
            
        except Exception as e:
            self.logger.error(f"Error collecting historical data for {ticker}: {e}")
            return None

    def _calculate_annual_returns(self, price_history: pd.DataFrame) -> Dict:
        """Calculate annual returns for each calendar year"""
        annual_returns = {}
        
        for year in price_history.index.year.unique():
            year_data = price_history[price_history.index.year == year]
            if len(year_data) > 1:
                start_price = year_data['Close'].iloc[0]
                end_price = year_data['Close'].iloc[-1]
                annual_return = ((end_price - start_price) / start_price) * 100
                annual_returns[int(year)] = round(annual_return, 2)
        
        return annual_returns

    def _extract_quarterly_metrics(self, financials, balance_sheet, cashflow) -> List[Dict]:
        """Extract key metrics from quarterly data"""
        quarterly_metrics = []
        
        if financials is not None and not financials.empty:
            for date in financials.columns:
                try:
                    metrics = {
                        'date': date.strftime('%Y-%m-%d'),
                        'quarter': f"Q{date.quarter} {date.year}",
                        'revenue': financials.loc['Total Revenue', date] if 'Total Revenue' in financials.index else None,
                        'net_income': financials.loc['Net Income', date] if 'Net Income' in financials.index else None,
                        'gross_profit': financials.loc['Gross Profit', date] if 'Gross Profit' in financials.index else None,
                        'operating_income': financials.loc['Operating Income', date] if 'Operating Income' in financials.index else None,
                        'ebitda': financials.loc['EBITDA', date] if 'EBITDA' in financials.index else None,
                    }
                    
                    # Add balance sheet items
                    if balance_sheet is not None and date in balance_sheet.columns:
                        metrics['total_assets'] = balance_sheet.loc['Total Assets', date] if 'Total Assets' in balance_sheet.index else None
                        metrics['total_liabilities'] = balance_sheet.loc['Total Liabilities Net Minority Interest', date] if 'Total Liabilities Net Minority Interest' in balance_sheet.index else None
                        metrics['shareholders_equity'] = balance_sheet.loc['Stockholders Equity', date] if 'Stockholders Equity' in balance_sheet.index else None
                        metrics['cash_and_equivalents'] = balance_sheet.loc['Cash And Cash Equivalents', date] if 'Cash And Cash Equivalents' in balance_sheet.index else None
                        metrics['total_debt'] = balance_sheet.loc['Total Debt', date] if 'Total Debt' in balance_sheet.index else None
                    
                    # Add cashflow items
                    if cashflow is not None and date in cashflow.columns:
                        metrics['operating_cashflow'] = cashflow.loc['Operating Cash Flow', date] if 'Operating Cash Flow' in cashflow.index else None
                        metrics['free_cashflow'] = cashflow.loc['Free Cash Flow', date] if 'Free Cash Flow' in cashflow.index else None
                        metrics['capital_expenditures'] = cashflow.loc['Capital Expenditures', date] if 'Capital Expenditures' in cashflow.index else None
                    
                    # Calculate ratios
                    if metrics['revenue'] and metrics['net_income']:
                        metrics['net_margin'] = (metrics['net_income'] / metrics['revenue']) * 100
                    if metrics['total_assets'] and metrics['net_income']:
                        metrics['roa'] = (metrics['net_income'] / metrics['total_assets']) * 100
                    if metrics['shareholders_equity'] and metrics['net_income']:
                        metrics['roe'] = (metrics['net_income'] / metrics['shareholders_equity']) * 100
                    
                    quarterly_metrics.append(metrics)
                    
                except Exception as e:
                    self.logger.debug(f"Error extracting quarterly metrics for {date}: {e}")
                    continue
        
        return quarterly_metrics

    def _extract_annual_metrics(self, financials, balance_sheet, cashflow) -> List[Dict]:
        """Extract key metrics from annual data"""
        annual_metrics = []
        
        if financials is not None and not financials.empty:
            for date in financials.columns:
                try:
                    metrics = {
                        'date': date.strftime('%Y-%m-%d'),
                        'year': date.year,
                        'revenue': financials.loc['Total Revenue', date] if 'Total Revenue' in financials.index else None,
                        'net_income': financials.loc['Net Income', date] if 'Net Income' in financials.index else None,
                        'ebitda': financials.loc['EBITDA', date] if 'EBITDA' in financials.index else None,
                        'gross_profit': financials.loc['Gross Profit', date] if 'Gross Profit' in financials.index else None,
                        'operating_income': financials.loc['Operating Income', date] if 'Operating Income' in financials.index else None,
                    }
                    
                    # Add balance sheet items
                    if balance_sheet is not None and date in balance_sheet.columns:
                        metrics['total_assets'] = balance_sheet.loc['Total Assets', date] if 'Total Assets' in balance_sheet.index else None
                        metrics['total_liabilities'] = balance_sheet.loc['Total Liabilities Net Minority Interest', date] if 'Total Liabilities Net Minority Interest' in balance_sheet.index else None
                        metrics['shareholders_equity'] = balance_sheet.loc['Stockholders Equity', date] if 'Stockholders Equity' in balance_sheet.index else None
                        metrics['cash_and_equivalents'] = balance_sheet.loc['Cash And Cash Equivalents', date] if 'Cash And Cash Equivalents' in balance_sheet.index else None
                        metrics['total_debt'] = balance_sheet.loc['Total Debt', date] if 'Total Debt' in balance_sheet.index else None
                    
                    # Add cashflow items
                    if cashflow is not None and date in cashflow.columns:
                        metrics['operating_cashflow'] = cashflow.loc['Operating Cash Flow', date] if 'Operating Cash Flow' in cashflow.index else None
                        metrics['free_cashflow'] = cashflow.loc['Free Cash Flow', date] if 'Free Cash Flow' in cashflow.index else None
                        metrics['capital_expenditures'] = cashflow.loc['Capital Expenditures', date] if 'Capital Expenditures' in cashflow.index else None
                    
                    # Calculate growth rates
                    if len(annual_metrics) > 0:
                        prev_revenue = annual_metrics[-1].get('revenue')
                        prev_net_income = annual_metrics[-1].get('net_income')
                        prev_fcf = annual_metrics[-1].get('free_cashflow')
                        
                        if prev_revenue and metrics['revenue']:
                            metrics['revenue_growth'] = ((metrics['revenue'] - prev_revenue) / prev_revenue) * 100
                        if prev_net_income and metrics['net_income']:
                            metrics['net_income_growth'] = ((metrics['net_income'] - prev_net_income) / prev_net_income) * 100
                        if prev_fcf and metrics['free_cashflow']:
                            metrics['fcf_growth'] = ((metrics['free_cashflow'] - prev_fcf) / prev_fcf) * 100
                    
                    # Calculate ratios
                    if metrics['revenue'] and metrics['net_income']:
                        metrics['net_margin'] = (metrics['net_income'] / metrics['revenue']) * 100
                    if metrics['total_assets'] and metrics['net_income']:
                        metrics['roa'] = (metrics['net_income'] / metrics['total_assets']) * 100
                    if metrics['shareholders_equity'] and metrics['net_income']:
                        metrics['roe'] = (metrics['net_income'] / metrics['shareholders_equity']) * 100
                    if metrics['total_debt'] and metrics['shareholders_equity']:
                        metrics['debt_to_equity'] = metrics['total_debt'] / metrics['shareholders_equity']
                    
                    annual_metrics.append(metrics)
                    
                except Exception as e:
                    self.logger.debug(f"Error extracting annual metrics for {date}: {e}")
                    continue
        
        return annual_metrics

    def _calculate_trend_metrics(self, price_history: pd.DataFrame) -> Dict:
        """Calculate additional trend analysis metrics"""
        try:
            prices = price_history['Close']
            
            # Calculate percentiles
            current_price = prices.iloc[-1]
            price_percentile_1y = (prices.tail(252) < current_price).mean() * 100 if len(prices) >= 252 else None
            price_percentile_5y = (prices.tail(1260) < current_price).mean() * 100 if len(prices) >= 1260 else None
            price_percentile_all = (prices < current_price).mean() * 100
            
            # Calculate Z-scores
            price_zscore_1y = ((current_price - prices.tail(252).mean()) / prices.tail(252).std()) if len(prices) >= 252 else None
            price_zscore_5y = ((current_price - prices.tail(1260).mean()) / prices.tail(1260).std()) if len(prices) >= 1260 else None
            
            # Calculate CAGR
            years_available = len(prices) / 252
            if years_available >= 1:
                cagr_1y = ((current_price / prices.iloc[-252]) ** (1/1) - 1) * 100 if len(prices) >= 252 else None
                cagr_5y = ((current_price / prices.iloc[-1260]) ** (1/5) - 1) * 100 if len(prices) >= 1260 else None
                cagr_all = ((current_price / prices.iloc[0]) ** (1/years_available) - 1) * 100
            else:
                cagr_1y = cagr_5y = cagr_all = None
            
            # Calculate maximum drawdown
            rolling_max = prices.expanding().max()
            drawdown = (prices - rolling_max) / rolling_max
            max_drawdown = drawdown.min() * 100
            
            # Calculate Sharpe ratio (simplified)
            returns = prices.pct_change().dropna()
            sharpe_ratio = (returns.mean() / returns.std()) * np.sqrt(252) if len(returns) > 1 and returns.std() > 0 else None
            
            return {
                'price_percentile_1y': price_percentile_1y,
                'price_percentile_5y': price_percentile_5y,
                'price_percentile_all': price_percentile_all,
                'price_zscore_1y': price_zscore_1y,
                'price_zscore_5y': price_zscore_5y,
                'cagr_1y': cagr_1y,
                'cagr_5y': cagr_5y,
                'cagr_all': cagr_all,
                'max_drawdown': max_drawdown,
                'sharpe_ratio': sharpe_ratio,
                'years_of_data': years_available
            }
            
        except Exception as e:
            self.logger.debug(f"Error calculating trend metrics: {e}")
            return {}

    def _calculate_historical_pe_ratios(self, price_history: pd.DataFrame, quarterly_financials, annual_financials, quarterly_balance_sheet=None, annual_balance_sheet=None) -> List[Dict]:
        """Calculate historical P/E ratios by combining price data with earnings data"""
        try:
            pe_data = []
            
            # Get annual earnings and shares outstanding data
            annual_data = {}
            if annual_financials is not None and not annual_financials.empty:
                for date in annual_financials.columns:
                    try:
                        net_income = annual_financials.loc['Net Income', date] if 'Net Income' in annual_financials.index else None
                        
                        # Get shares outstanding from balance sheet
                        shares_outstanding = None
                        if annual_balance_sheet is not None and date in annual_balance_sheet.columns:
                            # Try different possible column names for shares outstanding
                            possible_shares_columns = [
                                'Ordinary Shares Number',
                                'Common Stock Shares Outstanding',
                                'Shares Outstanding',
                                'Number of Shares',
                                'Common Shares Outstanding'
                            ]
                            
                            for col in possible_shares_columns:
                                if col in annual_balance_sheet.index:
                                    shares_outstanding = annual_balance_sheet.loc[col, date]
                                    break
                        
                        if net_income and net_income > 0 and shares_outstanding and shares_outstanding > 0:
                            annual_data[date.year] = {
                                'net_income': net_income,
                                'shares_outstanding': shares_outstanding,
                                'eps': net_income / shares_outstanding,
                                'date': date
                            }
                    except Exception as e:
                        self.logger.debug(f"Error extracting annual data for {date}: {e}")
                        continue
            
            # Get quarterly earnings and shares outstanding data
            quarterly_data = {}
            if quarterly_financials is not None and not quarterly_financials.empty:
                for date in quarterly_financials.columns:
                    try:
                        net_income = quarterly_financials.loc['Net Income', date] if 'Net Income' in quarterly_financials.index else None
                        
                        # Get shares outstanding from quarterly balance sheet
                        shares_outstanding = None
                        if quarterly_balance_sheet is not None and date in quarterly_balance_sheet.columns:
                            # Try different possible column names for shares outstanding
                            possible_shares_columns = [
                                'Ordinary Shares Number',
                                'Common Stock Shares Outstanding',
                                'Shares Outstanding',
                                'Number of Shares',
                                'Common Shares Outstanding'
                            ]
                            
                            for col in possible_shares_columns:
                                if col in quarterly_balance_sheet.index:
                                    shares_outstanding = quarterly_balance_sheet.loc[col, date]
                                    break
                        
                        if net_income and net_income > 0 and shares_outstanding and shares_outstanding > 0:
                            # Annualize quarterly earnings
                            annualized_net_income = net_income * 4
                            quarterly_data[date] = {
                                'net_income': annualized_net_income,
                                'shares_outstanding': shares_outstanding,
                                'eps': annualized_net_income / shares_outstanding,
                                'date': date
                            }
                    except Exception as e:
                        self.logger.debug(f"Error extracting quarterly data for {date}: {e}")
                        continue
            
            # Calculate P/E ratios for each trading day
            for date, row in price_history.iterrows():
                current_price = row['Close']
                
                # Try to find the most recent earnings data
                earnings_data = None
                
                # First, try quarterly data (more recent)
                for q_date in sorted(quarterly_data.keys(), reverse=True):
                    if q_date <= date:
                        earnings_data = quarterly_data[q_date]
                        break
                
                # If no quarterly data, try annual data
                if earnings_data is None:
                    for year in sorted(annual_data.keys(), reverse=True):
                        if year <= date.year:
                            earnings_data = annual_data[year]
                            break
                
                # Calculate P/E ratio if we have earnings data
                if earnings_data and earnings_data['eps'] > 0:
                    pe_ratio = current_price / earnings_data['eps']
                    
                    pe_data.append({
                        'date': date.strftime('%Y-%m-%d'),
                        'price': current_price,
                        'earnings_per_share': earnings_data['eps'],
                        'pe_ratio': pe_ratio,
                        'earnings_date': earnings_data['date'].strftime('%Y-%m-%d'),
                        'earnings_type': 'quarterly' if earnings_data['date'] in quarterly_data else 'annual',
                        'shares_outstanding': earnings_data['shares_outstanding'],
                        'net_income': earnings_data['net_income']
                    })
            
            # Calculate P/E statistics
            if pe_data:
                pe_ratios = [item['pe_ratio'] for item in pe_data]
                
                # Add summary statistics
                pe_summary = {
                    'current_pe': pe_data[-1]['pe_ratio'] if pe_data else None,
                    'pe_mean': np.mean(pe_ratios),
                    'pe_median': np.median(pe_ratios),
                    'pe_std': np.std(pe_ratios),
                    'pe_min': np.min(pe_ratios),
                    'pe_max': np.max(pe_ratios),
                    'pe_percentile_25': np.percentile(pe_ratios, 25),
                    'pe_percentile_75': np.percentile(pe_ratios, 75),
                    'pe_percentile_90': np.percentile(pe_ratios, 90),
                    'pe_percentile_95': np.percentile(pe_ratios, 95),
                    'total_pe_data_points': len(pe_data)
                }
                
                # Calculate current P/E percentile
                if pe_summary['current_pe']:
                    pe_summary['current_pe_percentile'] = (np.array(pe_ratios) < pe_summary['current_pe']).mean() * 100
                
                # Add valuation assessment
                if pe_summary['current_pe']:
                    if pe_summary['current_pe'] < pe_summary['pe_percentile_25']:
                        pe_summary['valuation_assessment'] = 'Very Undervalued'
                    elif pe_summary['current_pe'] < pe_summary['pe_percentile_50']:
                        pe_summary['valuation_assessment'] = 'Undervalued'
                    elif pe_summary['current_pe'] < pe_summary['pe_percentile_75']:
                        pe_summary['valuation_assessment'] = 'Fairly Valued'
                    elif pe_summary['current_pe'] < pe_summary['pe_percentile_90']:
                        pe_summary['valuation_assessment'] = 'Overvalued'
                    else:
                        pe_summary['valuation_assessment'] = 'Very Overvalued'
                
                # Add summary to the data
                pe_data.append({
                    'date': 'SUMMARY',
                    'price': None,
                    'earnings_per_share': None,
                    'pe_ratio': None,
                    'earnings_date': None,
                    'earnings_type': 'SUMMARY',
                    'shares_outstanding': None,
                    'net_income': None,
                    'summary_stats': pe_summary
                })
            
            self.logger.info(f"Calculated {len(pe_data)} P/E ratio data points")
            return pe_data
            
        except Exception as e:
            self.logger.error(f"Error calculating historical P/E ratios: {e}")
            return []

    def _load_historical_progress(self) -> Dict:
        """Load historical data collection progress"""
        try:
            if os.path.exists(self.historical_progress_file):
                with open(self.historical_progress_file, 'r') as f:
                    return json.load(f)
            return {'completed': [], 'failed': [], 'last_updated': None}
        except Exception as e:
            self.logger.error(f"Error loading historical progress: {e}")
            return {'completed': [], 'failed': [], 'last_updated': None}

    def _save_historical_progress(self, progress: Dict):
        """Save historical data collection progress"""
        try:
            progress['last_updated'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            with open(self.historical_progress_file, 'w') as f:
                json.dump(progress, f, indent=2)
        except Exception as e:
            self.logger.error(f"Error saving historical progress: {e}")

    def _append_to_historical_file(self, historical_data: Dict):
        """Append historical data to the historical Excel file with maximum daily granularity"""
        try:
            # Create comprehensive daily trends data
            trends_data = []
            
            # Create a comprehensive daily dataset
            for i, (date, price, volume, high, low, open_price) in enumerate(zip(
                historical_data['dates'],
                historical_data['daily_prices'],
                historical_data['daily_volumes'],
                historical_data['daily_highs'],
                historical_data['daily_lows'],
                historical_data['daily_opens']
            )):
                # Base daily data
                daily_row = {
                    'Date': date,
                    'Ticker': historical_data['ticker'],
                    'Company_Name': historical_data['company_name'],
                    'Price': price,
                    'Volume': volume,
                    'High': high,
                    'Low': low,
                    'Open': open_price,
                    'Data_Type': 'Daily_Comprehensive'
                }
                
                # Add rolling metrics for this day
                if i < len(historical_data['rolling_52w_high']):
                    daily_row['52W_High'] = historical_data['rolling_52w_high'][i]
                    daily_row['52W_Low'] = historical_data['rolling_52w_low'][i]
                if i < len(historical_data['rolling_200d_ma']):
                    daily_row['200D_MA'] = historical_data['rolling_200d_ma'][i]
                if i < len(historical_data['rolling_50d_ma']):
                    daily_row['50D_MA'] = historical_data['rolling_50d_ma'][i]
                if i < len(historical_data['rolling_20d_ma']):
                    daily_row['20D_MA'] = historical_data['rolling_20d_ma'][i]
                if i < len(historical_data['rolling_volatility']):
                    daily_row['Volatility_30D'] = historical_data['rolling_volatility'][i]
                if i < len(historical_data['rolling_volatility_annualized']):
                    daily_row['Volatility_Annualized'] = historical_data['rolling_volatility_annualized'][i]
                
                # Add price change metrics for the most recent day
                if i == len(historical_data['dates']) - 1:
                    daily_row['Price_Change_1D'] = historical_data['price_change_1d']
                    daily_row['Price_Change_1W'] = historical_data['price_change_1w']
                    daily_row['Price_Change_1M'] = historical_data['price_change_1m']
                    daily_row['Price_Change_1Y'] = historical_data['price_change_1y']
                
                # Add P/E ratio data for this day if available
                pe_ratio_for_date = None
                eps_for_date = None
                earnings_date_for_date = None
                earnings_type_for_date = None
                shares_outstanding_for_date = None
                net_income_for_date = None
                
                # Find P/E data for this specific date
                for pe_data in historical_data['historical_pe_ratios']:
                    if pe_data['date'] == date:
                        pe_ratio_for_date = pe_data['pe_ratio']
                        eps_for_date = pe_data['earnings_per_share']
                        earnings_date_for_date = pe_data['earnings_date']
                        earnings_type_for_date = pe_data['earnings_type']
                        shares_outstanding_for_date = pe_data['shares_outstanding']
                        net_income_for_date = pe_data['net_income']
                        break
                
                daily_row['PE_Ratio'] = pe_ratio_for_date
                daily_row['EPS'] = eps_for_date
                daily_row['Earnings_Date'] = earnings_date_for_date
                daily_row['Earnings_Type'] = earnings_type_for_date
                daily_row['Shares_Outstanding'] = shares_outstanding_for_date
                daily_row['Net_Income'] = net_income_for_date
                
                # Add trend metrics
                daily_row['Price_Percentile_1Y'] = historical_data.get('price_percentile_1y')
                daily_row['Price_Percentile_5Y'] = historical_data.get('price_percentile_5y')
                daily_row['Price_Percentile_All'] = historical_data.get('price_percentile_all')
                daily_row['Price_ZScore_1Y'] = historical_data.get('price_zscore_1y')
                daily_row['Price_ZScore_5Y'] = historical_data.get('price_zscore_5y')
                daily_row['CAGR_1Y'] = historical_data.get('cagr_1y')
                daily_row['CAGR_5Y'] = historical_data.get('cagr_5y')
                daily_row['CAGR_All'] = historical_data.get('cagr_all')
                daily_row['Max_Drawdown'] = historical_data.get('max_drawdown')
                daily_row['Sharpe_Ratio'] = historical_data.get('sharpe_ratio')
                daily_row['Years_of_Data'] = historical_data.get('years_of_data')
                
                trends_data.append(daily_row)
            
            # Add quarterly financial data as separate entries
            for q_data in historical_data['quarterly_data']:
                trends_data.append({
                    'Date': q_data['date'],
                    'Ticker': historical_data['ticker'],
                    'Company_Name': historical_data['company_name'],
                    'Revenue': q_data['revenue'],
                    'Net_Income': q_data['net_income'],
                    'Gross_Profit': q_data['gross_profit'],
                    'Operating_Income': q_data['operating_income'],
                    'EBITDA': q_data['ebitda'],
                    'Total_Assets': q_data['total_assets'],
                    'Total_Liabilities': q_data['total_liabilities'],
                    'Shareholders_Equity': q_data['shareholders_equity'],
                    'Operating_Cashflow': q_data['operating_cashflow'],
                    'Free_Cashflow': q_data['free_cashflow'],
                    'Net_Margin': q_data.get('net_margin'),
                    'ROA': q_data.get('roa'),
                    'ROE': q_data.get('roe'),
                    'Data_Type': 'Quarterly_Financial'
                })
            
            # Add annual financial data as separate entries
            for a_data in historical_data['annual_data']:
                trends_data.append({
                    'Date': a_data['date'],
                    'Ticker': historical_data['ticker'],
                    'Company_Name': historical_data['company_name'],
                    'Revenue': a_data['revenue'],
                    'Net_Income': a_data['net_income'],
                    'EBITDA': a_data['ebitda'],
                    'Total_Assets': a_data['total_assets'],
                    'Total_Liabilities': a_data['total_liabilities'],
                    'Shareholders_Equity': a_data['shareholders_equity'],
                    'Operating_Cashflow': a_data['operating_cashflow'],
                    'Free_Cashflow': a_data['free_cashflow'],
                    'Revenue_Growth': a_data.get('revenue_growth'),
                    'Net_Income_Growth': a_data.get('net_income_growth'),
                    'FCF_Growth': a_data.get('fcf_growth'),
                    'Net_Margin': a_data.get('net_margin'),
                    'ROA': a_data.get('roa'),
                    'ROE': a_data.get('roe'),
                    'Debt_to_Equity': a_data.get('debt_to_equity'),
                    'Data_Type': 'Annual_Financial'
                })
            
            # Add P/E summary statistics as a separate entry
            pe_summary_data = None
            for pe_data in historical_data['historical_pe_ratios']:
                if pe_data['date'] == 'SUMMARY':
                    pe_summary_data = pe_data['summary_stats']
                    break
            
            if pe_summary_data:
                trends_data.append({
                    'Date': historical_data['data_end_date'],
                    'Ticker': historical_data['ticker'],
                    'Company_Name': historical_data['company_name'],
                    'Current_PE': pe_summary_data['current_pe'],
                    'PE_Mean': pe_summary_data['pe_mean'],
                    'PE_Median': pe_summary_data['pe_median'],
                    'PE_Std': pe_summary_data['pe_std'],
                    'PE_Min': pe_summary_data['pe_min'],
                    'PE_Max': pe_summary_data['pe_max'],
                    'PE_Percentile_25': pe_summary_data['pe_percentile_25'],
                    'PE_Percentile_75': pe_summary_data['pe_percentile_75'],
                    'PE_Percentile_90': pe_summary_data['pe_percentile_90'],
                    'PE_Percentile_95': pe_summary_data['pe_percentile_95'],
                    'PE_Percentile_Current': pe_summary_data['current_pe_percentile'],
                    'Valuation_Assessment': pe_summary_data['valuation_assessment'],
                    'Total_PE_Data_Points': pe_summary_data['total_pe_data_points'],
                    'Data_Type': 'PE_Summary'
                })
            
            # Convert to DataFrame and append to Excel
            df_trends = pd.DataFrame(trends_data)
            
            # Load existing file or create new one
            if os.path.exists(self.historical_file):
                with pd.ExcelWriter(self.historical_file, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                    df_trends.to_excel(writer, sheet_name='Historical_Trends', index=False, header=False, startrow=writer.sheets['Historical_Trends'].max_row)
            else:
                # Create new file with proper formatting
                self._create_formatted_historical_file(df_trends)
            
            self.logger.info(f"Appended {len(trends_data)} comprehensive records to historical file for {historical_data['ticker']}")
            self.logger.info(f"Daily records: {len(historical_data['dates'])}, Quarterly: {len(historical_data['quarterly_data'])}, Annual: {len(historical_data['annual_data'])}")
            
        except Exception as e:
            self.logger.error(f"Error appending to historical file: {e}")

    def _create_formatted_historical_file(self, df_trends: pd.DataFrame):
        """Create a new historical file with proper formatting and structure"""
        try:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Create Historical Trends sheet
            ws_trends = wb.create_sheet('Historical_Trends', 0)
            
            # Add headers with formatting
            headers = list(df_trends.columns)
            for col, header in enumerate(headers, 1):
                cell = ws_trends.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # Add data
            for row_idx, (_, row_data) in enumerate(df_trends.iterrows(), 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws_trends.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Format based on data type
                    if col_idx == 1:  # Date column
                        cell.alignment = Alignment(horizontal='center')
                    elif isinstance(value, (int, float)) and value is not None:
                        if 'PE' in headers[col_idx-1] or 'Ratio' in headers[col_idx-1]:
                            cell.number_format = '0.00'
                        elif 'Price' in headers[col_idx-1] or 'EPS' in headers[col_idx-1]:
                            cell.number_format = '$#,##0.00'
                        elif 'Volume' in headers[col_idx-1] or 'Shares' in headers[col_idx-1]:
                            cell.number_format = '#,##0'
                        elif 'Percent' in headers[col_idx-1] or 'Growth' in headers[col_idx-1]:
                            cell.number_format = '0.00%'
                        else:
                            cell.number_format = '#,##0.00'
            
            # Set column widths
            column_widths = {
                'A': 12,  # Date
                'B': 10,  # Ticker
                'C': 35,  # Company Name
                'D': 12,  # Price
                'E': 15,  # Volume
                'F': 12,  # High
                'G': 12,  # Low
                'H': 12,  # Open
                'I': 12,  # PE Ratio
                'J': 12,  # EPS
                'K': 15,  # Data Type
            }
            
            for col, width in column_widths.items():
                ws_trends.column_dimensions[col].width = width
            
            # Create Historical Summary sheet
            ws_summary = wb.create_sheet('Historical_Summary', 1)
            self._create_historical_summary_sheet(ws_summary)
            
            # Create Historical Analysis sheet
            ws_analysis = wb.create_sheet('Historical_Analysis', 2)
            self._create_historical_analysis_sheet(ws_analysis)
            
            # Save the workbook
            wb.save(self.historical_file)
            self.logger.info(f"Created formatted historical file: {self.historical_file}")
            
        except Exception as e:
            self.logger.error(f"Error creating formatted historical file: {e}")

    def _create_historical_summary_sheet(self, ws):
        """Create a summary sheet for historical data overview"""
        try:
            # Title
            ws['A1'] = "Historical Data Collection Summary"
            ws['A1'].font = Font(bold=True, size=16, color="366092")
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:F1')
            
            # Collection info
            ws['A3'] = "Collection Date:"
            ws['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws['A3'].font = Font(bold=True)
            
            ws['A4'] = "Total Stocks Tracked:"
            ws['B4'] = len(self.focus_stocks)
            ws['A4'].font = Font(bold=True)
            
            ws['A5'] = "Data Period:"
            ws['B5'] = f"Back to {VALUATION_CONFIG['historical_start_year']}"
            ws['A5'].font = Font(bold=True)
            
            ws['A6'] = "Max Historical Period:"
            ws['B6'] = VALUATION_CONFIG['max_historical_period']
            ws['A6'].font = Font(bold=True)
            
            # Data structure explanation
            ws['A8'] = "Data Structure:"
            ws['A8'].font = Font(bold=True, size=12)
            
            data_types = [
                ("Daily_Comprehensive", "Daily price data with rolling metrics, P/E ratios, and trend indicators"),
                ("Quarterly_Financial", "Quarterly financial statements (revenue, earnings, balance sheet)"),
                ("Annual_Financial", "Annual financial statements with growth rates"),
                ("PE_Summary", "P/E ratio statistics and valuation assessments")
            ]
            
            row = 9
            for data_type, description in data_types:
                ws[f'A{row}'] = data_type
                ws[f'B{row}'] = description
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Key metrics explanation
            ws['A14'] = "Key Metrics Tracked:"
            ws['A14'].font = Font(bold=True, size=12)
            
            metrics = [
                ("Price Data", "Daily OHLCV, rolling averages, volatility"),
                ("P/E Ratios", "Historical P/E with percentiles and valuation assessment"),
                ("Financial Data", "Revenue, earnings, cash flow, balance sheet items"),
                ("Trend Analysis", "Price percentiles, Z-scores, CAGR, drawdowns"),
                ("Moving Averages", "20-day, 50-day, 200-day moving averages"),
                ("Volatility", "30-day rolling volatility (daily and annualized)")
            ]
            
            row = 15
            for metric, description in metrics:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = description
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Usage instructions
            ws['A22'] = "Usage Instructions:"
            ws['A22'].font = Font(bold=True, size=12)
            
            instructions = [
                "1. Filter by Ticker to analyze specific stocks",
                "2. Filter by Data_Type to focus on specific data categories",
                "3. Sort by Date to see chronological progression",
                "4. Use P/E percentiles to identify valuation opportunities",
                "5. Compare current metrics to historical averages"
            ]
            
            row = 23
            for instruction in instructions:
                ws[f'A{row}'] = instruction
                row += 1
            
            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 50
            
        except Exception as e:
            self.logger.error(f"Error creating historical summary sheet: {e}")

    def _create_historical_analysis_sheet(self, ws):
        """Create an analysis sheet with sample queries and insights"""
        try:
            # Title
            ws['A1'] = "Historical Data Analysis Guide"
            ws['A1'].font = Font(bold=True, size=16, color="366092")
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:F1')
            
            # Sample analysis queries
            ws['A3'] = "Sample Analysis Queries:"
            ws['A3'].font = Font(bold=True, size=12)
            
            queries = [
                ("Find undervalued stocks", "Filter PE_Percentile_Current < 25 AND Valuation_Assessment = 'Very Undervalued'"),
                ("Identify high volatility stocks", "Filter Volatility_Annualized > 30 AND Data_Type = 'Daily_Comprehensive'"),
                ("Find stocks above 200-day MA", "Filter Price > 200D_MA AND Data_Type = 'Daily_Comprehensive'"),
                ("Identify growth stocks", "Filter Revenue_Growth > 10 AND Data_Type = 'Annual_Financial'"),
                ("Find dividend opportunities", "Filter PE_Ratio < 15 AND Price_Percentile_All < 50"),
                ("Identify momentum stocks", "Filter Price_Change_1Y > 20 AND above_200d_ma = TRUE")
            ]
            
            row = 4
            for query_name, query_filter in queries:
                ws[f'A{row}'] = query_name
                ws[f'B{row}'] = query_filter
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(size=10)
                row += 1
            
            # Key insights to look for
            ws['A12'] = "Key Insights to Look For:"
            ws['A12'].font = Font(bold=True, size=12)
            
            insights = [
                ("Mean Reversion", "Stocks with P/E ratios in bottom 25th percentile often outperform"),
                ("Momentum", "Stocks above 200-day MA tend to continue upward trends"),
                ("Volatility Clusters", "High volatility periods often precede major moves"),
                ("Earnings Quality", "Consistent revenue growth with improving margins"),
                ("Valuation Cycles", "P/E ratios tend to revert to historical means over time"),
                ("Seasonal Patterns", "Some stocks show consistent seasonal performance")
            ]
            
            row = 13
            for insight_name, insight_description in insights:
                ws[f'A{row}'] = insight_name
                ws[f'B{row}'] = insight_description
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Performance benchmarks
            ws['A21'] = "Performance Benchmarks:"
            ws['A21'].font = Font(bold=True, size=12)
            
            benchmarks = [
                ("Excellent CAGR", "> 15% annually over 5+ years"),
                ("Good CAGR", "10-15% annually over 5+ years"),
                ("Acceptable CAGR", "5-10% annually over 5+ years"),
                ("Low Volatility", "< 20% annualized volatility"),
                ("Moderate Volatility", "20-30% annualized volatility"),
                ("High Volatility", "> 30% annualized volatility"),
                ("Conservative P/E", "< 15"),
                ("Moderate P/E", "15-25"),
                ("Growth P/E", "25-40"),
                ("Speculative P/E", "> 40")
            ]
            
            row = 22
            for benchmark_name, benchmark_value in benchmarks:
                ws[f'A{row}'] = benchmark_name
                ws[f'B{row}'] = benchmark_value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            # Set column widths
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 50
            
        except Exception as e:
            self.logger.error(f"Error creating historical analysis sheet: {e}")

    def run_historical_collection(self):
        """Collect historical data progressively with checkpointing"""
        self.logger.info("Starting comprehensive historical data collection")
        
        # Load progress tracker
        progress = self._load_historical_progress()
        
        # Get stocks that haven't been completed
        remaining_stocks = {k: v for k, v in self.focus_stocks.items() if k not in progress.get('completed', [])}
        
        if not remaining_stocks:
            self.logger.info("All stocks have been processed for historical data collection")
            return
        
        self.logger.info(f"Processing {len(remaining_stocks)} remaining stocks for historical data collection")
        
        # Process stocks in batches
        batch_size = VALUATION_CONFIG['historical_batch_size']
        checkpoint_interval = VALUATION_CONFIG['historical_checkpoint_interval']
        
        stock_items = list(remaining_stocks.items())
        
        for i in tqdm(range(0, len(stock_items), batch_size), desc="Processing historical data batches"):
            batch = stock_items[i:i + batch_size]
            
            for ticker, company_name in batch:
                try:
                    self.logger.info(f"Collecting historical data for {ticker} ({company_name})")
                    
                    historical_data = self.collect_comprehensive_historical_data(ticker)
                    
                    if historical_data:
                        # Save to historical file
                        self._append_to_historical_file(historical_data)
                        
                        # Update progress
                        progress['completed'].append(ticker)
                        self.logger.info(f"Completed {ticker} - {len(progress['completed'])}/{len(self.focus_stocks)} total")
                    else:
                        progress['failed'].append(ticker)
                        self.logger.warning(f"Failed to collect data for {ticker}")
                    
                    # Save progress every checkpoint_interval stocks
                    if len(progress['completed']) % checkpoint_interval == 0:
                        self._save_historical_progress(progress)
                        self.logger.info(f"Progress checkpoint saved: {len(progress['completed'])} completed")
                
                except Exception as e:
                    self.logger.error(f"Error processing {ticker}: {e}")
                    progress['failed'].append(ticker)
                    continue
            
            # Rate limiting between batches
            if i + batch_size < len(stock_items):
                self.logger.info(f"Waiting {VALUATION_CONFIG['historical_batch_delay']} seconds before next batch...")
                time.sleep(VALUATION_CONFIG['historical_batch_delay'])
        
        # Final progress save
        self._save_historical_progress(progress)
        
        self.logger.info(f"Historical data collection completed!")
        self.logger.info(f"Successfully processed: {len(progress['completed'])} stocks")
        self.logger.info(f"Failed: {len(progress['failed'])} stocks")
        
        if progress['failed']:
            self.logger.warning(f"Failed stocks: {progress['failed']}")

    def analyze_historical_trends(self, ticker: str = None) -> Dict:
        """Analyze historical trends for a specific ticker or all tickers"""
        try:
            if not os.path.exists(self.historical_file):
                self.logger.error("Historical file not found. Run historical collection first.")
                return {}
            
            # Load historical data
            df_historical = pd.read_excel(self.historical_file, sheet_name='Historical_Trends')
            
            if ticker:
                df_ticker = df_historical[df_historical['Ticker'] == ticker]
                if df_ticker.empty:
                    self.logger.warning(f"No historical data found for {ticker}")
                    return {}
                
                return self._analyze_single_ticker_trends(df_ticker, ticker)
            else:
                # Analyze all tickers
                analysis_results = {}
                for ticker in df_historical['Ticker'].unique():
                    df_ticker = df_historical[df_historical['Ticker'] == ticker]
                    analysis_results[ticker] = self._analyze_single_ticker_trends(df_ticker, ticker)
                
                return analysis_results
                
        except Exception as e:
            self.logger.error(f"Error analyzing historical trends: {e}")
            return {}

    def _analyze_single_ticker_trends(self, df_ticker: pd.DataFrame, ticker: str) -> Dict:
        """Analyze trends for a single ticker"""
        try:
            # Get price data
            price_data = df_ticker[df_ticker['Data_Type'] == 'Daily_Price'].copy()
            price_data['Date'] = pd.to_datetime(price_data['Date'])
            price_data = price_data.sort_values('Date')
            
            if price_data.empty:
                return {'error': 'No price data available'}
            
            current_price = price_data['Price'].iloc[-1]
            
            # Calculate trend metrics
            analysis = {
                'ticker': ticker,
                'current_price': current_price,
                'data_start_date': price_data['Date'].iloc[0].strftime('%Y-%m-%d'),
                'data_end_date': price_data['Date'].iloc[-1].strftime('%Y-%m-%d'),
                'total_trading_days': len(price_data),
                
                # Price analysis
                'price_percentile_1y': self._calculate_percentile(price_data['Price'].tail(252), current_price),
                'price_percentile_5y': self._calculate_percentile(price_data['Price'].tail(1260), current_price),
                'price_percentile_all': self._calculate_percentile(price_data['Price'], current_price),
                
                # Performance metrics
                'total_return': ((current_price / price_data['Price'].iloc[0]) - 1) * 100,
                'annualized_return': self._calculate_cagr(price_data['Price'].iloc[0], current_price, len(price_data) / 252),
                'max_drawdown': self._calculate_max_drawdown(price_data['Price']),
                'volatility': price_data['Price'].pct_change().std() * np.sqrt(252) * 100,
                
                # Trend indicators
                'above_200d_ma': current_price > price_data['Price'].rolling(200).mean().iloc[-1],
                'above_50d_ma': current_price > price_data['Price'].rolling(50).mean().iloc[-1],
                'ma_trend': self._calculate_ma_trend(price_data['Price']),
            }
            
            # Add P/E analysis if available
            pe_data = df_ticker[df_ticker['Data_Type'] == 'PE_Ratio'].copy()
            if not pe_data.empty:
                pe_data['Date'] = pd.to_datetime(pe_data['Date'])
                pe_data = pe_data.sort_values('Date')
                
                current_pe = pe_data['PE_Ratio'].iloc[-1]
                analysis.update({
                    'current_pe_ratio': current_pe,
                    'pe_percentile_1y': self._calculate_percentile(pe_data['PE_Ratio'].tail(252), current_pe),
                    'pe_percentile_5y': self._calculate_percentile(pe_data['PE_Ratio'].tail(1260), current_pe),
                    'pe_percentile_all': self._calculate_percentile(pe_data['PE_Ratio'], current_pe),
                    'pe_mean': pe_data['PE_Ratio'].mean(),
                    'pe_median': pe_data['PE_Ratio'].median(),
                    'pe_std': pe_data['PE_Ratio'].std(),
                    'pe_min': pe_data['PE_Ratio'].min(),
                    'pe_max': pe_data['PE_Ratio'].max(),
                    'pe_trend': self._calculate_pe_trend(pe_data['PE_Ratio'])
                })
            
            return analysis
            
        except Exception as e:
            self.logger.error(f"Error analyzing trends for {ticker}: {e}")
            return {'error': str(e)}

    def _calculate_percentile(self, prices: pd.Series, current_price: float) -> float:
        """Calculate percentile of current price in historical range"""
        if len(prices) == 0:
            return None
        return (prices < current_price).mean() * 100

    def _calculate_cagr(self, start_price: float, end_price: float, years: float) -> float:
        """Calculate Compound Annual Growth Rate"""
        if years <= 0 or start_price <= 0:
            return None
        return ((end_price / start_price) ** (1/years) - 1) * 100

    def _calculate_max_drawdown(self, prices: pd.Series) -> float:
        """Calculate maximum drawdown"""
        rolling_max = prices.expanding().max()
        drawdown = (prices - rolling_max) / rolling_max
        return drawdown.min() * 100

    def _calculate_ma_trend(self, prices: pd.Series) -> str:
        """Calculate moving average trend"""
        try:
            ma_50 = prices.rolling(50).mean()
            ma_200 = prices.rolling(200).mean()
            
            if len(ma_50) < 2 or len(ma_200) < 2:
                return "Insufficient Data"
            
            ma_50_trend = "Up" if ma_50.iloc[-1] > ma_50.iloc[-2] else "Down"
            ma_200_trend = "Up" if ma_200.iloc[-1] > ma_200.iloc[-2] else "Down"
            
            if ma_50_trend == "Up" and ma_200_trend == "Up":
                return "Strong Uptrend"
            elif ma_50_trend == "Up" and ma_200_trend == "Down":
                return "Mixed Trend"
            elif ma_50_trend == "Down" and ma_200_trend == "Up":
                return "Mixed Trend"
            else:
                return "Downtrend"
                
        except Exception as e:
            return "Error"

    def _calculate_pe_trend(self, pe_ratios: pd.Series) -> str:
        """Calculate P/E ratio trend"""
        try:
            if len(pe_ratios) < 2:
                return "Insufficient Data"
            
            # Calculate recent trend (last 20 data points)
            recent_data = pe_ratios.tail(20)
            if len(recent_data) < 2:
                return "Insufficient Data"
            
            # Simple trend calculation
            first_half = recent_data.iloc[:len(recent_data)//2].mean()
            second_half = recent_data.iloc[len(recent_data)//2:].mean()
            
            if second_half > first_half * 1.05:  # 5% increase
                return "Increasing"
            elif second_half < first_half * 0.95:  # 5% decrease
                return "Decreasing"
            else:
                return "Stable"
                
        except Exception as e:
            return "Error"

def main():
    """Main function"""
    scraper = StockValuationScraper()
    
    # Check command line arguments for different modes
    import sys
    
    if len(sys.argv) > 1:
        mode = sys.argv[1].lower()
        
        if mode == 'historical':
            print("Starting historical data collection...")
            scraper.run_historical_collection()
        elif mode == 'analyze':
            if len(sys.argv) > 2:
                ticker = sys.argv[2]
                print(f"Analyzing historical trends for {ticker}...")
                analysis = scraper.analyze_historical_trends(ticker)
                print(f"Analysis for {ticker}:")
                for key, value in analysis.items():
                    print(f"  {key}: {value}")
            else:
                print("Analyzing historical trends for all stocks...")
                analysis = scraper.analyze_historical_trends()
                print(f"Analysis completed for {len(analysis)} stocks")
        elif mode == 'both':
            print("Running both historical collection and regular valuation...")
            scraper.run_historical_collection()
            scraper.run()
        else:
            print("Unknown mode. Available modes: historical, analyze, both")
            print("Running regular valuation...")
            scraper.run()
    else:
        print("Running regular valuation...")
        print("Available modes:")
        print("  python stock_valuation_scraper.py historical  - Collect historical data")
        print("  python stock_valuation_scraper.py analyze [ticker]  - Analyze trends")
        print("  python stock_valuation_scraper.py both  - Run both historical and regular")
        scraper.run()

if __name__ == "__main__":
    main()
