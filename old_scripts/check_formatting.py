import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def check_excel_formatting():
    """Check Excel file formatting and American stock performance"""
    excel_file = r"C:\Users\james\Downloads\Stock Valuation\stock_valuation_dataset.xlsx"
    
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        print("=== EXCEL FILE FORMATTING ANALYSIS ===")
        print(f"Available sheets: {wb.sheetnames}")
        
        # Check My Portfolio sheet
        if 'My Portfolio' in wb.sheetnames:
            ws_portfolio = wb['My Portfolio']
            print(f"\n--- MY PORTFOLIO SHEET ---")
            print(f"Max row: {ws_portfolio.max_row}")
            print(f"Max column: {ws_portfolio.max_column}")
            
            # Check headers
            headers = []
            for col in range(1, ws_portfolio.max_column + 1):
                cell_value = ws_portfolio.cell(row=1, column=col).value
                headers.append(cell_value)
            print(f"Headers: {headers}")
            
            # Check formatting on first few rows
            print("\nFormatting check (first 5 rows):")
            for row in range(1, min(6, ws_portfolio.max_row + 1)):
                row_data = []
                for col in range(1, min(6, ws_portfolio.max_column + 1)):
                    cell = ws_portfolio.cell(row=row, column=col)
                    fill_color = cell.fill.start_color.rgb if cell.fill.start_color else "None"
                    font_color = cell.font.color.rgb if cell.font.color else "None"
                    row_data.append(f"{cell.value} (fill:{fill_color}, font:{font_color})")
                print(f"Row {row}: {row_data}")
        
        # Check Prospects sheet
        if 'Prospects' in wb.sheetnames:
            ws_prospects = wb['Prospects']
            print(f"\n--- PROSPECTS SHEET ---")
            print(f"Max row: {ws_prospects.max_row}")
            print(f"Max column: {ws_prospects.max_column}")
            
            # Check headers
            headers = []
            for col in range(1, ws_prospects.max_column + 1):
                cell_value = ws_prospects.cell(row=1, column=col).value
                headers.append(cell_value)
            print(f"Headers: {headers}")
        
        # Load data for analysis
        df = pd.read_excel(excel_file, sheet_name='Valuation Data')
        latest_data = df.sort_values(by='timestamp').groupby('ticker').last().reset_index()
        
        print(f"\n=== AMERICAN STOCK ANALYSIS ===")
        american_stocks = ['NVDA', 'MSFT', 'AAPL', 'META', 'AMZN', 'GOOGL', 'TSM', 'AMD', 'INTC', 'NVO', 'LMT', 'NOC', 'AMAT', 'SNOW', 'BRK-B', 'BRK-A', 'IWM', 'RKLB']
        
        print(f"Found {len(american_stocks)} American stocks to analyze")
        
        for stock in american_stocks:
            stock_data = latest_data[latest_data['ticker'] == stock]
            if not stock_data.empty:
                row = stock_data.iloc[0]
                print(f"\n--- {stock} ({row.get('company_name', 'Unknown')}) ---")
                print(f"Current Price: ${row.get('current_price', 'N/A')}")
                
                # Peter Lynch
                lynch_status = row.get('lynch_valuation_status', 'N/A')
                lynch_delta = row.get('lynch_delta_percentage', 0)
                print(f"Peter Lynch: {lynch_status} ({lynch_delta:+.1f}%)")
                
                # DCF
                dcf_status = row.get('dcf_valuation_status', 'N/A')
                dcf_delta = row.get('dcf_delta_percentage', 0)
                print(f"DCF: {dcf_status} ({dcf_delta:+.1f}%)")
                
                # Munger Farm
                munger_status = row.get('munger_farm_assessment', 'N/A')
                munger_delta = row.get('munger_farm_delta', 0)
                print(f"Munger Farm: {munger_status} ({munger_delta:+.1f}%)")
                
                # Enhanced DCF
                enhanced_dcf_status = row.get('enhanced_dcf_status', 'N/A')
                enhanced_dcf_delta = row.get('enhanced_dcf_delta', 0)
                print(f"Enhanced DCF: {enhanced_dcf_status} ({enhanced_dcf_delta:+.1f}%)")
                
                # Relative Valuation
                relative_status = row.get('relative_valuation_status', 'N/A')
                relative_delta = row.get('relative_valuation_delta', 0)
                print(f"Relative Valuation: {relative_status} ({relative_delta:+.1f}%)")
            else:
                print(f"\n--- {stock} ---")
                print("No data found")
        
        print(f"\n=== FORMATTING CONSISTENCY CHECK ===")
        
        # Check if formatting is consistent across sheets
        if 'My Portfolio' in wb.sheetnames and 'Prospects' in wb.sheetnames:
            ws_portfolio = wb['My Portfolio']
            ws_prospects = wb['Prospects']
            
            # Check if both sheets have the same number of columns
            portfolio_cols = ws_portfolio.max_column
            prospects_cols = ws_prospects.max_column
            
            print(f"My Portfolio columns: {portfolio_cols}")
            print(f"Prospects columns: {prospects_cols}")
            
            if portfolio_cols == prospects_cols:
                print("✓ Column count is consistent")
            else:
                print("✗ Column count mismatch")
            
            # Check header consistency
            portfolio_headers = []
            prospects_headers = []
            
            for col in range(1, min(portfolio_cols, prospects_cols) + 1):
                portfolio_headers.append(ws_portfolio.cell(row=1, column=col).value)
                prospects_headers.append(ws_prospects.cell(row=1, column=col).value)
            
            if portfolio_headers == prospects_headers:
                print("✓ Headers are consistent")
            else:
                print("✗ Headers mismatch")
                print(f"Portfolio: {portfolio_headers}")
                print(f"Prospects: {prospects_headers}")
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")

if __name__ == "__main__":
    check_excel_formatting()
