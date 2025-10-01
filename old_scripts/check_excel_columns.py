import pandas as pd
from openpyxl import load_workbook

def check_excel_columns(excel_file):
    """Check what columns are actually available in the Excel file"""
    try:
        print("=== CHECKING EXCEL FILE COLUMNS ===")
        
        # Load the workbook
        wb = load_workbook(excel_file)
        print(f"Available sheets: {wb.sheetnames}")
        
        # Read the Valuation Data sheet
        df = pd.read_excel(excel_file, sheet_name='Valuation Data')
        print(f"\nTotal records: {len(df)}")
        print(f"Total columns: {len(df.columns)}")
        
        print("\nAll columns in Valuation Data sheet:")
        for i, col in enumerate(df.columns, 1):
            print(f"{i:2d}. {col}")
        
        # Check for new valuation method columns
        new_methods = [
            'enhanced_dcf_status', 'enhanced_dcf_delta', 'enhanced_dcf_intrinsic_value',
            'relative_valuation_status', 'relative_valuation_delta', 'ev_ebitda_multiple',
            'reverse_dcf_assessment', 'reverse_dcf_implied_growth', 'reverse_dcf_reasonable',
            'epv_assessment', 'epv_delta', 'epv_intrinsic_value',
            'rim_assessment', 'rim_delta', 'rim_intrinsic_value'
        ]
        
        print("\n=== NEW VALUATION METHOD COLUMNS ===")
        found_columns = []
        missing_columns = []
        
        for col in new_methods:
            if col in df.columns:
                found_columns.append(col)
                print(f"✓ {col}")
            else:
                missing_columns.append(col)
                print(f"✗ {col}")
        
        print(f"\nFound: {len(found_columns)}/{len(new_methods)} new method columns")
        
        if missing_columns:
            print(f"\nMissing columns: {missing_columns}")
        
        # Check latest data for a few stocks
        latest_data = df.sort_values(by='timestamp').groupby('ticker').last().reset_index()
        print(f"\nLatest data for {len(latest_data)} tickers")
        
        # Sample a few stocks to see what data they have
        print("\n=== SAMPLE STOCK DATA ===")
        sample_stocks = latest_data.head(3)
        for _, stock in sample_stocks.iterrows():
            print(f"\n{stock['ticker']} ({stock['company_name']}):")
            print(f"  Peter Lynch: {stock.get('lynch_valuation_status', 'N/A')}")
            print(f"  DCF: {stock.get('dcf_valuation_status', 'N/A')}")
            print(f"  Munger Farm: {stock.get('munger_7pct_assessment', 'N/A')}")
            print(f"  Enhanced DCF: {stock.get('enhanced_dcf_status', 'N/A')}")
            print(f"  Relative Valuation: {stock.get('relative_valuation_status', 'N/A')}")
            print(f"  Reverse DCF: {stock.get('reverse_dcf_assessment', 'N/A')}")
            print(f"  EPV: {stock.get('epv_assessment', 'N/A')}")
            print(f"  RIM: {stock.get('rim_assessment', 'N/A')}")
        
        return True
        
    except Exception as e:
        print(f"Error checking Excel columns: {e}")
        return False

if __name__ == "__main__":
    excel_file_path = r"C:\Users\james\Downloads\Stock Valuation\stock_valuation_dataset.xlsx"
    check_excel_columns(excel_file_path)
